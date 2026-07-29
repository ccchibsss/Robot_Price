#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
ПОЛНЫЙ МОНОЛИТНЫЙ РОБОТ ДЛЯ ЗАГРУЗКИ ПРАЙСОВ ИЗ ПОЧТЫ 
С БАЗОЙ ТОВАРОВ В GOOGLE SHEETS И ОТПРАВКОЙ В ЯНДЕКС МАРКЕТ
+ МОДУЛЬ АНАЛИТИКИ ПРАЙСОВ ОТ ВСЕХ ПОСТАВЩИКОВ
+ ИНТЕРАКТИВНЫЙ КОНСТРУКТОР МАППИНГА КОЛОНОК
+ РАСШИРЕННАЯ ВАЛИДАЦИЯ ДАННЫХ
+ УМНОЕ АВТООПРЕДЕЛЕНИЕ КОЛОНОК
+ ВЕДЕНИЕ ИСТОРИИ ИЗМЕНЕНИЙ МАППИНГА
+ СИСТЕМА УВЕДОМЛЕНИЙ
+ МНОГОУРОВНЕВОЕ КЭШИРОВАНИЕ
+ ОПТИМИЗИРОВАННАЯ ПАРАЛЛЕЛЬНАЯ ОБРАБОТКА
+ МОНИТОРИНГ И СТАТИСТИКА В РЕАЛЬНОМ ВРЕМЕНИ
ВЕРСИЯ 11.0 - ПОЛНЫЙ МОНОЛИТ С РАСШИРЕННЫМ ФУНКЦИОНАЛОМ
НИКАКИХ СОКРАЩЕНИЙ - 100% ПОЛНЫЙ КОД
"""

# ===================================================================
# БЛОК 1: ВСЕ ИМПОРТЫ (РАСШИРЕННЫЙ НАБОР)
# ===================================================================

import os
import sys
import re
import json
import time
import hashlib
import sqlite3
import threading
import queue
import smtplib
import imaplib
import email
import base64
import logging
import tempfile
import shutil
import csv
import traceback
import pickle
import gzip
import zipfile
import tarfile
import argparse
import signal
import platform
import uuid
import random
import string
import secrets
import hashlib
import functools
import itertools
import operator
import math
import statistics
from io import BytesIO, StringIO
from pathlib import Path
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Tuple, Any, Union, Callable, Set, Generator, Iterator, TypeVar, Generic
from dataclasses import dataclass, asdict, field, fields
from enum import Enum, auto
from email.policy import default as email_default_policy
from email.header import decode_header, Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.parser import Parser, BytesParser
from email.message import Message, EmailMessage
from email.utils import parsedate_to_datetime, formatdate, make_msgid
from collections import defaultdict, deque, OrderedDict, Counter, namedtuple
from functools import wraps, lru_cache, partial, reduce
from contextlib import contextmanager, suppress, redirect_stdout, redirect_stderr
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed, wait, FIRST_COMPLETED
import warnings
warnings.filterwarnings('ignore')

# Попытка импорта rarfile (опционально)
try:
    import rarfile
    RAR_SUPPORT = True
except ImportError:
    RAR_SUPPORT = False

# Попытка импорта дополнительных библиотек для улучшенной обработки
try:
    import magic
    MAGIC_SUPPORT = True
except ImportError:
    MAGIC_SUPPORT = False

try:
    import chardet
    CHARDET_SUPPORT = True
except ImportError:
    CHARDET_SUPPORT = False

# Сторонние библиотеки
try:
    import streamlit as st
    import pandas as pd
    import numpy as np
    import requests
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    from google.oauth2.service_account import Credentials as GoogleCredentials
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload, MediaIoBaseUpload
except ImportError as e:
    print(f"Ошибка импорта: {e}")
    print("Установите зависимости: pip install streamlit pandas numpy requests plotly beautifulsoup4 openpyxl gspread oauth2client google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client python-magic chardet")
    sys.exit(1)


# ===================================================================
# БЛОК 2: РАСШИРЕННЫЕ ПЕРЕЧИСЛЕНИЯ И СТРУКТУРЫ ДАННЫХ
# ===================================================================

class LogLevel(Enum):
    """Уровни логирования"""
    DEBUG = auto()
    INFO = auto()
    WARNING = auto()
    ERROR = auto()
    CRITICAL = auto()
    SUCCESS = auto()

class TaskStatus(Enum):
    """Статусы выполнения задач"""
    PENDING = "pending"
    RUNNING = "running"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"
    PARTIAL = "partial"

class FileFormat(Enum):
    """Поддерживаемые форматы файлов"""
    EXCEL_XLSX = "xlsx"
    EXCEL_XLS = "xls"
    CSV = "csv"
    XML = "xml"
    JSON = "json"
    TXT = "txt"
    ZIP = "zip"
    TAR = "tar"
    GZ = "gz"
    UNKNOWN = "unknown"

class PriceChangeDirection(Enum):
    """Направление изменения цены"""
    INCREASED = "increased"
    DECREASED = "decreased"
    UNCHANGED = "unchanged"
    NEW = "new"
    REMOVED = "removed"

@dataclass
class PriceChangeRecord:
    """Запись об изменении цены"""
    sku: str
    old_price: float
    new_price: float
    old_stock: int
    new_stock: int
    supplier: str
    direction: PriceChangeDirection
    change_percent: float
    change_absolute: float
    timestamp: datetime = field(default_factory=datetime.now)
    
    def to_dict(self) -> Dict:
        """Преобразование в словарь"""
        data = asdict(self)
        data['direction'] = self.direction.value
        data['timestamp'] = self.timestamp.isoformat()
        return data

@dataclass
class ValidationResult:
    """Результат валидации данных"""
    is_valid: bool = True
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    statistics: Dict[str, Any] = field(default_factory=dict)
    
    def add_error(self, error: str) -> None:
        """Добавление ошибки"""
        self.errors.append(error)
        self.is_valid = False
    
    def add_warning(self, warning: str) -> None:
        """Добавление предупреждения"""
        self.warnings.append(warning)
    
    def merge(self, other: 'ValidationResult') -> 'ValidationResult':
        """Объединение результатов валидации"""
        self.errors.extend(other.errors)
        self.warnings.extend(other.warnings)
        self.statistics.update(other.statistics)
        self.is_valid = self.is_valid and other.is_valid
        return self

@dataclass
class MappingHistory:
    """История изменений маппинга"""
    id: str = field(default_factory=lambda: str(uuid.uuid4()))
    supplier_name: str = ''
    mapping: Dict[str, str] = field(default_factory=dict)
    created_at: datetime = field(default_factory=datetime.now)
    created_by: str = 'system'
    comment: str = ''
    version: int = 1
    
    def to_dict(self) -> Dict:
        """Преобразование в словарь"""
        return {
            'id': self.id,
            'supplier_name': self.supplier_name,
            'mapping': self.mapping,
            'created_at': self.created_at.isoformat(),
            'created_by': self.created_by,
            'comment': self.comment,
            'version': self.version
        }

@dataclass
class NotificationConfig:
    """Настройки уведомлений"""
    enabled: bool = False
    email_notifications: bool = False
    email_recipients: List[str] = field(default_factory=list)
    telegram_enabled: bool = False
    telegram_bot_token: str = ''
    telegram_chat_id: str = ''
    notify_on_success: bool = True
    notify_on_failure: bool = True
    notify_on_price_changes: bool = True
    min_price_change_percent: float = 5.0
    notify_on_new_products: bool = True
    notify_on_out_of_stock: bool = True
    daily_summary: bool = True
    summary_time: str = "09:00"
    
    def to_dict(self) -> Dict:
        """Преобразование в словарь"""
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'NotificationConfig':
        """Создание из словаря"""
        return cls(**data)


# ===================================================================
# БЛОК 3: РАСШИРЕННАЯ КОНФИГУРАЦИЯ С ПОСТАВЩИКАМИ И МАППИНГОМ
# ===================================================================

@dataclass
class SupplierConfig:
    """Расширенная конфигурация одного поставщика с маппингом колонок"""
    name: str = ''
    email: str = ''
    email_password: str = ''
    imap_server: str = 'imap.mail.ru'
    imap_port: int = 993
    subject_filter: str = ''
    sender_filter: str = ''
    enabled: bool = True
    last_sync: Optional[str] = None
    priority: int = 0  # Приоритет поставщика (выше = важнее)
    auto_apply_markup: bool = False  # Автоматически применять наценку
    custom_markup: float = 0.0  # Индивидуальная наценка
    min_order_amount: float = 0.0  # Минимальная сумма заказа
    delivery_days: int = 0  # Срок доставки в днях
    payment_terms: str = ''  # Условия оплаты
    contact_person: str = ''  # Контактное лицо
    contact_phone: str = ''  # Телефон
    notes: str = ''  # Заметки
    rating: float = 0.0  # Рейтинг поставщика (0-5)
    last_quality_check: Optional[str] = None  # Дата последней проверки качества
    
    # Маппинг колонок (сохраняется после настройки)
    column_mapping: Dict[str, str] = field(default_factory=lambda: {
        'sku': '',
        'price': '',
        'stock': '',
        'brand': '',
        'name': '',
        'category': '',
        'description': '',
        'weight': '',
        'dimensions': '',
        'barcode': '',
        'country': '',
        'warranty': '',
        'min_order_qty': ''
    })
    
    # Дополнительные правила обработки
    processing_rules: Dict[str, Any] = field(default_factory=lambda: {
        'skip_rows': 0,  # Пропустить первые N строк
        'skip_empty_rows': True,  # Пропускать пустые строки
        'price_multiplier': 1.0,  # Множитель цены
        'stock_formula': '',  # Формула расчета остатка
        'price_rounding': 2,  # Округление цены
        'currency_conversion': 1.0,  # Конвертация валюты
        'min_price_threshold': 0,  # Минимальная допустимая цена
        'max_price_threshold': 999999999,  # Максимальная допустимая цена
        'min_stock_threshold': 0,  # Минимальный остаток
        'deduplicate_by': 'sku',  # Дедупликация по полю
        'validation_rules': {
            'check_sku_format': True,  # Проверка формата SKU
            'check_price_range': True,  # Проверка диапазона цен
            'check_stock_positive': True,  # Проверка положительного остатка
            'check_brand_required': False  # Бренд обязателен
        }
    })
    
    # Пример данных для отображения
    sample_data: List[Dict] = field(default_factory=list)
    
    # Статистика поставщика
    stats: Dict[str, Any] = field(default_factory=lambda: {
        'total_files_processed': 0,
        'total_products_loaded': 0,
        'last_successful_sync': None,
        'last_error': None,
        'average_response_time': 0,
        'success_rate': 0,
        'total_errors': 0
    })
    
    def to_dict(self) -> Dict:
        """Расширенное преобразование в словарь"""
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'SupplierConfig':
        """Создание из словаря с обратной совместимостью"""
        # Удаляем неизвестные поля для обратной совместимости
        known_fields = {f.name for f in fields(cls)}
        filtered_data = {k: v for k, v in data.items() if k in known_fields}
        
        # Добавляем значения по умолчанию для новых полей
        instance = cls()
        for field_info in fields(cls):
            if field_info.name not in filtered_data:
                filtered_data[field_info.name] = field_info.default_factory() if callable(field_info.default_factory) else field_info.default
        
        return cls(**filtered_data)
    
    def calculate_success_rate(self) -> float:
        """Расчет процента успешных синхронизаций"""
        total = self.stats['total_files_processed'] + self.stats['total_errors']
        if total == 0:
            return 0.0
        return (self.stats['total_files_processed'] / total) * 100
    
    def update_stats(self, success: bool, products_count: int = 0) -> None:
        """Обновление статистики поставщика"""
        self.stats['total_files_processed'] += 1
        if success:
            self.stats['total_products_loaded'] += products_count
            self.stats['last_successful_sync'] = datetime.now().isoformat()
            self.stats['last_error'] = None
        else:
            self.stats['total_errors'] += 1
            self.stats['last_error'] = datetime.now().isoformat()
        
        self.stats['success_rate'] = self.calculate_success_rate()
    
    def is_valid_email(self) -> bool:
        """Проверка валидности email"""
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(email_pattern, self.email))
    
    def get_effective_markup(self, global_markup: float) -> float:
        """Получение эффективной наценки"""
        if self.auto_apply_markup and self.custom_markup > 0:
            return self.custom_markup
        return global_markup


@dataclass
class Config:
    """Расширенный класс конфигурации приложения с новыми возможностями"""
    # Версия конфигурации для миграций
    config_version: int = 2
    
    # Почтовые настройки (главный ящик)
    imap_server: str = 'imap.mail.ru'
    imap_port: int = 993
    email_user: str = ''
    email_pass: str = ''
    email_search_days: int = 1
    email_max_emails: int = 50
    email_mark_as_read: bool = False  # Помечать письма как прочитанные
    email_delete_after_processing: bool = False  # Удалять письма после обработки
    email_ssl_verify: bool = True  # Проверка SSL сертификата
    
    # Настройки Яндекс Маркета
    yandex_token: str = ''
    campaign_id: int = 0
    warehouse_id: int = 0
    yandex_api_url: str = 'https://api.partner.market.yandex.ru'
    yandex_auto_update: bool = True  # Автоматически обновлять на маркете
    yandex_update_interval_minutes: int = 60  # Интервал обновления
    
    # Настройки Google Sheets
    google_sheet_id: str = ''
    google_sheet_name: str = 'Товары'
    google_credentials_json: str = 'google_credentials.json'
    google_scope: List[str] = field(default_factory=lambda: [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
        'https://www.googleapis.com/auth/spreadsheets'
    ])
    google_auto_backup: bool = True  # Автоматическое резервное копирование
    google_backup_interval_hours: int = 24  # Интервал резервного копирования
    
    # Настройки обработки
    markup_percent: float = 20.0
    min_stock_threshold: int = 1
    max_offers_per_request: int = 500
    request_timeout: int = 30
    retry_count: int = 3
    retry_delay: int = 5
    retry_backoff_multiplier: float = 2.0  # Множитель задержки при повторных попытках
    
    # Настройки параллельной обработки
    max_workers: int = 4  # Максимальное количество потоков
    parallel_downloads: bool = True  # Параллельная загрузка
    parallel_parsing: bool = True  # Параллельный парсинг
    chunk_size: int = 1000  # Размер чанка для обработки
    
    # Поставщики (хранятся с маппингом)
    suppliers: List[Dict] = field(default_factory=list)
    suppliers_file: str = 'suppliers.json'
    
    # Настройки уведомлений
    notification_config: Dict = field(default_factory=lambda: NotificationConfig().to_dict())
    
    # Пути к файлам
    db_path: str = 'robot_data.db'
    log_path: str = 'robot.log'
    uploads_dir: str = 'uploads'
    archive_dir: str = 'archive'
    config_json: str = 'config.json'
    cache_dir: str = 'cache'
    temp_dir: str = 'temp'
    backup_dir: str = 'backups'
    analysis_dir: str = 'analysis_results'
    mapping_history_dir: str = 'mapping_history'
    reports_dir: str = 'reports'
    exports_dir: str = 'exports'
    
    # Глобальный маппинг (для автоопределения) - расширенный
    column_mapping: Dict[str, List[str]] = field(default_factory=lambda: {
        'sku': ['артикул', 'sku', 'код', 'id', 'номер', 'article', 'artikul', 'vendor_code', 'product_id', 'item_id', 'код_товара', 'ид', 'арт', 'art', 'code', 'product_code'],
        'price': ['цена', 'price', 'cost', 'розница', 'розничная', 'розничная цена', 'цена_розница', 'цена_розн', 'retail_price', 'selling_price', 'sales_price', 'стоимость', 'прайс', 'ценарозн'],
        'stock': ['остаток', 'stock', 'количество', 'quantity', 'кол-во', 'наличие', 'колво', 'qty', 'available', 'in_stock', 'остатки', 'склад', 'запас'],
        'brand': ['бренд', 'brand', 'производитель', 'manufacturer', 'vendor', 'make', 'марка', 'manufact', 'произв', 'брэнд'],
        'name': ['название', 'name', 'товар', 'product', 'наименование', 'наим', 'product_name', 'title', 'description', 'описание', 'назв', 'продукт'],
        'category': ['категория', 'category', 'раздел', 'section', 'cat', 'группа', 'group', 'тип', 'type', 'вид', 'kind'],
        'description': ['описание', 'description', 'desc', 'детали', 'details', 'характеристики', 'спецификация'],
        'weight': ['вес', 'weight', 'масса', 'mass', 'кг', 'kg'],
        'dimensions': ['размеры', 'dimensions', 'габариты', 'size', 'длина', 'ширина', 'высота'],
        'barcode': ['штрихкод', 'barcode', 'ean', 'upc', 'баркод', 'штрих-код', 'ean13'],
        'country': ['страна', 'country', 'производство', 'origin', 'made_in'],
        'warranty': ['гарантия', 'warranty', 'срок гарантии', 'guarantee']
    })
    
    # Настройки валидации
    validation_config: Dict = field(default_factory=lambda: {
        'max_sku_length': 100,
        'min_price': 0.01,
        'max_price': 999999999,
        'max_stock': 999999,
        'allowed_currencies': ['RUB', 'USD', 'EUR'],
        'required_fields': ['sku', 'price'],
        'optional_fields': ['stock', 'brand', 'name', 'category'],
        'auto_fix_price_rounding': True,
        'auto_fix_stock_negative': True,
        'auto_fix_sku_uppercase': True,
        'auto_remove_duplicates': True
    })
    
    def __post_init__(self):
        """Пост-инициализация для создания директорий"""
        self._create_directories()
    
    def _create_directories(self) -> None:
        """Создание всех необходимых директорий"""
        directories = [
            self.uploads_dir, self.archive_dir, self.cache_dir,
            self.temp_dir, self.backup_dir, self.analysis_dir,
            self.mapping_history_dir, self.reports_dir, self.exports_dir
        ]
        for directory in directories:
            Path(directory).mkdir(parents=True, exist_ok=True)
    
    def save(self) -> None:
        """Сохраняет конфигурацию в JSON файл с бэкапом"""
        try:
            if os.path.exists(self.config_json):
                backup_filename = f"{self.config_json}.{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
                backup_path = os.path.join(self.backup_dir, backup_filename)
                shutil.copy2(self.config_json, backup_path)
                
                # Удаляем старые бэкапы (оставляем последние 10)
                backup_files = sorted(Path(self.backup_dir).glob(f"{self.config_json}.*.bak"))
                if len(backup_files) > 10:
                    for old_backup in backup_files[:-10]:
                        old_backup.unlink()
            
            self._save_suppliers()
            
            config_dict = asdict(self)
            config_dict.pop('suppliers', None)
            config_dict.pop('suppliers_file', None)
            
            with open(self.config_json, 'w', encoding='utf-8') as f:
                json.dump(config_dict, f, ensure_ascii=False, indent=2, default=str)
                
        except Exception as e:
            print(f"Ошибка сохранения конфигурации: {e}")
            raise
    
    def _save_suppliers(self) -> None:
        """Сохраняет список поставщиков с маппингом в отдельный файл"""
        try:
            # Создаем бэкап перед сохранением
            if os.path.exists(self.suppliers_file):
                backup_file = f"{self.suppliers_file}.{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
                shutil.copy2(self.suppliers_file, backup_file)
            
            with open(self.suppliers_file, 'w', encoding='utf-8') as f:
                json.dump(self.suppliers, f, ensure_ascii=False, indent=2, default=str)
                
        except Exception as e:
            print(f"Ошибка сохранения поставщиков: {e}")
    
    def _load_suppliers(self) -> None:
        """Загружает список поставщиков с маппингом из файла"""
        try:
            if os.path.exists(self.suppliers_file):
                with open(self.suppliers_file, 'r', encoding='utf-8') as f:
                    self.suppliers = json.load(f)
                    
                # Обеспечиваем обратную совместимость с новыми полями
                updated_suppliers = []
                for supplier_data in self.suppliers:
                    supplier = SupplierConfig.from_dict(supplier_data)
                    updated_suppliers.append(supplier.to_dict())
                self.suppliers = updated_suppliers
                
        except Exception as e:
            print(f"Ошибка загрузки поставщиков: {e}")
            self.suppliers = []
    
    def get_supplier_mapping(self, supplier_name: str) -> Optional[Dict[str, str]]:
        """Получает маппинг колонок для поставщика"""
        for supplier in self.suppliers:
            if supplier.get('name') == supplier_name:
                return supplier.get('column_mapping', {})
        return None
    
    def set_supplier_mapping(self, supplier_name: str, mapping: Dict[str, str], 
                           created_by: str = 'system', comment: str = '') -> bool:
        """Устанавливает маппинг колонок для поставщика с сохранением истории"""
        for i, supplier in enumerate(self.suppliers):
            if supplier.get('name') == supplier_name:
                # Сохраняем старый маппинг в историю
                old_mapping = supplier.get('column_mapping', {})
                if old_mapping and old_mapping != mapping:
                    self._save_mapping_history(supplier_name, old_mapping, mapping, created_by, comment)
                
                # Обновляем маппинг
                supplier['column_mapping'] = mapping
                self._save_suppliers()
                return True
        return False
    
    def _save_mapping_history(self, supplier_name: str, old_mapping: Dict[str, str],
                            new_mapping: Dict[str, str], created_by: str, comment: str) -> None:
        """Сохраняет историю изменений маппинга"""
        try:
            history_dir = Path(self.mapping_history_dir)
            history_dir.mkdir(parents=True, exist_ok=True)
            
            history_file = history_dir / f"{supplier_name}_mapping_history.json"
            
            # Загружаем существующую историю
            history = []
            if history_file.exists():
                with open(history_file, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            
            # Определяем следующую версию
            next_version = len(history) + 1
            
            # Создаем запись в истории
            history_record = MappingHistory(
                supplier_name=supplier_name,
                mapping=new_mapping,
                created_by=created_by,
                comment=comment,
                version=next_version
            )
            
            # Добавляем информацию об изменениях
            changes = []
            for key in set(list(old_mapping.keys()) + list(new_mapping.keys())):
                old_value = old_mapping.get(key, '')
                new_value = new_mapping.get(key, '')
                if old_value != new_value:
                    changes.append({
                        'field': key,
                        'old_value': old_value,
                        'new_value': new_value
                    })
            
            history_record_dict = history_record.to_dict()
            history_record_dict['changes'] = changes
            history.append(history_record_dict)
            
            # Сохраняем историю
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2, default=str)
                
        except Exception as e:
            print(f"Ошибка сохранения истории маппинга: {e}")
    
    def get_mapping_history(self, supplier_name: str) -> List[Dict]:
        """Получает историю изменений маппинга для поставщика"""
        try:
            history_file = Path(self.mapping_history_dir) / f"{supplier_name}_mapping_history.json"
            if history_file.exists():
                with open(history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Ошибка загрузки истории маппинга: {e}")
        return []
    
    def get_supplier_by_name(self, name: str) -> Optional[SupplierConfig]:
        """Получает объект поставщика по имени"""
        for supplier_data in self.suppliers:
            if supplier_data.get('name') == name:
                return SupplierConfig.from_dict(supplier_data)
        return None
    
    def get_active_suppliers(self) -> List[SupplierConfig]:
        """Получает список активных поставщиков"""
        active_suppliers = []
        for supplier_data in self.suppliers:
            if supplier_data.get('enabled', True):
                active_suppliers.append(SupplierConfig.from_dict(supplier_data))
        return sorted(active_suppliers, key=lambda x: x.priority, reverse=True)
    
    @classmethod
    def load(cls) -> 'Config':
        """Загружает конфигурацию из JSON файла с миграцией"""
        config = cls()
        try:
            if os.path.exists(config.config_json):
                with open(config.config_json, 'r', encoding='utf-8') as f:
                    config_dict = json.load(f)
                
                # Миграция конфигурации при необходимости
                config_version = config_dict.get('config_version', 1)
                if config_version < 2:
                    config_dict = cls._migrate_config_v1_to_v2(config_dict)
                
                # Применяем настройки
                for key, value in config_dict.items():
                    if hasattr(config, key):
                        setattr(config, key, value)
                        
        except json.JSONDecodeError as e:
            print(f"Ошибка парсинга JSON: {e}")
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
        
        config._load_suppliers()
        return config
    
    @classmethod
    def _migrate_config_v1_to_v2(cls, config_dict: Dict) -> Dict:
        """Миграция конфигурации с версии 1 на версию 2"""
        # Добавляем новые поля с значениями по умолчанию
        config_dict['config_version'] = 2
        config_dict.setdefault('notification_config', NotificationConfig().to_dict())
        config_dict.setdefault('validation_config', Config().validation_config)
        config_dict.setdefault('max_workers', 4)
        config_dict.setdefault('parallel_downloads', True)
        config_dict.setdefault('parallel_parsing', True)
        
        # Обновляем column_mapping для новых полей
        if 'column_mapping' in config_dict:
            default_mapping = Config().column_mapping
            for key, values in default_mapping.items():
                if key not in config_dict['column_mapping']:
                    config_dict['column_mapping'][key] = values
        
        return config_dict


# ===================================================================
# БЛОК 4: РАСШИРЕННАЯ СИСТЕМА ЛОГГИРОВАНИЯ
# ===================================================================

class StreamlitLogger:
    """Расширенный класс для логирования с поддержкой уровней и форматирования"""
    
    LOG_COLORS = {
        LogLevel.DEBUG: '#808080',      # Серый
        LogLevel.INFO: '#0000FF',       # Синий
        LogLevel.WARNING: '#FFA500',    # Оранжевый
        LogLevel.ERROR: '#FF0000',      # Красный
        LogLevel.CRITICAL: '#8B0000',   # Темно-красный
        LogLevel.SUCCESS: '#00FF00'     # Зеленый
    }
    
    def __init__(self, max_logs: int = 2000, max_file_size_mb: int = 10, 
                 backup_count: int = 5, enable_colors: bool = True):
        self.logs: List[Dict[str, Any]] = []
        self.max_logs = max_logs
        self.max_file_size_mb = max_file_size_mb
        self.backup_count = backup_count
        self.enable_colors = enable_colors
        self.log_path = 'robot.log'
        self._init_log_file()
        self._log_queue = queue.Queue()
        self._flush_thread = threading.Thread(target=self._flush_logs, daemon=True)
        self._flush_thread.start()
        self.lock = threading.Lock()
    
    def _init_log_file(self) -> None:
        """Инициализация лог-файла с ротацией"""
        try:
            if os.path.exists(self.log_path):
                file_size = os.path.getsize(self.log_path) / (1024 * 1024)
                if file_size > self.max_file_size_mb:
                    self._rotate_logs()
            
            with open(self.log_path, 'a', encoding='utf-8') as f:
                f.write(f"\n{'='*60}\n")
                f.write(f"ЛОГИРОВАНИЕ ЗАПУЩЕНО: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Версия: 11.0 | Платформа: {platform.platform()}\n")
                f.write(f"{'='*60}\n")
        except Exception as e:
            print(f"Ошибка инициализации лог-файла: {e}")
    
    def _rotate_logs(self) -> None:
        """Ротация лог-файлов"""
        try:
            for i in range(self.backup_count - 1, 0, -1):
                src = f"{self.log_path}.{i}.bak"
                dst = f"{self.log_path}.{i+1}.bak"
                if os.path.exists(src):
                    shutil.move(src, dst)
            
            if os.path.exists(self.log_path):
                shutil.move(self.log_path, f"{self.log_path}.1.bak")
        except Exception as e:
            print(f"Ошибка ротации логов: {e}")
    
    def _flush_logs(self) -> None:
        """Фоновая запись логов в файл"""
        while True:
            try:
                while not self._log_queue.empty():
                    log_entry = self._log_queue.get_nowait()
                    with open(self.log_path, 'a', encoding='utf-8') as f:
                        f.write(log_entry + '\n')
                time.sleep(1)
            except Exception:
                time.sleep(5)
    
    def log(self, message: str, level: LogLevel = LogLevel.INFO, 
            context: Optional[Dict] = None) -> None:
        """Расширенное логирование с контекстом"""
        timestamp = datetime.now()
        log_entry = {
            'timestamp': timestamp.isoformat(),
            'level': level.name,
            'message': message,
            'context': context or {}
        }
        
        with self.lock:
            self.logs.append(log_entry)
            if len(self.logs) > self.max_logs:
                self.logs = self.logs[-self.max_logs:]
        
        # Форматируем для файла
        context_str = f" | Context: {json.dumps(context, ensure_ascii=False)}" if context else ""
        file_entry = f"[{timestamp.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}] [{level.name:8s}] {message}{context_str}"
        
        self._log_queue.put(file_entry)
    
    def debug(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование отладочной информации"""
        self.log(message, LogLevel.DEBUG, context)
    
    def info(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование информационного сообщения"""
        self.log(message, LogLevel.INFO, context)
    
    def warning(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование предупреждения"""
        self.log(message, LogLevel.WARNING, context)
    
    def error(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование ошибки"""
        self.log(message, LogLevel.ERROR, context)
    
    def success(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование успешного выполнения"""
        self.log(message, LogLevel.SUCCESS, context)
    
    def critical(self, message: str, context: Optional[Dict] = None) -> None:
        """Логирование критической ошибки"""
        self.log(message, LogLevel.CRITICAL, context)
    
    def exception(self, message: str, exc: Exception = None, 
                 context: Optional[Dict] = None) -> None:
        """Логирование исключения с трассировкой"""
        error_context = context or {}
        if exc:
            error_context['exception_type'] = type(exc).__name__
            error_context['exception_args'] = str(exc.args)
            error_context['traceback'] = traceback.format_exc()
        self.log(message, LogLevel.ERROR, error_context)
    
    def get_logs(self, last_n: Optional[int] = None, 
                level: Optional[LogLevel] = None) -> List[Dict]:
        """Получение логов с фильтрацией"""
        logs = self.logs.copy()
        
        if level:
            logs = [log for log in logs if log['level'] == level.name]
        
        if last_n:
            logs = logs[-last_n:]
        
        return logs
    
    def get_logs_by_timerange(self, start_time: datetime, 
                             end_time: Optional[datetime] = None) -> List[Dict]:
        """Получение логов за временной диапазон"""
        if end_time is None:
            end_time = datetime.now()
        
        filtered_logs = []
        for log in self.logs:
            log_time = datetime.fromisoformat(log['timestamp'])
            if start_time <= log_time <= end_time:
                filtered_logs.append(log)
        
        return filtered_logs
    
    def clear(self) -> None:
        """Очистка логов"""
        with self.lock:
            self.logs = []
    
    def get_stats(self) -> Dict[str, int]:
        """Получение статистики по логам"""
        stats = {
            'total': len(self.logs),
            LogLevel.DEBUG.name.lower(): 0,
            LogLevel.INFO.name.lower(): 0,
            LogLevel.WARNING.name.lower(): 0,
            LogLevel.ERROR.name.lower(): 0,
            LogLevel.CRITICAL.name.lower(): 0,
            LogLevel.SUCCESS.name.lower(): 0
        }
        
        for log in self.logs:
            level_name = log['level'].lower()
            if level_name in stats:
                stats[level_name] += 1
        
        return stats
    
    def export_logs(self, format: str = 'json') -> str:
        """Экспорт логов в различные форматы"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'json':
            filename = f"logs_export_{timestamp}.json"
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.logs, f, ensure_ascii=False, indent=2)
        elif format == 'csv':
            filename = f"logs_export_{timestamp}.csv"
            with open(filename, 'w', encoding='utf-8', newline='') as f:
                if self.logs:
                    writer = csv.DictWriter(f, fieldnames=self.logs[0].keys())
                    writer.writeheader()
                    writer.writerows(self.logs)
        else:
            filename = f"logs_export_{timestamp}.txt"
            with open(filename, 'w', encoding='utf-8') as f:
                for log in self.logs:
                    f.write(f"[{log['timestamp']}] [{log['level']}] {log['message']}\n")
        
        return filename
    
    def search_logs(self, query: str, case_sensitive: bool = False) -> List[Dict]:
        """Поиск по логам"""
        results = []
        search_query = query if case_sensitive else query.lower()
        
        for log in self.logs:
            message = log['message'] if case_sensitive else log['message'].lower()
            if search_query in message:
                results.append(log)
        
        return results


# ===================================================================
# БЛОК 5: РАСШИРЕННАЯ БАЗА ДАННЫХ SQLITE
# ===================================================================

class Database:
    """Расширенный класс для работы с SQLite базой данных с новыми таблицами"""
    
    SCHEMA_VERSION = 3
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._init_db()
        self._check_schema_version()
    
    def _check_schema_version(self) -> None:
        """Проверка и обновление версии схемы БД"""
        try:
            with self.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("PRAGMA user_version")
                current_version = cursor.fetchone()[0]
                
                if current_version < self.SCHEMA_VERSION:
                    self._migrate_schema(current_version)
                    cursor.execute(f"PRAGMA user_version = {self.SCHEMA_VERSION}")
                    conn.commit()
        except Exception as e:
            print(f"Ошибка проверки версии схемы: {e}")
    
    def _migrate_schema(self, from_version: int) -> None:
        """Миграция схемы БД"""
        if from_version < 2:
            self._migrate_v1_to_v2()
        if from_version < 3:
            self._migrate_v2_to_v3()
    
    def _migrate_v1_to_v2(self) -> None:
        """Миграция с версии 1 на версию 2"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS supplier_stats (
                    supplier_name TEXT PRIMARY KEY,
                    total_files INTEGER DEFAULT 0,
                    total_products INTEGER DEFAULT 0,
                    total_errors INTEGER DEFAULT 0,
                    last_sync TIMESTAMP,
                    avg_price REAL DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
    
    def _migrate_v2_to_v3(self) -> None:
        """Миграция с версии 2 на версию 3"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Таблица для хранения метаданных файлов
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS file_metadata (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_hash TEXT UNIQUE NOT NULL,
                    original_filename TEXT NOT NULL,
                    file_size INTEGER,
                    mime_type TEXT,
                    encoding TEXT,
                    row_count INTEGER,
                    column_count INTEGER,
                    supplier TEXT,
                    processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Таблица для истории уведомлений
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS notification_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    notification_type TEXT NOT NULL,
                    recipient TEXT,
                    subject TEXT,
                    message TEXT,
                    status TEXT,
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    error_message TEXT
                )
            ''')
            
            conn.commit()
    
    def _init_db(self) -> None:
        """Инициализация всех таблиц базы данных"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            
            # Включаем поддержку внешних ключей
            cursor.execute("PRAGMA foreign_keys = ON")
            
            # Таблица обработанных файлов (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS processed_files (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_hash TEXT UNIQUE NOT NULL,
                    filename TEXT NOT NULL,
                    original_filename TEXT,
                    supplier TEXT,
                    supplier_email TEXT,
                    processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    offers_count INTEGER DEFAULT 0,
                    valid_offers INTEGER DEFAULT 0,
                    invalid_offers INTEGER DEFAULT 0,
                    new_products INTEGER DEFAULT 0,
                    updated_products INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'processed',
                    file_size INTEGER DEFAULT 0,
                    file_format TEXT,
                    encoding_used TEXT,
                    processing_time REAL DEFAULT 0,
                    error_message TEXT,
                    retry_count INTEGER DEFAULT 0,
                    metadata JSON
                )
            ''')
            
            # Индексы для ускорения запросов
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_hash ON processed_files(file_hash)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_supplier ON processed_files(supplier)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_processed_at ON processed_files(processed_at)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_status ON processed_files(status)')
            
            # Таблица истории аналитики (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS analysis_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    analysis_id TEXT UNIQUE,
                    supplier_count INTEGER,
                    total_files INTEGER,
                    total_products INTEGER,
                    unique_skus INTEGER,
                    new_skus INTEGER,
                    price_changes INTEGER,
                    min_price REAL,
                    max_price REAL,
                    avg_price REAL,
                    median_price REAL,
                    export_file TEXT,
                    duration_seconds REAL,
                    status TEXT,
                    error_message TEXT,
                    analysis_config JSON
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_analysis_date ON analysis_history(run_date)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_analysis_id ON analysis_history(analysis_id)')
            
            # Таблица ценовых сравнений (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS price_comparisons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    analysis_id TEXT,
                    sku TEXT NOT NULL,
                    brand TEXT,
                    name TEXT,
                    category TEXT,
                    supplier TEXT NOT NULL,
                    supplier_email TEXT,
                    price REAL,
                    original_price REAL,
                    currency TEXT DEFAULT 'RUB',
                    stock INTEGER,
                    source_file TEXT,
                    analyzed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_best_price BOOLEAN DEFAULT 0,
                    price_rank INTEGER,
                    FOREIGN KEY (analysis_id) REFERENCES analysis_history(analysis_id)
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_comparisons_sku ON price_comparisons(sku)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_comparisons_analysis ON price_comparisons(analysis_id)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_comparisons_supplier ON price_comparisons(supplier)')
            
            # Таблица истории цен (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS price_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sku TEXT NOT NULL,
                    supplier TEXT,
                    old_price REAL,
                    new_price REAL,
                    old_stock INTEGER,
                    new_stock INTEGER,
                    change_percent REAL,
                    change_direction TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    source TEXT DEFAULT 'auto',
                    analysis_id TEXT,
                    metadata JSON
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_history_sku ON price_history(sku)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_history_date ON price_history(updated_at)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_history_supplier ON price_history(supplier)')
            
            # Таблица статистики запусков (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS run_stats (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_id TEXT UNIQUE,
                    run_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    files_found INTEGER DEFAULT 0,
                    files_processed INTEGER DEFAULT 0,
                    files_skipped INTEGER DEFAULT 0,
                    files_with_errors INTEGER DEFAULT 0,
                    offers_sent INTEGER DEFAULT 0,
                    offers_failed INTEGER DEFAULT 0,
                    products_updated INTEGER DEFAULT 0,
                    products_added INTEGER DEFAULT 0,
                    products_deleted INTEGER DEFAULT 0,
                    errors INTEGER DEFAULT 0,
                    warnings INTEGER DEFAULT 0,
                    duration_seconds REAL DEFAULT 0,
                    status TEXT DEFAULT 'success',
                    run_type TEXT DEFAULT 'manual',
                    triggered_by TEXT,
                    config_snapshot JSON
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_stats_date ON run_stats(run_date)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_stats_status ON run_stats(status)')
            
            # Таблица кэша отправленных офферов (расширенная)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS sent_offers_cache (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sku TEXT UNIQUE NOT NULL,
                    last_price REAL,
                    last_stock INTEGER,
                    last_sent TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    sent_count INTEGER DEFAULT 1,
                    last_response_code INTEGER,
                    last_response_message TEXT,
                    price_history JSON
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_cache_sku ON sent_offers_cache(sku)')
            
            # Таблица для хранения конфигураций маппинга
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS mapping_configs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    supplier_name TEXT NOT NULL,
                    mapping JSON NOT NULL,
                    version INTEGER DEFAULT 1,
                    is_active BOOLEAN DEFAULT 1,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    created_by TEXT DEFAULT 'system',
                    comment TEXT
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_mapping_supplier ON mapping_configs(supplier_name)')
            
            # Таблица для алертов и уведомлений
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS alerts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    alert_type TEXT NOT NULL,
                    severity TEXT DEFAULT 'info',
                    title TEXT NOT NULL,
                    message TEXT,
                    data JSON,
                    is_read BOOLEAN DEFAULT 0,
                    is_resolved BOOLEAN DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    resolved_at TIMESTAMP,
                    assigned_to TEXT
                )
            ''')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_type ON alerts(alert_type)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_severity ON alerts(severity)')
            cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_unread ON alerts(is_read, is_resolved)')
            
            conn.commit()
    
    @contextmanager
    def get_connection(self):
        """Контекстный менеджер для соединения с БД"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()
    
    def is_file_processed(self, file_content: bytes) -> bool:
        """Проверка, был ли файл уже обработан (по хешу)"""
        file_hash = hashlib.md5(file_content).hexdigest()
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id FROM processed_files WHERE file_hash = ? AND status != 'error'",
                (file_hash,)
            )
            return cursor.fetchone() is not None
    
    def mark_file_processed(self, filename: str, file_content: bytes, 
                           offers_count: int, supplier: str = '',
                           status: str = 'processed', error_message: str = '',
                           valid_offers: int = 0, invalid_offers: int = 0,
                           new_products: int = 0, updated_products: int = 0,
                           processing_time: float = 0, metadata: Optional[Dict] = None) -> None:
        """Расширенная маркировка файла как обработанного"""
        file_hash = hashlib.md5(file_content).hexdigest()
        file_size = len(file_content)
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO processed_files 
                (file_hash, filename, supplier, offers_count, valid_offers, 
                 invalid_offers, new_products, updated_products, status, 
                 file_size, error_message, processing_time, metadata, processed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (file_hash, filename, supplier, offers_count, valid_offers,
                  invalid_offers, new_products, updated_products, status,
                  file_size, error_message, processing_time, 
                  json.dumps(metadata) if metadata else None))
            conn.commit()
    
    def save_analysis(self, supplier_count: int, total_products: int, 
                     unique_skus: int, export_file: str, status: str = 'success',
                     analysis_id: Optional[str] = None, duration: float = 0,
                     new_skus: int = 0, price_changes: int = 0,
                     min_price: float = 0, max_price: float = 0,
                     avg_price: float = 0, median_price: float = 0,
                     total_files: int = 0, analysis_config: Optional[Dict] = None) -> None:
        """Расширенное сохранение результатов анализа"""
        if analysis_id is None:
            analysis_id = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO analysis_history 
                (analysis_id, supplier_count, total_files, total_products, unique_skus, 
                 new_skus, price_changes, min_price, max_price, avg_price, median_price,
                 export_file, duration_seconds, status, analysis_config)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (analysis_id, supplier_count, total_files, total_products, unique_skus,
                  new_skus, price_changes, min_price, max_price, avg_price, median_price,
                  export_file, duration, status, 
                  json.dumps(analysis_config) if analysis_config else None))
            conn.commit()
    
    def get_analysis_history(self, days: int = 30) -> pd.DataFrame:
        """Получение истории анализов"""
        with self.get_connection() as conn:
            return pd.read_sql_query('''
                SELECT * FROM analysis_history 
                WHERE run_date > datetime('now', '-' || ? || ' days')
                ORDER BY run_date DESC
            ''', conn, params=(days,))
    
    def save_price_comparison(self, sku: str, supplier: str, price: float, 
                             stock: int, analysis_id: str, brand: str = '',
                             name: str = '', category: str = '',
                             is_best_price: bool = False, price_rank: int = 0,
                             original_price: float = 0, currency: str = 'RUB',
                             source_file: str = '', supplier_email: str = '') -> None:
        """Расширенное сохранение сравнения цен"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO price_comparisons 
                (analysis_id, sku, brand, name, category, supplier, supplier_email,
                 price, original_price, currency, stock, source_file, is_best_price, price_rank)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (analysis_id, sku, brand, name, category, supplier, supplier_email,
                  price, original_price, currency, stock, source_file, is_best_price, price_rank))
            conn.commit()
    
    def add_price_history(self, sku: str, old_price: float, new_price: float,
                         old_stock: Optional[int] = None, new_stock: Optional[int] = None,
                         source: str = 'auto', supplier: str = '',
                         analysis_id: Optional[str] = None,
                         metadata: Optional[Dict] = None) -> None:
        """Расширенное добавление записи в историю цен"""
        change_percent = 0
        if old_price > 0:
            change_percent = ((new_price - old_price) / old_price) * 100
        
        if new_price > old_price:
            change_direction = PriceChangeDirection.INCREASED.value
        elif new_price < old_price:
            change_direction = PriceChangeDirection.DECREASED.value
        else:
            change_direction = PriceChangeDirection.UNCHANGED.value
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO price_history 
                (sku, supplier, old_price, new_price, old_stock, new_stock,
                 change_percent, change_direction, source, analysis_id, metadata)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (sku, supplier, old_price, new_price, old_stock, new_stock,
                  change_percent, change_direction, source, analysis_id,
                  json.dumps(metadata) if metadata else None))
            conn.commit()
    
    def update_sent_cache(self, sku: str, price: float, stock: int,
                         response_code: int = 200, response_message: str = 'OK') -> None:
        """Расширенное обновление кэша отправленных офферов"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Получаем текущие данные для истории
            current = self.get_sent_cache(sku)
            price_history = []
            if current:
                price_history = [{'price': current[0], 'stock': current[1], 
                                'timestamp': datetime.now().isoformat()}]
            
            cursor.execute('''
                INSERT OR REPLACE INTO sent_offers_cache 
                (sku, last_price, last_stock, last_sent, sent_count,
                 last_response_code, last_response_message, price_history)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP, 
                        COALESCE((SELECT sent_count + 1 FROM sent_offers_cache WHERE sku = ?), 1),
                        ?, ?, ?)
            ''', (sku, price, stock, sku, response_code, response_message,
                  json.dumps(price_history)))
            conn.commit()
    
    def get_sent_cache(self, sku: str) -> Optional[Tuple[float, int]]:
        """Получение данных из кэша отправленных офферов"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT last_price, last_stock FROM sent_offers_cache WHERE sku = ?",
                (sku,)
            )
            row = cursor.fetchone()
            return (row['last_price'], row['last_stock']) if row else None
    
    def add_run_stats(self, files_processed: int, offers_sent: int,
                     products_updated: int, products_added: int,
                     errors: int, duration: float, status: str = 'success',
                     run_id: Optional[str] = None, run_type: str = 'manual',
                     triggered_by: str = '', files_found: int = 0,
                     files_skipped: int = 0, files_with_errors: int = 0,
                     offers_failed: int = 0, products_deleted: int = 0,
                     warnings: int = 0, config_snapshot: Optional[Dict] = None) -> str:
        """Расширенное сохранение статистики запуска"""
        if run_id is None:
            run_id = f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO run_stats 
                (run_id, files_found, files_processed, files_skipped, files_with_errors,
                 offers_sent, offers_failed, products_updated, products_added, products_deleted,
                 errors, warnings, duration_seconds, status, run_type, triggered_by, config_snapshot)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (run_id, files_found, files_processed, files_skipped, files_with_errors,
                  offers_sent, offers_failed, products_updated, products_added, products_deleted,
                  errors, warnings, duration, status, run_type, triggered_by,
                  json.dumps(config_snapshot) if config_snapshot else None))
            conn.commit()
        
        return run_id
    
    def get_stats(self, days: int = 30) -> pd.DataFrame:
        """Получение статистики запусков"""
        with self.get_connection() as conn:
            return pd.read_sql_query('''
                SELECT run_date, files_processed, offers_sent, products_updated, products_added,
                       errors, duration_seconds, status, run_type
                FROM run_stats 
                WHERE run_date > datetime('now', '-' || ? || ' days')
                ORDER BY run_date DESC
            ''', conn, params=(days,))
    
    def get_price_history(self, sku: Optional[str] = None, days: int = 30,
                         supplier: Optional[str] = None) -> pd.DataFrame:
        """Расширенное получение истории цен"""
        query = '''
            SELECT * FROM price_history 
            WHERE updated_at > datetime('now', '-' || ? || ' days')
        '''
        params = [days]
        
        if sku:
            query += ' AND sku = ?'
            params.append(sku)
        
        if supplier:
            query += ' AND supplier = ?'
            params.append(supplier)
        
        query += ' ORDER BY updated_at DESC LIMIT 1000'
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=params)
    
    def get_processed_files(self, days: int = 7) -> pd.DataFrame:
        """Получение списка обработанных файлов"""
        with self.get_connection() as conn:
            return pd.read_sql_query('''
                SELECT filename, supplier, processed_at, offers_count, 
                       valid_offers, invalid_offers, status, file_size, processing_time
                FROM processed_files 
                WHERE processed_at > datetime('now', '-' || ? || ' days')
                ORDER BY processed_at DESC
            ''', conn, params=(days,))
    
    def create_alert(self, alert_type: str, title: str, message: str = '',
                    severity: str = 'info', data: Optional[Dict] = None,
                    assigned_to: str = '') -> int:
        """Создание алерта"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO alerts 
                (alert_type, severity, title, message, data, assigned_to)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (alert_type, severity, title, message,
                  json.dumps(data) if data else None, assigned_to))
            conn.commit()
            return cursor.lastrowid
    
    def get_unresolved_alerts(self, alert_type: Optional[str] = None) -> pd.DataFrame:
        """Получение неразрешенных алертов"""
        query = '''
            SELECT * FROM alerts 
            WHERE is_resolved = 0
        '''
        params = []
        
        if alert_type:
            query += ' AND alert_type = ?'
            params.append(alert_type)
        
        query += ' ORDER BY created_at DESC'
        
        with self.get_connection() as conn:
            return pd.read_sql_query(query, conn, params=params)
    
    def resolve_alert(self, alert_id: int) -> None:
        """Разрешение алерта"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE alerts 
                SET is_resolved = 1, resolved_at = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (alert_id,))
            conn.commit()
    
    def get_database_stats(self) -> Dict[str, Any]:
        """Получение статистики базы данных"""
        stats = {}
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Размер базы данных
            stats['db_size_mb'] = os.path.getsize(self.db_path) / (1024 * 1024)
            
            # Количество записей в основных таблицах
            tables = ['processed_files', 'analysis_history', 'price_comparisons',
                     'price_history', 'run_stats', 'sent_offers_cache', 'alerts']
            
            for table in tables:
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                stats[f'{table}_count'] = cursor.fetchone()[0]
            
            # Статистика по статусам
            cursor.execute("SELECT status, COUNT(*) FROM processed_files GROUP BY status")
            stats['file_statuses'] = dict(cursor.fetchall())
            
            cursor.execute("SELECT status, COUNT(*) FROM run_stats GROUP BY status")
            stats['run_statuses'] = dict(cursor.fetchall())
        
        return stats
# ===================================================================
# БЛОК 6: РАСШИРЕННАЯ ТОВАРНАЯ БАЗА В GOOGLE SHEETS
# ===================================================================

class GoogleSheetsDatabase:
    """Расширенный класс для работы с товарной базой в Google Sheets"""
    
    REQUIRED_COLUMNS = ['артикул', 'цена_базовая']
    OPTIONAL_COLUMNS = ['бренд', 'название', 'цена_розница', 'остаток', 'категория', 
                       'описание', 'вес', 'размеры', 'штрихкод', 'страна', 'гарантия']
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.client = None
        self.sheet = None
        self.df: Optional[pd.DataFrame] = None
        self.last_backup_time: Optional[datetime] = None
        self.cache: Dict[str, Any] = {}
        self.cache_ttl = 300  # 5 минут кэширования
        self._connect()
        self._load_data()
    
    def _connect(self) -> None:
        """Расширенное подключение к Google Sheets с проверкой и переподключением"""
        try:
            self.logger.info("Подключение к Google Sheets...")
            
            if not os.path.exists(self.config.google_credentials_json):
                self.logger.error(f"Файл {self.config.google_credentials_json} не найден")
                self._create_default_credentials()
                return
            
            scope = self.config.google_scope
            creds = ServiceAccountCredentials.from_json_keyfile_name(
                self.config.google_credentials_json, 
                scope
            )
            
            self.client = gspread.authorize(creds)
            
            # Проверяем подключение
            if self.config.google_sheet_id:
                try:
                    test_sheet = self.client.open_by_key(self.config.google_sheet_id)
                    self.logger.success(f"Подключение к Google Sheets успешно. Таблица: {test_sheet.title}")
                except Exception as e:
                    self.logger.warning(f"Не удалось открыть таблицу: {e}")
                    self._create_new_spreadsheet()
            else:
                self.logger.warning("ID Google Sheets не указан, создаем новую таблицу")
                self._create_new_spreadsheet()
            
        except Exception as e:
            self.logger.error(f"Ошибка подключения к Google Sheets: {e}")
            raise
    
    def _create_default_credentials(self) -> None:
        """Создание файла с учетными данными по умолчанию"""
        try:
            default_creds = {
                "type": "service_account",
                "project_id": "your-project-id",
                "private_key_id": "your-private-key-id",
                "private_key": "-----BEGIN PRIVATE KEY-----\nYOUR-PRIVATE-KEY\n-----END PRIVATE KEY-----\n",
                "client_email": "your-service-account@your-project.iam.gserviceaccount.com",
                "client_id": "your-client-id",
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/your-service-account%40your-project.iam.gserviceaccount.com"
            }
            
            with open(self.config.google_credentials_json, 'w', encoding='utf-8') as f:
                json.dump(default_creds, f, indent=2)
            
            self.logger.warning(f"Создан шаблон {self.config.google_credentials_json}. Заполните реальные данные!")
        except Exception as e:
            self.logger.error(f"Ошибка создания файла учетных данных: {e}")
    
    def _create_new_spreadsheet(self) -> None:
        """Создание новой таблицы Google Sheets"""
        try:
            if not self.client:
                self.logger.error("Нет подключения к Google Sheets")
                return
            
            spreadsheet = self.client.create('База товаров робота')
            self.config.google_sheet_id = spreadsheet.id
            self.config.save()
            
            self.logger.success(f"Создана новая таблица: {spreadsheet.id}")
            self.sheet = spreadsheet
            
            # Создаем лист с товарами
            self._create_default_sheet()
            
        except Exception as e:
            self.logger.error(f"Ошибка создания таблицы: {e}")
    
    def _load_data(self) -> None:
        """Расширенная загрузка данных из Google Sheets с кэшированием"""
        try:
            # Проверяем кэш
            cache_key = f"gs_data_{self.config.google_sheet_id}_{self.config.google_sheet_name}"
            if cache_key in self.cache:
                cache_time, cached_data = self.cache[cache_key]
                if time.time() - cache_time < self.cache_ttl:
                    self.df = cached_data.copy()
                    self.logger.debug("Данные загружены из кэша")
                    return
            
            if not self.client:
                self._connect()
            
            if not self.config.google_sheet_id:
                self.logger.warning("ID Google Sheets не указан")
                return
            
            self.sheet = self.client.open_by_key(self.config.google_sheet_id)
            worksheet = self.sheet.worksheet(self.config.google_sheet_name)
            data = worksheet.get_all_values()
            
            if not data:
                self.logger.warning("Таблица пуста")
                self._create_default_sheet()
                return
            
            headers = data[0]
            rows = data[1:]
            headers = [h.strip().lower() for h in headers]
            
            self.df = pd.DataFrame(rows, columns=headers)
            
            # Сохраняем в кэш
            self.cache[cache_key] = (time.time(), self.df.copy())
            
            self._validate_and_fix()
            
            self.logger.info(f"Загружено {len(self.df)} товаров из Google Sheets")
            
        except gspread.exceptions.SpreadsheetNotFound:
            self.logger.error(f"Таблица {self.config.google_sheet_id} не найдена")
            self._create_new_spreadsheet()
        except gspread.exceptions.WorksheetNotFound:
            self.logger.warning(f"Лист {self.config.google_sheet_name} не найден, создаем новый")
            self._create_default_sheet()
        except Exception as e:
            self.logger.error(f"Ошибка загрузки данных из Google Sheets: {e}")
            self._create_default_sheet()
    
    def _validate_and_fix(self) -> None:
        """Расширенная валидация и исправление данных"""
        if self.df is None:
            return
        
        changes_made = False
        
        # Проверяем обязательные колонки
        missing = [col for col in self.REQUIRED_COLUMNS if col not in self.df.columns]
        if missing:
            for col in missing:
                self.df[col] = ''
            changes_made = True
            self.logger.warning(f"Добавлены отсутствующие колонки: {missing}")
        
        # Добавляем опциональные колонки если их нет
        for col in self.OPTIONAL_COLUMNS:
            if col not in self.df.columns:
                self.df[col] = ''
                changes_made = True
        
        # Очистка и нормализация данных
        self.df['артикул'] = self.df['артикул'].astype(str).str.strip().str.upper()
        self.df = self.df[self.df['артикул'] != '']
        
        # Удаляем дубликаты
        duplicates = self.df['артикул'].duplicated()
        if duplicates.any():
            self.logger.warning(f"Найдено {duplicates.sum()} дубликатов артикулов")
            self.df = self.df.drop_duplicates(subset=['артикул'], keep='first')
            changes_made = True
        
        # Нормализация цен
        self.df['цена_базовая'] = pd.to_numeric(self.df['цена_базовая'], errors='coerce').fillna(0)
        self.df['цена_розница'] = pd.to_numeric(self.df['цена_розница'], errors='coerce').fillna(self.df['цена_базовая'])
        
        # Если розничная цена равна 0, устанавливаем её равной базовой с наценкой
        zero_retail_mask = self.df['цена_розница'] == 0
        if zero_retail_mask.any():
            self.df.loc[zero_retail_mask, 'цена_розница'] = self.df.loc[zero_retail_mask, 'цена_базовая'] * (1 + self.config.markup_percent / 100)
            changes_made = True
        
        # Нормализация остатков
        self.df['остаток'] = pd.to_numeric(self.df['остаток'], errors='coerce').fillna(0).astype(int)
        
        # Исправляем отрицательные остатки
        negative_stock = self.df['остаток'] < 0
        if negative_stock.any():
            self.logger.warning(f"Исправлено {negative_stock.sum()} отрицательных остатков")
            self.df.loc[negative_stock, 'остаток'] = 0
            changes_made = True
        
        # Заполняем пустые текстовые поля
        text_columns = ['бренд', 'название', 'категория', 'описание', 'страна', 'гарантия']
        for col in text_columns:
            if col in self.df.columns:
                self.df[col] = self.df[col].fillna('').astype(str).str.strip()
        
        # Нормализация числовых полей
        numeric_columns = ['вес']
        for col in numeric_columns:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0)
        
        if changes_made:
            self._save()
    
    def _create_default_sheet(self) -> None:
        """Создание листа с примерами товаров и форматированием"""
        try:
            if not self.client or not self.config.google_sheet_id:
                return
            
            sheet = self.client.open_by_key(self.config.google_sheet_id)
            
            try:
                # Удаляем существующий лист если есть
                worksheet = sheet.worksheet(self.config.google_sheet_name)
                sheet.del_worksheet(worksheet)
            except:
                pass
            
            # Создаем новый лист
            worksheet = sheet.add_worksheet(
                self.config.google_sheet_name, 
                rows=1000, 
                cols=len(self.REQUIRED_COLUMNS) + len(self.OPTIONAL_COLUMNS) + 5
            )
            
            # Добавляем заголовки с форматированием
            headers = self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS + ['дата_обновления', 'источник']
            header_range = worksheet.range(1, 1, 1, len(headers))
            
            for i, header in enumerate(headers):
                header_range[i].value = header
            
            worksheet.update_cells(header_range)
            
            # Форматируем заголовки
            worksheet.format('A1:Z1', {
                'backgroundColor': {'red': 0.2, 'green': 0.6, 'blue': 0.8},
                'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
                'horizontalAlignment': 'CENTER'
            })
            
            # Добавляем примеры товаров
            sample_data = [
                ['SKU001', 1000.00, 'Apple', 'iPhone 15 Pro', 1300.00, 10, 'Смартфоны', 
                 'Новейший смартфон Apple', 0.2, '15x7x0.8', '1234567890123', 'Китай', '12 месяцев',
                 datetime.now().strftime('%Y-%m-%d'), 'Пример'],
                ['SKU002', 800.00, 'Samsung', 'Galaxy S24', 1000.00, 5, 'Смартфоны',
                 'Флагманский смартфон Samsung', 0.18, '14x6.5x0.7', '2345678901234', 'Вьетнам', '12 месяцев',
                 datetime.now().strftime('%Y-%m-%d'), 'Пример'],
                ['SKU003', 400.00, 'Xiaomi', 'Redmi Note 13', 500.00, 8, 'Смартфоны',
                 'Доступный смартфон Xiaomi', 0.19, '16x7.5x0.8', '3456789012345', 'Китай', '12 месяцев',
                 datetime.now().strftime('%Y-%m-%d'), 'Пример'],
                ['SKU004', 250.00, 'Sony', 'WH-1000XM5', 350.00, 15, 'Наушники',
                 'Беспроводные наушники с шумоподавлением', 0.25, '20x18x5', '4567890123456', 'Япония', '6 месяцев',
                 datetime.now().strftime('%Y-%m-%d'), 'Пример'],
                ['SKU005', 1500.00, 'Dell', 'XPS 15', 1800.00, 3, 'Ноутбуки',
                 'Мощный ноутбук для работы', 2.0, '35x24x1.5', '5678901234567', 'США', '24 месяца',
                 datetime.now().strftime('%Y-%m-%d'), 'Пример']
            ]
            
            for i, row in enumerate(sample_data, start=2):
                cell_range = worksheet.range(i, 1, i, len(row))
                for j, value in enumerate(row):
                    cell_range[j].value = value
                worksheet.update_cells(cell_range)
            
            self.logger.info("Создан лист с примерами товаров и форматированием")
            self._load_data()
            
        except Exception as e:
            self.logger.error(f"Ошибка создания листа: {e}")
    
    def _save(self) -> None:
        """Расширенное сохранение данных в Google Sheets с бэкапом"""
        try:
            if self.df is None or self.df.empty:
                return
            
            # Создаем бэкап если нужно
            if self.config.google_auto_backup:
                if self.last_backup_time is None or \
                   (datetime.now() - self.last_backup_time).total_seconds() > self.config.google_backup_interval_hours * 3600:
                    self._backup_data()
            
            worksheet = self.sheet.worksheet(self.config.google_sheet_name)
            
            # Очищаем лист (кроме заголовков)
            worksheet.clear()
            
            # Подготавливаем данные для записи
            headers = list(self.df.columns)
            worksheet.append_row(headers)
            
            # Записываем данные чанками для производительности
            chunk_size = 100
            for start_idx in range(0, len(self.df), chunk_size):
                end_idx = min(start_idx + chunk_size, len(self.df))
                chunk = self.df.iloc[start_idx:end_idx]
                
                rows_data = []
                for _, row in chunk.iterrows():
                    row_values = []
                    for col in headers:
                        value = row[col]
                        # Конвертируем специальные типы
                        if isinstance(value, (datetime, date)):
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, (np.integer,)):
                            value = int(value)
                        elif isinstance(value, (np.floating,)):
                            value = float(value)
                        elif isinstance(value, (np.bool_,)):
                            value = bool(value)
                        elif pd.isna(value):
                            value = ''
                        row_values.append(str(value))
                    rows_data.append(row_values)
                
                # Добавляем данные
                worksheet.append_rows(rows_data)
            
            # Обновляем кэш
            cache_key = f"gs_data_{self.config.google_sheet_id}_{self.config.google_sheet_name}"
            self.cache[cache_key] = (time.time(), self.df.copy())
            
            self.logger.info(f"Данные сохранены в Google Sheets: {len(self.df)} товаров")
            
        except gspread.exceptions.APIError as e:
            self.logger.error(f"Ошибка API Google при сохранении: {e}")
            self._handle_api_error(e)
        except Exception as e:
            self.logger.error(f"Ошибка сохранения в Google Sheets: {e}")
    
    def _handle_api_error(self, error: Exception) -> None:
        """Обработка ошибок API Google"""
        # Ждем и пробуем переподключиться
        time.sleep(5)
        try:
            self._connect()
            self._save()
        except Exception as e:
            self.logger.critical(f"Не удалось сохранить данные после ошибки API: {e}")
    
    def _backup_data(self) -> None:
        """Создание резервной копии данных"""
        try:
            if self.df is None:
                return
            
            backup_dir = Path(self.config.backup_dir)
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = backup_dir / f"gs_backup_{timestamp}.parquet"
            
            # Сохраняем в Parquet для эффективности
            self.df.to_parquet(backup_file)
            
            # Также сохраняем в CSV для совместимости
            csv_file = backup_dir / f"gs_backup_{timestamp}.csv"
            self.df.to_csv(csv_file, index=False, encoding='utf-8-sig')
            
            self.last_backup_time = datetime.now()
            self.logger.info(f"Создан бэкап: {backup_file}")
            
            # Удаляем старые бэкапы (оставляем последние 10)
            backup_files = sorted(backup_dir.glob("gs_backup_*.parquet"))
            if len(backup_files) > 10:
                for old_backup in backup_files[:-10]:
                    old_backup.unlink()
                    # Удаляем соответствующий CSV
                    csv_old = old_backup.with_suffix('.csv')
                    if csv_old.exists():
                        csv_old.unlink()
                    
        except Exception as e:
            self.logger.error(f"Ошибка создания бэкапа: {e}")
    
    def restore_from_backup(self, backup_file: str) -> bool:
        """Восстановление данных из бэкапа"""
        try:
            if backup_file.endswith('.parquet'):
                self.df = pd.read_parquet(backup_file)
            elif backup_file.endswith('.csv'):
                self.df = pd.read_csv(backup_file, encoding='utf-8-sig')
            else:
                self.logger.error(f"Неподдерживаемый формат бэкапа: {backup_file}")
                return False
            
            self._save()
            self.logger.success(f"Данные восстановлены из {backup_file}")
            return True
        except Exception as e:
            self.logger.error(f"Ошибка восстановления из бэкапа: {e}")
            return False
    
    def get_product_by_sku(self, sku: str) -> Optional[Dict]:
        """Получение товара по артикулу"""
        if self.df is None:
            return None
        
        product = self.df[self.df['артикул'] == str(sku).strip().upper()]
        if not product.empty:
            return product.iloc[0].to_dict()
        return None
    
    def get_all_products(self) -> pd.DataFrame:
        """Получение всех товаров"""
        if self.df is None:
            return pd.DataFrame()
        return self.df.copy()
    
    def update_product(self, sku: str, **kwargs) -> bool:
        """Расширенное обновление товара с валидацией"""
        if self.df is None:
            return False
        
        sku = str(sku).strip().upper()
        mask = self.df['артикул'] == sku
        
        if not mask.any():
            # Товар не найден, создаем новый
            new_row = {col: '' for col in self.df.columns}
            new_row['артикул'] = sku
            new_row.update(kwargs)
            new_row['дата_обновления'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            self.df = pd.concat([self.df, pd.DataFrame([new_row])], ignore_index=True)
            self._save()
            return True
        
        # Обновляем существующий товар
        for key, value in kwargs.items():
            if key in self.df.columns:
                self.df.loc[mask, key] = value
        
        self.df.loc[mask, 'дата_обновления'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        self._save()
        return True
    
    def batch_update_products(self, updates: List[Dict]) -> Tuple[int, int, List[str]]:
        """Пакетное обновление товаров"""
        updated = 0
        added = 0
        errors = []
        
        for update in updates:
            try:
                sku = update.get('sku', '')
                if not sku:
                    continue
                
                # Удаляем sku из словаря обновления
                update_data = {k: v for k, v in update.items() if k != 'sku'}
                
                if self.update_product(sku, **update_data):
                    if self.get_product_by_sku(sku):
                        updated += 1
                    else:
                        added += 1
            except Exception as e:
                errors.append(f"Ошибка обновления {update.get('sku', 'unknown')}: {str(e)}")
        
        return updated, added, errors
    
    def get_products_for_yandex(self, markup_percent: float = 0, 
                               min_stock_threshold: int = 1) -> List[Dict]:
        """Расширенная подготовка товаров для Яндекс Маркета"""
        if self.df is None:
            return []
        
        offers = []
        
        for _, row in self.df.iterrows():
            try:
                sku = str(row['артикул']).strip()
                if not sku:
                    continue
                
                # Рассчитываем цену
                base_price = float(row.get('цена_базовая', 0))
                retail_price = float(row.get('цена_розница', 0))
                
                if retail_price <= 0:
                    retail_price = base_price * (1 + markup_percent / 100)
                elif markup_percent > 0 and retail_price == base_price:
                    retail_price = base_price * (1 + markup_percent / 100)
                
                # Определяем остаток для отправки
                stock = int(row.get('остаток', 0))
                if stock < min_stock_threshold:
                    stock_to_send = 0
                else:
                    stock_to_send = stock
                
                # Подготавливаем оффер
                offer = {
                    'sku': sku,
                    'price': round(retail_price, 2),
                    'stock': stock_to_send,
                    'brand': str(row.get('бренд', '')),
                    'name': str(row.get('название', '')),
                    'description': str(row.get('описание', '')),
                    'category': str(row.get('категория', '')),
                    'barcode': str(row.get('штрихкод', '')),
                    'weight': float(row.get('вес', 0)),
                    'country': str(row.get('страна', '')),
                    'warranty': str(row.get('гарантия', ''))
                }
                
                offers.append(offer)
                
            except Exception as e:
                self.logger.error(f"Ошибка подготовки товара {row.get('артикул', 'unknown')}: {e}")
                continue
        
        return offers
    
    def get_stats(self) -> Dict:
        """Расширенная статистика товарной базы"""
        if self.df is None or self.df.empty:
            return {
                'total_products': 0,
                'total_brands': 0,
                'total_categories': 0,
                'avg_price': 0,
                'median_price': 0,
                'total_stock': 0,
                'min_price': 0,
                'max_price': 0,
                'products_with_stock': 0,
                'products_without_stock': 0,
                'products_with_price': 0,
                'products_without_price': 0,
                'avg_markup': 0,
                'total_value': 0,
                'last_update': None
            }
        
        prices = pd.to_numeric(self.df['цена_розница'], errors='coerce')
        base_prices = pd.to_numeric(self.df['цена_базовая'], errors='coerce')
        stocks = pd.to_numeric(self.df['остаток'], errors='coerce').fillna(0)
        
        # Расчет наценки
        markup = ((prices - base_prices) / base_prices * 100).replace([np.inf, -np.inf], np.nan)
        
        return {
            'total_products': len(self.df),
            'total_brands': len(self.df['бренд'].dropna().unique()),
            'total_categories': len(self.df['категория'].dropna().unique()),
            'avg_price': round(prices.mean(), 2),
            'median_price': round(prices.median(), 2),
            'total_stock': int(stocks.sum()),
            'min_price': round(prices.min(), 2),
            'max_price': round(prices.max(), 2),
            'products_with_stock': int((stocks > 0).sum()),
            'products_without_stock': int((stocks == 0).sum()),
            'products_with_price': int((prices > 0).sum()),
            'products_without_price': int((prices == 0).sum()),
            'avg_markup': round(markup.mean(), 2),
            'total_value': round((prices * stocks).sum(), 2),
            'last_update': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def search_products(self, query: str, fields: Optional[List[str]] = None) -> pd.DataFrame:
        """Расширенный поиск товаров по нескольким полям"""
        if self.df is None:
            return pd.DataFrame()
        
        query = query.lower().strip()
        if not query:
            return self.df.copy()
        
        if fields is None:
            fields = ['артикул', 'бренд', 'название', 'категория', 'описание']
        
        mask = pd.Series(False, index=self.df.index)
        
        for field in fields:
            if field in self.df.columns:
                mask |= self.df[field].astype(str).str.lower().str.contains(query, na=False)
        
        return self.df[mask].copy()
    
    def export_to_excel(self, filename: Optional[str] = None) -> str:
        """Экспорт базы в Excel с форматированием"""
        if self.df is None:
            return ''
        
        if filename is None:
            filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        export_path = Path(self.config.exports_dir) / filename
        export_path.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            self.df.to_excel(writer, sheet_name='Товары', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Товары']
            
            # Форматирование
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Автоподбор ширины колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Добавляем автофильтр
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(self.df.columns))}{len(self.df) + 1}"
        
        self.logger.info(f"База экспортирована в {export_path}")
        return str(export_path)
    
    def get_price_distribution(self) -> Dict:
        """Получение распределения цен для аналитики"""
        if self.df is None:
            return {}
        
        prices = pd.to_numeric(self.df['цена_розница'], errors='coerce').dropna()
        
        if prices.empty:
            return {}
        
        return {
            'min': float(prices.min()),
            'max': float(prices.max()),
            'mean': float(prices.mean()),
            'median': float(prices.median()),
            'std': float(prices.std()),
            'q25': float(prices.quantile(0.25)),
            'q75': float(prices.quantile(0.75)),
            'histogram': np.histogram(prices, bins=10)[0].tolist(),
            'bin_edges': np.histogram(prices, bins=10)[1].tolist()
        }
    
    def get_category_stats(self) -> pd.DataFrame:
        """Статистика по категориям"""
        if self.df is None:
            return pd.DataFrame()
        
        stats = self.df.groupby('категория').agg({
            'артикул': 'count',
            'цена_розница': ['mean', 'min', 'max'],
            'остаток': 'sum'
        }).round(2)
        
        stats.columns = ['Количество', 'Средняя цена', 'Мин. цена', 'Макс. цена', 'Общий остаток']
        return stats.reset_index()
    
    def refresh_data(self) -> None:
        """Принудительное обновление данных из Google Sheets"""
        cache_key = f"gs_data_{self.config.google_sheet_id}_{self.config.google_sheet_name}"
        if cache_key in self.cache:
            del self.cache[cache_key]
        self._load_data()
    
    def clear_cache(self) -> None:
        """Очистка кэша"""
        self.cache.clear()
        self.logger.info("Кэш очищен")


# ===================================================================
# БЛОК 7: РАСШИРЕННЫЙ КЛИЕНТ ЯНДЕКС МАРКЕТ
# ===================================================================

class YandexMarketClient:
    """Расширенный клиент для работы с API Яндекс Маркета"""
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': f'Bearer {self.config.yandex_token}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        })
        self.base_url = self.config.yandex_api_url
        self.rate_limit_delay = 0.1  # Задержка между запросами для соблюдения лимитов
        self.last_request_time = 0
        self.request_count = 0
        self.max_requests_per_minute = 100
        self.request_window_start = time.time()
    
    def _rate_limit(self) -> None:
        """Соблюдение ограничений частоты запросов"""
        current_time = time.time()
        
        # Сбрасываем счетчик каждую минуту
        if current_time - self.request_window_start > 60:
            self.request_count = 0
            self.request_window_start = current_time
        
        # Если превысили лимит, ждем
        if self.request_count >= self.max_requests_per_minute:
            wait_time = 60 - (current_time - self.request_window_start) + 1
            self.logger.warning(f"Превышен лимит запросов, ожидание {wait_time:.1f} секунд")
            time.sleep(wait_time)
            self.request_count = 0
            self.request_window_start = time.time()
        
        # Минимальная задержка между запросами
        time_since_last = current_time - self.last_request_time
        if time_since_last < self.rate_limit_delay:
            time.sleep(self.rate_limit_delay - time_since_last)
        
        self.request_count += 1
        self.last_request_time = time.time()
    
    def _make_request(self, method: str, endpoint: str, data: Optional[Dict] = None,
                     params: Optional[Dict] = None, retry: bool = True,
                     timeout: Optional[int] = None) -> Tuple[int, Dict, Dict]:
        """Расширенный метод выполнения HTTP запросов"""
        url = f"{self.base_url}{endpoint}"
        retries = self.config.retry_count if retry else 1
        timeout = timeout or self.config.request_timeout
        
        headers = {}
        response_headers = {}
        
        for attempt in range(retries):
            try:
                self._rate_limit()
                
                if method.upper() == 'GET':
                    response = self.session.get(url, params=params, timeout=timeout)
                elif method.upper() == 'POST':
                    response = self.session.post(url, json=data, params=params, timeout=timeout)
                elif method.upper() == 'PUT':
                    response = self.session.put(url, json=data, params=params, timeout=timeout)
                elif method.upper() == 'DELETE':
                    response = self.session.delete(url, params=params, timeout=timeout)
                elif method.upper() == 'PATCH':
                    response = self.session.patch(url, json=data, params=params, timeout=timeout)
                else:
                    raise ValueError(f"Неподдерживаемый метод: {method}")
                
                response_headers = dict(response.headers)
                
                # Логируем информацию о запросе для отладки
                self.logger.debug(
                    f"API запрос: {method} {endpoint} -> Статус: {response.status_code}",
                    context={
                        'status_code': response.status_code,
                        'response_size': len(response.content),
                        'attempt': attempt + 1,
                        'url': url
                    }
                )
                
                try:
                    response_data = response.json() if response.content else {}
                except json.JSONDecodeError:
                    response_data = {
                        'error': 'Invalid JSON response',
                        'raw_response': response.text[:500]
                    }
                
                # Успешный ответ
                if 200 <= response.status_code < 300:
                    return response.status_code, response_data, response_headers
                
                # Клиентские ошибки (4xx)
                if 400 <= response.status_code < 500:
                    error_msg = response_data.get('error', response_data.get('message', f'HTTP {response.status_code}'))
                    
                    # Не повторяем для некоторых ошибок
                    if response.status_code in [400, 401, 403, 404, 409, 422]:
                        self.logger.error(f"API ошибка {response.status_code}: {error_msg}")
                        return response.status_code, response_data, response_headers
                    
                    if attempt < retries - 1:
                        delay = self.config.retry_delay * (self.config.retry_backoff_multiplier ** attempt)
                        self.logger.warning(f"Повторная попытка через {delay:.1f}с (попытка {attempt + 1}/{retries})")
                        time.sleep(delay)
                        continue
                
                # Серверные ошибки (5xx)
                if response.status_code >= 500:
                    if attempt < retries - 1:
                        delay = self.config.retry_delay * (self.config.retry_backoff_multiplier ** attempt)
                        self.logger.warning(f"Серверная ошибка, повтор через {delay:.1f}с")
                        time.sleep(delay)
                        continue
                
                return response.status_code, response_data, response_headers
                
            except requests.exceptions.Timeout:
                if attempt < retries - 1:
                    delay = self.config.retry_delay * (self.config.retry_backoff_multiplier ** attempt)
                    self.logger.warning(f"Таймаут, повтор через {delay:.1f}с")
                    time.sleep(delay)
                    continue
                return 408, {'error': 'Request timeout'}, {}
                
            except requests.exceptions.ConnectionError as e:
                if attempt < retries - 1:
                    delay = self.config.retry_delay * (self.config.retry_backoff_multiplier ** attempt)
                    self.logger.warning(f"Ошибка соединения, повтор через {delay:.1f}с")
                    time.sleep(delay)
                    continue
                return 503, {'error': f'Connection error: {str(e)}'}, {}
                
            except Exception as e:
                if attempt < retries - 1:
                    time.sleep(self.config.retry_delay)
                    continue
                return 500, {'error': str(e)}, {}
        
        return 500, {'error': 'Max retries exceeded'}, {}
    
    def test_connection(self) -> Tuple[bool, str]:
        """Расширенная проверка подключения к API"""
        try:
            # Проверяем наличие токена и ID кампании
            if not self.config.yandex_token:
                return False, "Не указан OAuth токен"
            
            if not self.config.campaign_id:
                return False, "Не указан ID кампании"
            
            # Проверяем подключение
            status, data, headers = self._make_request(
                'GET', 
                f'/campaigns/{self.config.campaign_id}'
            )
            
            if status == 200:
                campaign_name = data.get('campaign', {}).get('domain', 'Неизвестно')
                return True, f"Подключение успешно. Кампания: {campaign_name}"
            elif status == 401:
                return False, "Ошибка авторизации. Проверьте OAuth токен"
            elif status == 403:
                return False, "Нет доступа к кампании. Проверьте права"
            elif status == 404:
                return False, f"Кампания с ID {self.config.campaign_id} не найдена"
            else:
                error_msg = data.get('error', data.get('message', f'Статус {status}'))
                return False, f"Ошибка подключения: {error_msg}"
                
        except Exception as e:
            return False, f"Исключение при проверке: {str(e)}"
    
    def get_campaign_info(self) -> Optional[Dict]:
        """Получение информации о кампании"""
        status, data, headers = self._make_request(
            'GET',
            f'/campaigns/{self.config.campaign_id}'
        )
        
        if status == 200:
            return data.get('campaign', {})
        return None
    
    def update_prices(self, offers: List[Dict]) -> Tuple[bool, List[Dict], int]:
        """Расширенное обновление цен с детальной статистикой"""
        if not offers:
            return True, [], 0
        
        errors = []
        success_count = 0
        endpoint = f'/campaigns/{self.config.campaign_id}/offer-prices/updates'
        
        for i in range(0, len(offers), self.config.max_offers_per_request):
            batch = offers[i:i + self.config.max_offers_per_request]
            
            # Подготавливаем данные для отправки
            price_items = []
            for offer in batch:
                item = {
                    "offerId": offer['sku'],
                    "price": {
                        "value": float(offer['price']),
                        "currencyId": "RUR"
                    }
                }
                price_items.append(item)
            
            payload = {"offers": price_items}
            
            status, data, headers = self._make_request('POST', endpoint, payload)
            
            if status == 200:
                batch_success = len(batch)
                success_count += batch_success
                self.logger.info(f"Цены обновлены для {batch_success} товаров (пакет {i // self.config.max_offers_per_request + 1})")
            elif status == 202:
                # Запрос принят, но требует времени на обработку
                self.logger.info(f"Запрос на обновление цен принят (пакет {i // self.config.max_offers_per_request + 1})")
                success_count += len(batch)
            else:
                error_msg = data.get('error', data.get('message', f'Статус {status}'))
                errors.append({
                    'batch': i // self.config.max_offers_per_request + 1,
                    'error': error_msg,
                    'offers_count': len(batch),
                    'status_code': status
                })
                self.logger.error(f"Ошибка обновления цен: {error_msg}")
        
        return len(errors) == 0, errors, success_count
    
    def update_stocks(self, offers: List[Dict]) -> Tuple[bool, List[Dict], int]:
        """Расширенное обновление остатков с детальной статистикой"""
        if not offers:
            return True, [], 0
        
        errors = []
        success_count = 0
        endpoint = f'/campaigns/{self.config.campaign_id}/offers/stocks'
        
        for i in range(0, len(offers), self.config.max_offers_per_request):
            batch = offers[i:i + self.config.max_offers_per_request]
            
            # Подготавливаем данные для отправки
            stock_items = []
            for offer in batch:
                item = {
                    "offerId": offer['sku'],
                    "stocks": [
                        {
                            "warehouseId": self.config.warehouse_id,
                            "count": int(offer['stock'])
                        }
                    ]
                }
                stock_items.append(item)
            
            payload = {"skus": stock_items}
            
            status, data, headers = self._make_request('PUT', endpoint, payload)
            
            if status == 200:
                batch_success = len(batch)
                success_count += batch_success
                self.logger.info(f"Остатки обновлены для {batch_success} товаров")
            elif status == 202:
                self.logger.info(f"Запрос на обновление остатков принят")
                success_count += len(batch)
            else:
                error_msg = data.get('error', data.get('message', f'Статус {status}'))
                errors.append({
                    'batch': i // self.config.max_offers_per_request + 1,
                    'error': error_msg,
                    'offers_count': len(batch),
                    'status_code': status
                })
                self.logger.error(f"Ошибка обновления остатков: {error_msg}")
        
        return len(errors) == 0, errors, success_count
    
    def get_offer_info(self, sku: str) -> Optional[Dict]:
        """Получение информации о конкретном оффере"""
        status, data, headers = self._make_request(
            'GET',
            f'/campaigns/{self.config.campaign_id}/offers/{sku}'
        )
        
        if status == 200:
            return data.get('offer', {})
        return None
    
    def get_all_offers(self, limit: int = 1000) -> List[Dict]:
        """Получение списка всех офферов"""
        all_offers = []
        page_token = None
        
        while True:
            params = {'limit': min(limit, 200)}
            if page_token:
                params['page_token'] = page_token
            
            status, data, headers = self._make_request(
                'GET',
                f'/campaigns/{self.config.campaign_id}/offers',
                params=params
            )
            
            if status == 200:
                offers = data.get('offers', [])
                all_offers.extend(offers)
                
                page_token = data.get('pager', {}).get('nextPageToken')
                if not page_token or len(all_offers) >= limit:
                    break
            else:
                break
        
        return all_offers[:limit]
    
    def get_price_suggestions(self, skus: List[str]) -> Dict:
        """Получение рекомендованных цен"""
        if not skus:
            return {}
        
        endpoint = f'/campaigns/{self.config.campaign_id}/offer-prices/suggestions'
        payload = {"offers": [{"offerId": sku} for sku in skus]}
        
        status, data, headers = self._make_request('POST', endpoint, payload)
        
        if status == 200:
            return data
        return {}
    
    def get_stats(self) -> Dict:
        """Получение общей статистики по кампании"""
        stats = {
            'prices_updated_today': 0,
            'stocks_updated_today': 0,
            'total_offers': 0,
            'active_offers': 0,
            'errors_today': 0
        }
        
        try:
            # Получаем общую статистику
            status, data, headers = self._make_request(
                'GET',
                f'/campaigns/{self.config.campaign_id}/stats/main'
            )
            
            if status == 200:
                main_stats = data.get('mainStats', {})
                stats['total_offers'] = main_stats.get('totalOffers', 0)
                stats['active_offers'] = main_stats.get('activeOffers', 0)
            
            # Получаем историю операций за сегодня
            today = datetime.now().strftime('%Y-%m-%d')
            status, data, headers = self._make_request(
                'GET',
                f'/campaigns/{self.config.campaign_id}/offer-mapping-entries/updates',
                params={'fromDate': today}
            )
            
            if status == 200:
                updates = data.get('result', {}).get('offerMappingEntries', [])
                stats['prices_updated_today'] = len(updates)
                
        except Exception as e:
            self.logger.error(f"Ошибка получения статистики: {e}")
        
        return stats
    
    def bulk_update_prices_and_stocks(self, offers: List[Dict]) -> Dict:
        """Массовое обновление цен и остатков"""
        result = {
            'prices': {'success': True, 'errors': [], 'updated': 0},
            'stocks': {'success': True, 'errors': [], 'updated': 0}
        }
        
        # Обновляем цены
        prices_success, prices_errors, prices_count = self.update_prices(offers)
        result['prices'] = {
            'success': prices_success,
            'errors': prices_errors,
            'updated': prices_count
        }
        
        # Обновляем остатки
        stocks_success, stocks_errors, stocks_count = self.update_stocks(offers)
        result['stocks'] = {
            'success': stocks_success,
            'errors': stocks_errors,
            'updated': stocks_count
        }
        
        return result
    
    def validate_offers(self, offers: List[Dict]) -> List[Dict]:
        """Валидация офферов перед отправкой"""
        validated = []
        errors = []
        
        for offer in offers:
            offer_errors = []
            
            # Проверка обязательных полей
            if not offer.get('sku'):
                offer_errors.append("Отсутствует артикул")
            
            price = offer.get('price', 0)
            if price <= 0:
                offer_errors.append(f"Некорректная цена: {price}")
            elif price > 999999999:
                offer_errors.append(f"Слишком большая цена: {price}")
            
            stock = offer.get('stock', -1)
            if stock < 0:
                offer_errors.append(f"Отрицательный остаток: {stock}")
            
            if offer_errors:
                errors.append({
                    'sku': offer.get('sku', 'unknown'),
                    'errors': offer_errors
                })
            else:
                validated.append(offer)
        
        if errors:
            self.logger.warning(f"Найдено {len(errors)} некорректных офферов")
            for error in errors[:10]:  # Логируем первые 10 ошибок
                self.logger.debug(f"Оффер {error['sku']}: {', '.join(error['errors'])}")
        
        return validated

# ===================================================================
# БЛОК 8: РАСШИРЕННЫЙ ЗАГРУЗЧИК ИЗ ПОЧТЫ
# ===================================================================

class EmailDownloader:
    """Расширенный класс для скачивания вложений из почты через IMAP"""
    
    ALLOWED_EXTENSIONS = ['.xls', '.xlsx', '.csv', '.xml', '.json', '.txt', '.xlsm', '.xlsb', '.ods']
    ALLOWED_MIME_TYPES = [
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'text/csv',
        'application/xml',
        'text/xml',
        'application/json',
        'text/plain',
        'application/vnd.oasis.opendocument.spreadsheet'
    ]
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.mail = None
        self.connection_attempts = 0
        self.max_connection_attempts = 3
        self.connection_timeout = 30
    
    def connect(self) -> bool:
        """Расширенное подключение к почтовому серверу с повторными попытками"""
        self.connection_attempts = 0
        
        while self.connection_attempts < self.max_connection_attempts:
            try:
                self.connection_attempts += 1
                self.logger.info(
                    f"Подключение к {self.config.imap_server}:{self.config.imap_port} "
                    f"(попытка {self.connection_attempts}/{self.max_connection_attempts})"
                )
                
                # Создаем подключение с таймаутом
                if self.config.imap_port == 993:
                    self.mail = imaplib.IMAP4_SSL(
                        self.config.imap_server, 
                        self.config.imap_port,
                        timeout=self.connection_timeout
                    )
                else:
                    self.mail = imaplib.IMAP4(
                        self.config.imap_server, 
                        self.config.imap_port,
                        timeout=self.connection_timeout
                    )
                
                # Отключаем проверку SSL если нужно
                if not self.config.email_ssl_verify:
                    self.logger.warning("Проверка SSL сертификата отключена")
                    import ssl
                    ssl_context = ssl.create_default_context()
                    ssl_context.check_hostname = False
                    ssl_context.verify_mode = ssl.CERT_NONE
                    if self.config.imap_port == 993:
                        self.mail = imaplib.IMAP4_SSL(
                            self.config.imap_server,
                            self.config.imap_port,
                            ssl_context=ssl_context,
                            timeout=self.connection_timeout
                        )
                
                # Логинимся
                self.mail.login(self.config.email_user, self.config.email_pass)
                
                # Выбираем папку INBOX
                status, data = self.mail.select('INBOX')
                if status != 'OK':
                    raise Exception(f"Не удалось выбрать папку INBOX: {data}")
                
                # Получаем информацию о папке
                status, mailbox_info = self.mail.status('INBOX', '(MESSAGES UNSEEN RECENT)')
                if status == 'OK':
                    info_str = mailbox_info[0].decode('utf-8', errors='ignore')
                    self.logger.info(f"Состояние почтового ящика: {info_str}")
                
                self.logger.info("Подключение к почте успешно установлено")
                return True
                
            except imaplib.IMAP4.error as e:
                error_msg = str(e)
                if 'AUTHENTICATIONFAILED' in error_msg.upper() or 'LOGIN FAILED' in error_msg.upper():
                    self.logger.error(f"Ошибка аутентификации: неверный логин или пароль")
                    return False
                elif 'Temporary authentication failure' in error_msg:
                    self.logger.warning(f"Временная ошибка аутентификации, ожидание...")
                    time.sleep(5 * self.connection_attempts)
                else:
                    self.logger.error(f"Ошибка IMAP: {error_msg}")
                    
            except imaplib.IMAP4.abort as e:
                self.logger.error(f"Соединение прервано: {e}")
                time.sleep(3 * self.connection_attempts)
                
            except socket.timeout:
                self.logger.error(f"Таймаут подключения (попытка {self.connection_attempts})")
                time.sleep(2 * self.connection_attempts)
                
            except socket.gaierror as e:
                self.logger.error(f"Ошибка разрешения имени сервера: {e}")
                return False
                
            except Exception as e:
                self.logger.error(f"Неожиданная ошибка подключения: {e}")
                time.sleep(5 * self.connection_attempts)
        
        self.logger.error(f"Не удалось подключиться после {self.max_connection_attempts} попыток")
        return False
    
    def disconnect(self) -> None:
        """Безопасное отключение от почтового сервера"""
        if self.mail:
            try:
                self.mail.close()
                self.mail.logout()
                self.logger.info("Отключение от почтового сервера выполнено")
            except Exception as e:
                self.logger.warning(f"Ошибка при отключении: {e}")
            finally:
                self.mail = None
    
    def check_connection(self) -> bool:
        """Проверка активности соединения и переподключение при необходимости"""
        if not self.mail:
            return self.connect()
        
        try:
            status, _ = self.mail.noop()
            if status != 'OK':
                self.logger.warning("Соединение потеряно, переподключение...")
                self.disconnect()
                return self.connect()
            return True
        except Exception:
            self.logger.warning("Ошибка проверки соединения, переподключение...")
            try:
                self.disconnect()
            except:
                pass
            return self.connect()
    
    def search_emails(self, custom_criteria: Optional[List[str]] = None) -> List[Dict]:
        """Расширенный поиск писем с дополнительными критериями"""
        if not self.check_connection():
            return []
        
        try:
            criteria = custom_criteria or []
            
            # Добавляем фильтр по отправителю
            if self.config.sender_filter:
                criteria.append(f'FROM "{self.config.sender_filter}"')
            
            # Добавляем фильтр по теме
            if self.config.subject_filter:
                criteria.append(f'SUBJECT "{self.config.subject_filter}"')
            
            # Добавляем фильтр по дате
            since_date = (datetime.now() - timedelta(days=self.config.email_search_days)).strftime("%d-%b-%Y")
            criteria.append(f'SINCE {since_date}')
            
            # Добавляем фильтр непрочитанных если нужно
            criteria.append('UNSEEN')
            
            # Формируем строку поиска
            search_criteria = ' '.join(criteria) if criteria else 'ALL'
            
            self.logger.debug(f"Критерии поиска: {search_criteria}")
            
            # Выполняем поиск
            status, messages = self.mail.search(None, search_criteria)
            
            if status != 'OK':
                self.logger.error(f"Ошибка поиска писем: {messages}")
                return []
            
            if not messages[0]:
                self.logger.info("Писем по заданным критериям не найдено")
                return []
            
            email_ids = messages[0].split()
            
            # Ограничиваем количество обрабатываемых писем
            if len(email_ids) > self.config.email_max_emails:
                self.logger.info(f"Найдено {len(email_ids)} писем, обрабатываем последние {self.config.email_max_emails}")
                email_ids = email_ids[-self.config.email_max_emails:]
            else:
                self.logger.info(f"Найдено {len(email_ids)} писем для обработки")
            
            # Получаем информацию о письмах
            email_list = []
            for msg_id in email_ids:
                try:
                    # Получаем ENVELOPE для быстрого получения метаданных
                    status, msg_data = self.mail.fetch(msg_id, '(ENVELOPE FLAGS INTERNALDATE)')
                    if status != 'OK':
                        self.logger.warning(f"Не удалось получить данные письма {msg_id}")
                        continue
                    
                    # Парсим ответ сервера
                    envelope_data = None
                    flags = []
                    internal_date = None
                    
                    for part in msg_data:
                        if isinstance(part, tuple):
                            envelope_str = part[0].decode('utf-8', errors='ignore') if isinstance(part[0], bytes) else str(part[0])
                            
                            # Извлекаем флаги
                            if b'FLAGS' in part[0] if isinstance(part[0], bytes) else 'FLAGS' in envelope_str:
                                flags_match = re.search(r'FLAGS\s*\(([^)]*)\)', envelope_str)
                                if flags_match:
                                    flags = flags_match.group(1).split()
                            
                            # Извлекаем INTERNALDATE
                            if b'INTERNALDATE' in part[0] if isinstance(part[0], bytes) else 'INTERNALDATE' in envelope_str:
                                date_match = re.search(r'INTERNALDATE\s*"([^"]*)"', envelope_str)
                                if date_match:
                                    internal_date = date_match.group(1)
                            
                            envelope_data = envelope_str
                    
                    # Получаем ENVELOPE отдельно если нужно
                    status, envelope_response = self.mail.fetch(msg_id, '(ENVELOPE)')
                    envelope = None
                    if status == 'OK' and envelope_response[0]:
                        import email
                        msg = email.message_from_bytes(envelope_response[0][1])
                        envelope = {
                            'subject': msg.get('Subject', ''),
                            'from': msg.get('From', ''),
                            'date': msg.get('Date', ''),
                            'message_id': msg.get('Message-ID', '')
                        }
                    
                    # Декодируем тему
                    subject = ''
                    if envelope and envelope.get('subject'):
                        subject_parts = decode_header(envelope['subject'])
                        subject = ''
                        for part, encoding in subject_parts:
                            if isinstance(part, bytes):
                                try:
                                    subject += part.decode(encoding or 'utf-8', errors='ignore')
                                except:
                                    subject += part.decode('utf-8', errors='ignore')
                            else:
                                subject += str(part)
                    
                    from_addr = ''
                    if envelope and envelope.get('from'):
                        from_addr = envelope['from']
                    
                    email_date = None
                    if envelope and envelope.get('date'):
                        try:
                            email_date = parsedate_to_datetime(envelope['date'])
                        except:
                            email_date = None
                    
                    if not email_date and internal_date:
                        try:
                            from email.utils import parsedate_to_datetime as parse_dt
                            email_date = parse_dt(internal_date)
                        except:
                            email_date = datetime.now()
                    
                    is_read = b'\\Seen' in str(flags).encode() if flags else False
                    
                    email_list.append({
                        'id': msg_id.decode() if isinstance(msg_id, bytes) else msg_id,
                        'subject': subject,
                        'from': from_addr,
                        'date': email_date,
                        'is_read': is_read,
                        'flags': flags,
                        'size': len(str(envelope_data)) if envelope_data else 0
                    })
                    
                except Exception as e:
                    self.logger.error(f"Ошибка обработки письма {msg_id}: {e}")
                    continue
            
            # Сортируем по дате (новые сначала)
            email_list.sort(key=lambda x: x.get('date') or datetime.min, reverse=True)
            
            return email_list
            
        except imaplib.IMAP4.error as e:
            self.logger.error(f"Ошибка IMAP при поиске: {e}")
            return []
        except Exception as e:
            self.logger.error(f"Неожиданная ошибка при поиске писем: {e}")
            return []
    
    def download_attachments(self, email_data: Dict) -> List[Tuple[str, bytes, str]]:
        """Расширенное скачивание вложений с определением MIME-типов"""
        if not self.check_connection():
            return []
        
        attachments = []
        
        try:
            msg_id = email_data['id']
            
            # Получаем полное содержимое письма
            status, msg_data = self.mail.fetch(msg_id, '(RFC822)')
            
            if status != 'OK':
                self.logger.error(f"Не удалось получить содержимое письма {msg_id}")
                return []
            
            # Парсим письмо
            msg = email.message_from_bytes(msg_data[0][1], policy=email_default_policy)
            
            # Получаем основную информацию о письме
            subject = msg.get('Subject', '')
            from_addr = msg.get('From', '')
            date_str = msg.get('Date', '')
            
            self.logger.debug(
                f"Обработка вложений письма от {from_addr}: {subject[:100]}",
                context={'message_id': msg_id, 'from': from_addr, 'subject': subject[:100]}
            )
            
            # Проходим по всем частям письма
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get('Content-Disposition', ''))
                content_transfer_encoding = part.get('Content-Transfer-Encoding', '')
                
                # Определяем, является ли часть вложением
                is_attachment = False
                filename = None
                
                # Проверяем Content-Disposition
                if 'attachment' in content_disposition.lower():
                    is_attachment = True
                    filename = part.get_filename()
                
                # Проверяем наличие имени файла
                if not filename:
                    filename = part.get_filename()
                
                # Если есть имя файла, считаем вложением
                if filename:
                    is_attachment = True
                
                if not is_attachment or not filename:
                    continue
                
                # Декодируем имя файла
                try:
                    decoded_parts = decode_header(filename)
                    filename = ''
                    for decoded_part, encoding in decoded_parts:
                        if isinstance(decoded_part, bytes):
                            try:
                                filename += decoded_part.decode(encoding or 'utf-8', errors='ignore')
                            except:
                                filename += decoded_part.decode('utf-8', errors='ignore')
                        else:
                            filename += str(decoded_part)
                except Exception as e:
                    self.logger.debug(f"Ошибка декодирования имени файла: {e}")
                
                # Очищаем имя файла от недопустимых символов
                filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                
                # Проверяем расширение файла
                ext = os.path.splitext(filename)[1].lower()
                
                if ext not in self.ALLOWED_EXTENSIONS:
                    # Проверяем MIME-тип
                    if content_type not in self.ALLOWED_MIME_TYPES:
                        self.logger.debug(f"Пропускаем файл {filename} (тип: {content_type}, расширение: {ext})")
                        continue
                
                # Извлекаем содержимое
                try:
                    content = part.get_payload(decode=True)
                    if not content:
                        self.logger.warning(f"Пустое содержимое вложения: {filename}")
                        continue
                    
                    # Проверяем размер файла
                    file_size = len(content)
                    if file_size == 0:
                        self.logger.warning(f"Файл {filename} имеет нулевой размер")
                        continue
                    
                    if file_size > 100 * 1024 * 1024:  # 100 MB
                        self.logger.warning(f"Файл {filename} слишком большой ({file_size / 1024 / 1024:.1f} MB), пропускаем")
                        continue
                    
                    attachments.append((filename, content, content_type))
                    
                    self.logger.info(
                        f"Скачан файл: {filename}",
                        context={
                            'size': file_size,
                            'size_mb': round(file_size / 1024 / 1024, 2),
                            'mime_type': content_type,
                            'encoding': content_transfer_encoding
                        }
                    )
                    
                except Exception as e:
                    self.logger.error(f"Ошибка извлечения содержимого {filename}: {e}")
                    continue
            
            # Помечаем письмо как прочитанное если нужно
            if self.config.email_mark_as_read:
                try:
                    self.mail.store(msg_id, '+FLAGS', '\\Seen')
                    self.logger.debug(f"Письмо {msg_id} помечено как прочитанное")
                except Exception as e:
                    self.logger.warning(f"Не удалось пометить письмо как прочитанное: {e}")
            
            # Удаляем письмо если нужно
            if self.config.email_delete_after_processing:
                try:
                    self.mail.store(msg_id, '+FLAGS', '\\Deleted')
                    self.mail.expunge()
                    self.logger.info(f"Письмо {msg_id} удалено после обработки")
                except Exception as e:
                    self.logger.warning(f"Не удалось удалить письмо: {e}")
            
            return attachments
            
        except imaplib.IMAP4.error as e:
            self.logger.error(f"Ошибка IMAP при скачивании вложений: {e}")
            return []
        except Exception as e:
            self.logger.error(f"Ошибка скачивания вложений из письма {email_data.get('id', 'unknown')}: {e}")
            return []
    
    def get_mailbox_stats(self) -> Dict:
        """Получение статистики почтового ящика"""
        stats = {
            'total_messages': 0,
            'unread_messages': 0,
            'recent_messages': 0,
            'mailbox_size': 0,
            'last_check': datetime.now().isoformat()
        }
        
        if not self.check_connection():
            return stats
        
        try:
            # Получаем статус папки INBOX
            status, data = self.mail.status('INBOX', '(MESSAGES UNSEEN RECENT)')
            if status == 'OK':
                status_str = data[0].decode('utf-8', errors='ignore')
                
                # Парсим статистику
                messages_match = re.search(r'MESSAGES\s+(\d+)', status_str)
                unseen_match = re.search(r'UNSEEN\s+(\d+)', status_str)
                recent_match = re.search(r'RECENT\s+(\d+)', status_str)
                
                if messages_match:
                    stats['total_messages'] = int(messages_match.group(1))
                if unseen_match:
                    stats['unread_messages'] = int(unseen_match.group(1))
                if recent_match:
                    stats['recent_messages'] = int(recent_match.group(1))
            
            # Получаем квоту если поддерживается
            try:
                status, quota_data = self.mail.getquotaroot('INBOX')
                if status == 'OK' and quota_data:
                    quota_str = quota_data[0].decode('utf-8', errors='ignore') if isinstance(quota_data[0], bytes) else str(quota_data[0])
                    # Парсим информацию о квоте
                    usage_match = re.search(r'STORAGE\s+(\d+)\s+(\d+)', quota_str)
                    if usage_match:
                        stats['mailbox_usage_kb'] = int(usage_match.group(1))
                        stats['mailbox_limit_kb'] = int(usage_match.group(2))
                        stats['mailbox_usage_percent'] = round(
                            int(usage_match.group(1)) / int(usage_match.group(2)) * 100, 2
                        ) if int(usage_match.group(2)) > 0 else 0
            except:
                pass
            
        except Exception as e:
            self.logger.error(f"Ошибка получения статистики ящика: {e}")
        
        return stats
    
    def search_by_date_range(self, start_date: datetime, end_date: datetime,
                           sender: Optional[str] = None, subject: Optional[str] = None) -> List[Dict]:
        """Поиск писем за определенный период"""
        if not self.check_connection():
            return []
        
        try:
            criteria = []
            
            # Добавляем фильтры
            if sender:
                criteria.append(f'FROM "{sender}"')
            if subject:
                criteria.append(f'SUBJECT "{subject}"')
            
            # Формируем критерии даты
            start_str = start_date.strftime("%d-%b-%Y")
            end_str = end_date.strftime("%d-%b-%Y")
            criteria.append(f'SINCE {start_str}')
            criteria.append(f'BEFORE {end_str}')
            
            search_criteria = ' '.join(criteria)
            
            return self.search_emails([search_criteria])
            
        except Exception as e:
            self.logger.error(f"Ошибка поиска по диапазону дат: {e}")
            return []
    
    def download_all_attachments_from_folder(self, folder: str = 'INBOX', 
                                           max_emails: int = 50) -> List[Tuple[str, bytes, str]]:
        """Скачивание всех вложений из указанной папки"""
        all_attachments = []
        
        if not self.check_connection():
            return all_attachments
        
        try:
            # Выбираем папку
            status, data = self.mail.select(folder)
            if status != 'OK':
                self.logger.error(f"Не удалось выбрать папку {folder}")
                return all_attachments
            
            # Ищем все письма
            status, messages = self.mail.search(None, 'ALL')
            if status != 'OK' or not messages[0]:
                return all_attachments
            
            email_ids = messages[0].split()
            
            # Ограничиваем количество
            if len(email_ids) > max_emails:
                email_ids = email_ids[-max_emails:]
            
            for msg_id in email_ids:
                email_data = {
                    'id': msg_id.decode() if isinstance(msg_id, bytes) else msg_id,
                    'subject': '',
                    'from': '',
                    'date': None,
                    'is_read': True
                }
                
                attachments = self.download_attachments(email_data)
                all_attachments.extend(attachments)
                
                if len(all_attachments) >= max_emails:
                    break
            
            return all_attachments
            
        except Exception as e:
            self.logger.error(f"Ошибка скачивания вложений из папки {folder}: {e}")
            return all_attachments


# ===================================================================
# БЛОК 9: РАСШИРЕННЫЙ ПАРСЕР ПРАЙСОВ
# ===================================================================

class PriceParser:
    """Расширенный класс для парсинга прайс-файлов с автоопределением форматов"""
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.supported_encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'cp1252', 'latin-1', 
                                   'iso-8859-1', 'koi8-r', 'cp866', 'mac-cyrillic']
    
    def detect_encoding(self, content: bytes) -> str:
        """Автоматическое определение кодировки файла"""
        if CHARDET_SUPPORT:
            try:
                result = chardet.detect(content[:10000])  # Анализируем первые 10KB
                encoding = result.get('encoding', 'utf-8')
                confidence = result.get('confidence', 0)
                
                if confidence > 0.7:
                    self.logger.debug(f"Определена кодировка: {encoding} (уверенность: {confidence:.2f})")
                    return encoding
            except Exception as e:
                self.logger.debug(f"Ошибка определения кодировки chardet: {e}")
        
        # Пробуем популярные кодировки
        for encoding in self.supported_encodings:
            try:
                content[:100].decode(encoding)
                return encoding
            except:
                continue
        
        return 'utf-8'
    
    def detect_format(self, filename: str, content: Optional[bytes] = None) -> FileFormat:
        """Расширенное определение формата файла"""
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        
        format_map = {
            'xlsx': FileFormat.EXCEL_XLSX,
            'xls': FileFormat.EXCEL_XLS,
            'xlsm': FileFormat.EXCEL_XLSX,
            'xlsb': FileFormat.EXCEL_XLS,
            'ods': FileFormat.EXCEL_XLSX,
            'csv': FileFormat.CSV,
            'xml': FileFormat.XML,
            'json': FileFormat.JSON,
            'txt': FileFormat.TXT,
            'zip': FileFormat.ZIP,
            'tar': FileFormat.TAR,
            'gz': FileFormat.GZ,
            'tgz': FileFormat.GZ
        }
        
        if ext in format_map:
            return format_map[ext]
        
        # Если расширение не определено, пробуем по содержимому
        if content and MAGIC_SUPPORT:
            try:
                mime = magic.Magic(mime=True)
                mime_type = mime.from_buffer(content[:1024])
                
                mime_format_map = {
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': FileFormat.EXCEL_XLSX,
                    'application/vnd.ms-excel': FileFormat.EXCEL_XLS,
                    'text/csv': FileFormat.CSV,
                    'text/xml': FileFormat.XML,
                    'application/xml': FileFormat.XML,
                    'application/json': FileFormat.JSON,
                    'text/plain': FileFormat.TXT,
                    'application/zip': FileFormat.ZIP,
                    'application/x-tar': FileFormat.TAR,
                    'application/gzip': FileFormat.GZ
                }
                
                if mime_type in mime_format_map:
                    return mime_format_map[mime_type]
            except Exception as e:
                self.logger.debug(f"Ошибка определения MIME-типа: {e}")
        
        return FileFormat.UNKNOWN
    
    def parse(self, filename: str, content: bytes, 
             column_mapping: Optional[Dict[str, str]] = None,
             processing_rules: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
        """Расширенный парсинг файлов с применением правил обработки"""
        self.logger.debug(
            f"Парсинг файла: {filename}",
            context={'size': len(content), 'has_mapping': column_mapping is not None}
        )
        
        file_format = self.detect_format(filename, content)
        
        try:
            # Определяем кодировку для текстовых форматов
            if file_format in [FileFormat.CSV, FileFormat.TXT, FileFormat.XML, FileFormat.JSON]:
                encoding = self.detect_encoding(content)
            else:
                encoding = 'utf-8'
            
            # Парсим в зависимости от формата
            if file_format in [FileFormat.EXCEL_XLSX, FileFormat.EXCEL_XLS]:
                df = self._parse_excel(content)
            elif file_format == FileFormat.CSV:
                df = self._parse_csv(content, encoding)
            elif file_format == FileFormat.XML:
                df = self._parse_xml(content, encoding)
            elif file_format == FileFormat.JSON:
                df = self._parse_json(content, encoding)
            elif file_format == FileFormat.TXT:
                df = self._parse_txt(content, encoding)
            else:
                # Пробуем все форматы по очереди
                df = self._try_all_formats(content, encoding)
            
            if df is None or df.empty:
                raise ValueError(f"Не удалось распарсить файл {filename}")
            
            # Применяем правила обработки
            if processing_rules:
                df = self._apply_processing_rules(df, processing_rules)
            
            # Применяем маппинг если передан
            if column_mapping and any(column_mapping.values()):
                df = self._apply_mapping(df, column_mapping)
            else:
                df = self._clean_dataframe(df)
            
            # Валидируем результат
            validation = self.validate_data(df)
            if not validation.is_valid:
                self.logger.warning(
                    f"Проблемы валидации данных: {len(validation.errors)} ошибок, {len(validation.warnings)} предупреждений",
                    context={'errors': validation.errors[:5], 'warnings': validation.warnings[:5]}
                )
            
            self.logger.info(f"Файл {filename} успешно распарсен: {len(df)} строк")
            return df
            
        except Exception as e:
            self.logger.error(f"Ошибка парсинга {filename}: {e}")
            raise
    
    def _try_all_formats(self, content: bytes, encoding: str) -> pd.DataFrame:
        """Попытка распарсить файл всеми доступными способами"""
        parsers = [
            ('Excel', lambda: self._parse_excel(content)),
            ('CSV', lambda: self._parse_csv(content, encoding)),
            ('XML', lambda: self._parse_xml(content, encoding)),
            ('JSON', lambda: self._parse_json(content, encoding)),
            ('TXT', lambda: self._parse_txt(content, encoding))
        ]
        
        for format_name, parser_func in parsers:
            try:
                df = parser_func()
                if df is not None and not df.empty and len(df.columns) > 1:
                    self.logger.info(f"Файл успешно распознан как {format_name}")
                    return df
            except Exception as e:
                self.logger.debug(f"Не удалось распарсить как {format_name}: {e}")
                continue
        
        raise ValueError("Не удалось определить формат файла")
    
    def _parse_excel(self, content: bytes) -> pd.DataFrame:
        """Расширенный парсинг Excel файлов"""
        try:
            # Пробуем разные движки
            engines = ['openpyxl', 'xlrd', 'odf']
            
            for engine in engines:
                try:
                    if engine == 'openpyxl':
                        df = pd.read_excel(BytesIO(content), engine='openpyxl', na_filter=False)
                    elif engine == 'xlrd':
                        df = pd.read_excel(BytesIO(content), engine='xlrd', na_filter=False)
                    elif engine == 'odf':
                        df = pd.read_excel(BytesIO(content), engine='odf', na_filter=False)
                    
                    if df is not None and not df.empty:
                        return df
                except Exception:
                    continue
            
            # Пробуем читать все листы
            try:
                all_sheets = pd.read_excel(BytesIO(content), sheet_name=None, engine='openpyxl')
                if all_sheets:
                    # Берем первый непустой лист
                    for sheet_name, sheet_df in all_sheets.items():
                        if not sheet_df.empty:
                            self.logger.info(f"Прочитан лист: {sheet_name}")
                            return sheet_df
            except:
                pass
            
            raise ValueError("Не удалось прочитать Excel файл ни одним движком")
            
        except Exception as e:
            raise ValueError(f"Ошибка парсинга Excel: {e}")
    
    def _parse_csv(self, content: bytes, encoding: str) -> pd.DataFrame:
        """Расширенный парсинг CSV файлов с автоопределением разделителя"""
        separators = [';', ',', '\t', '|', ':']
        
        # Декодируем содержимое
        text_content = content.decode(encoding, errors='ignore')
        
        # Определяем разделитель
        best_sep = None
        best_score = 0
        
        for sep in separators:
            lines = text_content.split('\n')[:10]  # Анализируем первые 10 строк
            if len(lines) < 2:
                continue
            
            # Считаем количество колонок
            col_counts = [len(line.split(sep)) for line in lines if line.strip()]
            if not col_counts:
                continue
            
            # Проверяем консистентность
            most_common = max(set(col_counts), key=col_counts.count)
            consistency = col_counts.count(most_common) / len(col_counts)
            
            if consistency > best_score and most_common > 1:
                best_score = consistency
                best_sep = sep
        
        if best_sep is None:
            best_sep = ';'
        
        self.logger.debug(f"Определен разделитель CSV: '{best_sep}' (консистентность: {best_score:.2f})")
        
        try:
            df = pd.read_csv(
                BytesIO(content),
                encoding=encoding,
                sep=best_sep,
                engine='python',
                on_bad_lines='skip',
                na_filter=False,
                skip_blank_lines=True
            )
            
            if len(df.columns) > 1:
                return df
            
            # Если не получилось, пробуем другие кодировки
            for enc in self.supported_encodings:
                if enc == encoding:
                    continue
                try:
                    df = pd.read_csv(
                        BytesIO(content),
                        encoding=enc,
                        sep=best_sep,
                        engine='python',
                        on_bad_lines='skip'
                    )
                    if len(df.columns) > 1:
                        self.logger.info(f"CSV прочитан с кодировкой {enc}")
                        return df
                except:
                    continue
            
        except Exception as e:
            self.logger.debug(f"Ошибка чтения CSV с разделителем '{best_sep}': {e}")
        
        raise ValueError("Не удалось прочитать CSV файл")
    
    def _parse_xml(self, content: bytes, encoding: str) -> pd.DataFrame:
        """Расширенный парсинг XML файлов"""
        try:
            from xml.etree import ElementTree as ET
            
            text = content.decode(encoding, errors='ignore')
            root = ET.fromstring(text)
            
            # Ищем товары по различным тегам
            product_tags = ['offer', 'item', 'product', 'товар', 'good', 'position']
            
            for tag in product_tags:
                elements = root.findall(f'.//{tag}')
                if elements:
                    rows = []
                    for elem in elements:
                        row = {}
                        # Извлекаем атрибуты
                        for attr, value in elem.attrib.items():
                            row[f'@{attr}'] = value
                        # Извлекаем дочерние элементы
                        for child in elem:
                            if child.text and child.text.strip():
                                row[child.tag] = child.text.strip()
                            # Рекурсивно извлекаем вложенные элементы
                            for subchild in child:
                                if subchild.text and subchild.text.strip():
                                    row[f'{child.tag}_{subchild.tag}'] = subchild.text.strip()
                        if row:
                            rows.append(row)
                    
                    if rows:
                        return pd.DataFrame(rows)
            
            # Пробуем найти табличную структуру
            table_tags = ['table', 'data', 'list', 'rows', 'items']
            for table_tag in table_tags:
                table = root.find(f'.//{table_tag}')
                if table is not None:
                    rows = []
                    row_tags = ['row', 'tr', 'item', 'element']
                    for row_tag in row_tags:
                        row_elements = table.findall(f'.//{row_tag}')
                        if row_elements:
                            for row_elem in row_elements:
                                row = {}
                                for cell in row_elem:
                                    if cell.text and cell.text.strip():
                                        row[cell.tag] = cell.text.strip()
                                if row:
                                    rows.append(row)
                            break
                    
                    if rows:
                        return pd.DataFrame(rows)
            
            raise ValueError("Не найдены данные о товарах в XML")
            
        except ET.ParseError as e:
            raise ValueError(f"Ошибка парсинга XML: {e}")
    
    def _parse_json(self, content: bytes, encoding: str) -> pd.DataFrame:
        """Расширенный парсинг JSON файлов"""
        try:
            text = content.decode(encoding, errors='ignore')
            data = json.loads(text)
            
            # Обрабатываем разные структуры JSON
            if isinstance(data, list):
                return pd.DataFrame(data)
            elif isinstance(data, dict):
                # Ищем массив с данными
                array_keys = ['offers', 'items', 'products', 'data', 'rows', 'list', 'results', 'records']
                for key in array_keys:
                    if key in data and isinstance(data[key], list):
                        df = pd.DataFrame(data[key])
                        if not df.empty:
                            return df
                
                # Может быть вложенная структура
                for key, value in data.items():
                    if isinstance(value, dict):
                        for sub_key in array_keys:
                            if sub_key in value and isinstance(value[sub_key], list):
                                df = pd.DataFrame(value[sub_key])
                                if not df.empty:
                                    return df
                
                # Если один объект, делаем DataFrame из него
                return pd.DataFrame([data])
            
            raise ValueError("Не удалось распарсить JSON")
            
        except json.JSONDecodeError as e:
            raise ValueError(f"Ошибка парсинга JSON: {e}")
    
    def _parse_txt(self, content: bytes, encoding: str) -> pd.DataFrame:
        """Расширенный парсинг текстовых файлов"""
        try:
            text = content.decode(encoding, errors='ignore')
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            if not lines:
                raise ValueError("Файл пуст")
            
            # Определяем разделитель
            separators = ['\t', ';', ',', '|', '  ']  # Пробел последний
            
            best_sep = None
            best_score = 0
            
            for sep in separators:
                col_counts = [len(line.split(sep)) for line in lines[:20]]
                if not col_counts:
                    continue
                
                most_common = max(set(col_counts), key=col_counts.count)
                consistency = col_counts.count(most_common) / len(col_counts)
                
                if consistency > best_score and most_common > 1:
                    best_score = consistency
                    best_sep = sep
            
            if best_sep:
                return pd.read_csv(StringIO(text), sep=best_sep, engine='python')
            
            # Пробуем fixed-width формат
            df = pd.read_fwf(StringIO(text))
            if len(df.columns) > 1:
                return df
            
            # Последняя попытка - читаем как одну колонку
            return pd.DataFrame(lines, columns=['data'])
            
        except Exception as e:
            raise ValueError(f"Ошибка парсинга текстового файла: {e}")
    
    def _apply_processing_rules(self, df: pd.DataFrame, rules: Dict[str, Any]) -> pd.DataFrame:
        """Применение правил обработки к DataFrame"""
        if df.empty:
            return df
        
        # Пропускаем строки
        skip_rows = rules.get('skip_rows', 0)
        if skip_rows > 0:
            df = df.iloc[skip_rows:]
        
        # Пропускаем пустые строки
        if rules.get('skip_empty_rows', True):
            df = df.dropna(how='all')
        
        # Применяем множитель цены
        price_multiplier = rules.get('price_multiplier', 1.0)
        if price_multiplier != 1.0 and 'price' in df.columns:
            df['price'] = pd.to_numeric(df['price'], errors='coerce') * price_multiplier
        
        return df
    
    def _apply_mapping(self, df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
        """Расширенное применение маппинга колонок"""
        if df.empty:
            return df
        
        # Создаем новый DataFrame с переименованными колонками
        new_df = pd.DataFrame()
        
        for target_col, source_col in mapping.items():
            if source_col and source_col in df.columns:
                new_df[target_col] = df[source_col]
            elif source_col and source_col.strip() == '':
                new_df[target_col] = ''
            else:
                # Ищем похожую колонку
                if source_col:
                    similar_cols = [col for col in df.columns if source_col.lower() in col.lower()]
                    if similar_cols:
                        new_df[target_col] = df[similar_cols[0]]
                        self.logger.debug(f"Автоопределение: {target_col} -> {similar_cols[0]}")
                    else:
                        new_df[target_col] = ''
                else:
                    new_df[target_col] = ''
        
        # Очистка и нормализация
        return self._clean_dataframe(new_df)
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Расширенная очистка и нормализация DataFrame"""
        if df.empty:
            return df
        
        # Удаляем полностью пустые строки и колонки
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        
        # Нормализуем названия колонок
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        # Применяем глобальный маппинг для автоопределения
        column_map = {}
        for target, variants in self.config.column_mapping.items():
            for col in df.columns:
                col_clean = col.lower().strip()
                for variant in variants:
                    if variant in col_clean or col_clean in variant:
                        if col not in column_map:  # Не переопределяем уже назначенные
                            column_map[col] = target
                        break
                if col in column_map:
                    break
        
        # Переименовываем колонки
        for old, new in column_map.items():
            if old != new:
                df = df.rename(columns={old: new})
        
        # Нормализация данных
        if 'sku' in df.columns:
            df['sku'] = df['sku'].astype(str).str.strip()
            if self.config.validation_config.get('auto_fix_sku_uppercase', True):
                df['sku'] = df['sku'].str.upper()
            df = df[df['sku'] != '']
            df = df[df['sku'] != 'NAN']
            df = df[df['sku'] != 'NONE']
        
        if 'price' in df.columns:
            df['price'] = pd.to_numeric(df['price'], errors='coerce')
            if self.config.validation_config.get('auto_fix_price_rounding', True):
                df['price'] = df['price'].round(2)
            df = df[df['price'] > 0]
        
        if 'stock' in df.columns:
            df['stock'] = pd.to_numeric(df['stock'], errors='coerce').fillna(0).astype(int)
            if self.config.validation_config.get('auto_fix_stock_negative', True):
                df.loc[df['stock'] < 0, 'stock'] = 0
        
        if 'brand' in df.columns:
            df['brand'] = df['brand'].astype(str).str.strip()
        
        if 'name' in df.columns:
            df['name'] = df['name'].astype(str).str.strip()
        
        if 'category' in df.columns:
            df['category'] = df['category'].astype(str).str.strip()
        
        # Удаление дубликатов
        if 'sku' in df.columns and self.config.validation_config.get('auto_remove_duplicates', True):
            df = df.drop_duplicates(subset=['sku'], keep='first')
        
        return df
    
    def validate_data(self, df: pd.DataFrame) -> ValidationResult:
        """Расширенная валидация данных"""
        result = ValidationResult()
        
        if df.empty:
            result.add_error("DataFrame пуст")
            return result
        
        validation_config = self.config.validation_config
        total_rows = len(df)
        
        # Статистика
        result.statistics['total_rows'] = total_rows
        result.statistics['columns'] = list(df.columns)
        
        # Проверка обязательных колонок
        for field in validation_config.get('required_fields', ['sku', 'price']):
            if field not in df.columns:
                result.add_error(f"Отсутствует обязательная колонка: {field}")
        
        # Валидация SKU
        if 'sku' in df.columns:
            empty_sku = df['sku'].isna().sum() + (df['sku'] == '').sum()
            if empty_sku > 0:
                result.add_warning(f"Найдено {empty_sku} пустых артикулов")
            
            max_sku_length = validation_config.get('max_sku_length', 100)
            long_sku = (df['sku'].str.len() > max_sku_length).sum()
            if long_sku > 0:
                result.add_warning(f"Найдено {long_sku} артикулов длиннее {max_sku_length} символов")
            
            duplicate_sku = df['sku'].duplicated().sum()
            if duplicate_sku > 0:
                result.add_warning(f"Найдено {duplicate_sku} дубликатов артикулов")
            
            result.statistics['valid_skus'] = total_rows - empty_sku
        
        # Валидация цен
        if 'price' in df.columns:
            min_price = validation_config.get('min_price', 0.01)
            max_price = validation_config.get('max_price', 999999999)
            
            low_prices = (df['price'] < min_price).sum()
            if low_prices > 0:
                result.add_warning(f"Найдено {low_prices} цен ниже минимальной ({min_price})")
            
            high_prices = (df['price'] > max_price).sum()
            if high_prices > 0:
                result.add_warning(f"Найдено {high_prices} цен выше максимальной ({max_price})")
            
            result.statistics['price_range'] = {
                'min': float(df['price'].min()),
                'max': float(df['price'].max()),
                'mean': float(df['price'].mean()),
                'median': float(df['price'].median())
            }
        
        # Валидация остатков
        if 'stock' in df.columns:
            max_stock = validation_config.get('max_stock', 999999)
            high_stock = (df['stock'] > max_stock).sum()
            if high_stock > 0:
                result.add_warning(f"Найдено {high_stock} остатков выше максимального ({max_stock})")
            
            negative_stock = (df['stock'] < 0).sum()
            if negative_stock > 0:
                result.add_warning(f"Найдено {negative_stock} отрицательных остатков")
            
            result.statistics['stock_stats'] = {
                'total': int(df['stock'].sum()),
                'with_stock': int((df['stock'] > 0).sum()),
                'without_stock': int((df['stock'] == 0).sum())
            }
        
        return result
    
    def get_file_info(self, content: bytes, filename: str) -> Dict:
        """Получение информации о файле без полного парсинга"""
        info = {
            'filename': filename,
            'size': len(content),
            'size_mb': round(len(content) / 1024 / 1024, 2),
            'format': self.detect_format(filename, content).value,
            'encoding': None
        }
        
        if MAGIC_SUPPORT:
            try:
                mime = magic.Magic(mime=True)
                info['mime_type'] = mime.from_buffer(content[:1024])
            except:
                info['mime_type'] = 'unknown'
        
        try:
            info['encoding'] = self.detect_encoding(content)
        except:
            info['encoding'] = 'unknown'
        
        return info

# ===================================================================
# БЛОК 10: РАСШИРЕННЫЙ КОНСТРУКТОР МАППИНГА КОЛОНОК
# С ПОДДЕРЖКОЙ ЛОКАЛЬНОЙ ЗАГРУЗКИ ФАЙЛОВ И ЗАГРУЗКИ ИЗ ПОЧТЫ
# ===================================================================

class MappingConstructor:
    """
    Расширенный класс для интерактивного создания маппинга колонок
    с автоопределением, валидацией, историей изменений,
    поддержкой локальной загрузки файлов и загрузки из почты поставщика
    """
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.parser = PriceParser(config, logger)
        self.email_downloader = EmailDownloader(config, logger)
        self.mapping_templates: Dict[str, Dict[str, str]] = self._load_templates()
        
        # Кэш для хранения загруженных файлов
        self.uploaded_files_cache: Dict[str, Dict] = {}
        
        # Директория для временного хранения загруженных файлов
        self.upload_dir = Path(config.temp_dir) / 'mapping_uploads'
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    def _load_templates(self) -> Dict[str, Dict[str, str]]:
        """Загрузка шаблонов маппинга"""
        templates = {
            'standard': {
                'sku': 'Артикул',
                'price': 'Цена',
                'stock': 'Остаток',
                'brand': 'Бренд',
                'name': 'Название',
                'category': 'Категория',
                'description': 'Описание',
                'weight': 'Вес',
                'barcode': 'Штрихкод',
                'country': 'Страна',
                'warranty': 'Гарантия',
                'dimensions': 'Размеры',
                'min_order_qty': 'Мин. заказ'
            },
            'minimal': {
                'sku': 'Артикул',
                'price': 'Цена',
                'stock': 'Остаток'
            },
            'extended': {
                'sku': 'Артикул',
                'price': 'Цена',
                'stock': 'Остаток',
                'brand': 'Бренд',
                'name': 'Название',
                'category': 'Категория',
                'description': 'Описание',
                'weight': 'Вес',
                'barcode': 'Штрихкод',
                'country': 'Страна',
                'warranty': 'Гарантия',
                'dimensions': 'Размеры',
                'min_order_qty': 'Мин. заказ',
                'currency': 'Валюта',
                'vat': 'НДС'
            },
            'yandex_market': {
                'sku': 'offerId',
                'price': 'price',
                'stock': 'stock',
                'brand': 'vendor',
                'name': 'name',
                'category': 'category',
                'description': 'description',
                'barcode': 'barcode',
                'weight': 'weight',
                'dimensions': 'dimensions',
                'country': 'country_of_origin',
                'warranty': 'warranty_days',
                'vat': 'vat'
            },
            'ozon': {
                'sku': 'offer_id',
                'price': 'price',
                'stock': 'stock',
                'brand': 'brand',
                'name': 'name',
                'category': 'category',
                'barcode': 'barcode',
                'weight': 'weight',
                'dimensions': 'dimensions',
                'country': 'country',
                'warranty': 'warranty',
                'vat': 'vat'
            },
            'wildberries': {
                'sku': 'Артикул',
                'price': 'Цена',
                'stock': 'Остаток',
                'brand': 'Бренд',
                'name': 'Наименование',
                'category': 'Категория',
                'description': 'Описание',
                'barcode': 'Штрихкод',
                'weight': 'Вес',
                'dimensions': 'Габариты',
                'country': 'Страна',
                'vat': 'Ставка НДС'
            },
            '1c_export': {
                'sku': 'Код',
                'price': 'Цена',
                'stock': 'Остаток',
                'brand': 'Производитель',
                'name': 'Наименование',
                'category': 'Группа',
                'description': 'Описание',
                'barcode': 'Штрихкод',
                'weight': 'Вес',
                'country': 'СтранаПроисхождения',
                'warranty': 'Гарантия',
                'vat': 'СтавкаНДС',
                'currency': 'Валюта'
            }
        }
        
        # Загружаем пользовательские шаблоны из файла если есть
        templates_file = Path(self.config.cache_dir) / 'mapping_templates.json'
        if templates_file.exists():
            try:
                with open(templates_file, 'r', encoding='utf-8') as f:
                    custom_templates = json.load(f)
                    templates.update(custom_templates)
                    self.logger.info(f"Загружено {len(custom_templates)} пользовательских шаблонов")
            except Exception as e:
                self.logger.warning(f"Ошибка загрузки пользовательских шаблонов: {e}")
        
        return templates
    
    def save_templates(self) -> None:
        """Сохранение шаблонов маппинга в файл"""
        templates_file = Path(self.config.cache_dir) / 'mapping_templates.json'
        templates_file.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            # Сохраняем только пользовательские шаблоны (не системные)
            system_templates = ['standard', 'minimal', 'extended', 'yandex_market', 'ozon', 'wildberries', '1c_export']
            custom_templates = {k: v for k, v in self.mapping_templates.items() if k not in system_templates}
            
            with open(templates_file, 'w', encoding='utf-8') as f:
                json.dump(custom_templates, f, ensure_ascii=False, indent=2)
            
            self.logger.debug(f"Сохранено {len(custom_templates)} пользовательских шаблонов")
        except Exception as e:
            self.logger.error(f"Ошибка сохранения шаблонов: {e}")
    
    def save_uploaded_file(self, file_data: bytes, filename: str) -> str:
        """
        Сохранение загруженного файла во временную директорию
        
        Args:
            file_data: Бинарные данные файла
            filename: Имя файла
            
        Returns:
            Путь к сохраненному файлу
        """
        # Очищаем имя файла от недопустимых символов
        safe_filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
        
        # Добавляем временную метку для уникальности
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{safe_filename}"
        
        filepath = self.upload_dir / unique_filename
        
        try:
            with open(filepath, 'wb') as f:
                f.write(file_data)
            
            self.logger.info(f"Файл сохранен локально: {filepath}")
            
            # Сохраняем информацию в кэше
            self.uploaded_files_cache[unique_filename] = {
                'original_filename': filename,
                'filepath': str(filepath),
                'size': len(file_data),
                'uploaded_at': datetime.now().isoformat(),
                'file_format': self.parser.detect_format(filename, file_data).value
            }
            
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Ошибка сохранения файла {filename}: {e}")
            raise
    
    def load_file_from_local(self, filepath: str) -> Tuple[bytes, str]:
        """
        Загрузка файла с локального диска
        
        Args:
            filepath: Путь к файлу
            
        Returns:
            Кортеж (содержимое файла, имя файла)
        """
        filepath = Path(filepath)
        
        if not filepath.exists():
            raise FileNotFoundError(f"Файл не найден: {filepath}")
        
        if not filepath.is_file():
            raise ValueError(f"Указанный путь не является файлом: {filepath}")
        
        # Проверяем размер файла
        file_size = filepath.stat().st_size
        if file_size == 0:
            raise ValueError(f"Файл пуст: {filepath}")
        
        if file_size > 100 * 1024 * 1024:  # 100 MB
            raise ValueError(f"Файл слишком большой ({file_size / 1024 / 1024:.1f} MB): {filepath}")
        
        try:
            with open(filepath, 'rb') as f:
                content = f.read()
            
            filename = filepath.name
            
            self.logger.info(f"Файл загружен с локального диска: {filename} ({len(content)} байт)")
            
            return content, filename
            
        except Exception as e:
            self.logger.error(f"Ошибка чтения локального файла {filepath}: {e}")
            raise
    
    def load_file_from_email(self, supplier_name: str, max_emails: int = 10) -> List[Tuple[bytes, str, str]]:
        """
        Загрузка файлов из почты поставщика
        
        Args:
            supplier_name: Имя поставщика
            max_emails: Максимальное количество писем для проверки
            
        Returns:
            Список кортежей (содержимое, имя файла, тема письма)
        """
        supplier = self.config.get_supplier_by_name(supplier_name)
        
        if not supplier:
            raise ValueError(f"Поставщик '{supplier_name}' не найден в конфигурации")
        
        if not supplier.email or not supplier.email_password:
            raise ValueError(f"Для поставщика '{supplier_name}' не указаны email или пароль")
        
        # Создаем временный конфиг для подключения к почте поставщика
        temp_config = Config()
        temp_config.email_user = supplier.email
        temp_config.email_pass = supplier.email_password
        temp_config.imap_server = supplier.imap_server
        temp_config.imap_port = supplier.imap_port
        temp_config.email_search_days = self.config.email_search_days
        temp_config.email_max_emails = max_emails
        
        # Создаем временный загрузчик почты
        email_downloader = EmailDownloader(temp_config, self.logger)
        
        files = []
        
        try:
            if not email_downloader.connect():
                raise ConnectionError(f"Не удалось подключиться к почте поставщика {supplier_name}")
            
            # Ищем письма
            emails = email_downloader.search_emails()
            
            if not emails:
                self.logger.info(f"Нет писем от поставщика {supplier_name}")
                return files
            
            # Скачиваем вложения
            for email_data in emails[:max_emails]:
                attachments = email_downloader.download_attachments(email_data)
                
                for filename, content, mime_type in attachments:
                    files.append((content, filename, email_data.get('subject', '')))
                    
                    # Сохраняем локально
                    self.save_uploaded_file(content, filename)
                    
                    self.logger.info(
                        f"Файл загружен из почты поставщика {supplier_name}: "
                        f"{filename} (из письма: {email_data.get('subject', '')[:50]})"
                    )
            
        finally:
            email_downloader.disconnect()
        
        return files
    
    def load_file_from_url(self, url: str, filename: Optional[str] = None) -> Tuple[bytes, str]:
        """
        Загрузка файла по URL
        
        Args:
            url: URL файла
            filename: Имя файла (если не указано, извлекается из URL)
            
        Returns:
            Кортеж (содержимое файла, имя файла)
        """
        try:
            # Определяем имя файла из URL если не указано
            if filename is None:
                filename = url.split('/')[-1].split('?')[0]
                if not filename:
                    filename = f"downloaded_file_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Загружаем файл
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(url, headers=headers, timeout=30, stream=True)
            response.raise_for_status()
            
            # Проверяем размер
            content_length = response.headers.get('content-length')
            if content_length and int(content_length) > 100 * 1024 * 1024:
                raise ValueError(f"Файл слишком большой ({int(content_length) / 1024 / 1024:.1f} MB)")
            
            content = response.content
            
            if len(content) == 0:
                raise ValueError("Получен пустой файл")
            
            # Сохраняем локально
            self.save_uploaded_file(content, filename)
            
            self.logger.info(f"Файл загружен по URL: {filename} ({len(content)} байт)")
            
            return content, filename
            
        except requests.exceptions.RequestException as e:
            self.logger.error(f"Ошибка загрузки файла по URL {url}: {e}")
            raise
    
    def load_sample_from_supplier(self, supplier_name: str) -> Optional[Tuple[bytes, str]]:
        """
        Загрузка образца прайса из сохраненных файлов поставщика
        
        Args:
            supplier_name: Имя поставщика
            
        Returns:
            Кортеж (содержимое, имя файла) или None
        """
        # Ищем файлы в архиве
        archive_dir = Path(self.config.archive_dir)
        if archive_dir.exists():
            # Ищем файлы, связанные с поставщиком
            pattern = f"*{supplier_name}*"
            matching_files = list(archive_dir.glob(pattern))
            
            if matching_files:
                # Берем самый новый файл
                latest_file = max(matching_files, key=lambda f: f.stat().st_mtime)
                
                try:
                    content = latest_file.read_bytes()
                    filename = latest_file.name
                    
                    self.logger.info(f"Загружен образец из архива: {filename}")
                    
                    return content, filename
                except Exception as e:
                    self.logger.warning(f"Ошибка чтения файла из архива: {e}")
        
        # Ищем в директории загрузок
        uploads_dir = Path(self.config.uploads_dir)
        if uploads_dir.exists():
            matching_files = list(uploads_dir.glob(pattern))
            
            if matching_files:
                latest_file = max(matching_files, key=lambda f: f.stat().st_mtime)
                
                try:
                    content = latest_file.read_bytes()
                    filename = latest_file.name
                    
                    self.logger.info(f"Загружен образец из uploads: {filename}")
                    
                    return content, filename
                except Exception as e:
                    self.logger.warning(f"Ошибка чтения файла из uploads: {e}")
        
        return None
    
    def preview_file(self, content: bytes, filename: str) -> Tuple[pd.DataFrame, List[str], Dict]:
        """
        Расширенный предпросмотр файла с метаданными
        
        Args:
            content: Бинарное содержимое файла
            filename: Имя файла
            
        Returns:
            Кортеж (DataFrame, список колонок, метаданные)
        """
        try:
            # Получаем информацию о файле
            file_info = self.parser.get_file_info(content, filename)
            
            # Парсим файл (без маппинга, чтобы показать сырые данные)
            df = self.parser.parse(filename, content)
            
            # Определяем кодировку если возможно
            encoding = file_info.get('encoding', 'unknown')
            
            # Получаем метаданные
            metadata = {
                'file_info': file_info,
                'row_count': len(df),
                'column_count': len(df.columns),
                'columns': df.columns.tolist(),
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()},
                'sample_values': {},
                'null_counts': {},
                'unique_counts': {},
                'encoding': encoding,
                'file_size_formatted': f"{file_info.get('size', 0) / 1024:.1f} KB"
            }
            
            # Добавляем примеры значений для каждой колонки (первые 5 непустых)
            for col in df.columns:
                sample = df[col].dropna().head(5).tolist()
                metadata['sample_values'][col] = sample
                
                # Количество пустых значений
                null_count = df[col].isna().sum()
                metadata['null_counts'][col] = int(null_count)
                
                # Количество уникальных значений
                unique_count = df[col].nunique()
                metadata['unique_counts'][col] = int(unique_count)
            
            # Определяем возможные типы колонок
            column_types = {}
            for col in df.columns:
                col_data = df[col].dropna()
                if len(col_data) == 0:
                    column_types[col] = 'empty'
                    continue
                
                # Проверяем числовые типы
                numeric_data = pd.to_numeric(col_data, errors='coerce')
                numeric_ratio = numeric_data.notna().sum() / len(col_data)
                
                if numeric_ratio > 0.9:
                    if (numeric_data == numeric_data.astype(int)).sum() / len(numeric_data) > 0.9:
                        column_types[col] = 'integer'
                    else:
                        column_types[col] = 'decimal'
                else:
                    # Проверяем длину строк
                    str_data = col_data.astype(str)
                    avg_length = str_data.str.len().mean()
                    
                    if avg_length < 20:
                        column_types[col] = 'short_text'
                    elif avg_length < 100:
                        column_types[col] = 'text'
                    else:
                        column_types[col] = 'long_text'
            
            metadata['column_types'] = column_types
            
            return df, df.columns.tolist(), metadata
            
        except Exception as e:
            self.logger.error(f"Ошибка предпросмотра файла {filename}: {e}")
            raise
    
    def auto_detect_mapping(self, df: pd.DataFrame) -> Dict[str, str]:
        """
        Автоматическое определение маппинга на основе содержимого колонок
        
        Args:
            df: DataFrame с данными
            
        Returns:
            Словарь маппинга {целевое_поле: колонка_в_файле}
        """
        mapping = {}
        columns = df.columns.tolist()
        
        # Функция для проверки, является ли колонка числовой
        def get_numeric_ratio(col_data: pd.Series) -> float:
            """Возвращает долю числовых значений в колонке"""
            numeric_data = pd.to_numeric(col_data, errors='coerce')
            return numeric_data.notna().sum() / len(col_data) if len(col_data) > 0 else 0
        
        # Функция для проверки, является ли колонка целочисленной
        def is_integer_column(col_data: pd.Series) -> bool:
            """Проверяет, содержит ли колонка только целые числа"""
            numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
            if len(numeric_data) == 0:
                return False
            return (numeric_data == numeric_data.astype(int)).sum() / len(numeric_data) > 0.95
        
        # Функция для проверки уникальности
        def get_uniqueness(col_data: pd.Series) -> float:
            """Возвращает долю уникальных значений"""
            non_null = col_data.dropna()
            if len(non_null) == 0:
                return 0
            return non_null.nunique() / len(non_null)
        
        # Функция для проверки средней длины строки
        def get_avg_length(col_data: pd.Series) -> float:
            """Возвращает среднюю длину строковых значений"""
            str_data = col_data.dropna().astype(str)
            if len(str_data) == 0:
                return 0
            return str_data.str.len().mean()
        
        # Сначала ищем по названиям колонок
        for target, variants in self.config.column_mapping.items():
            best_match = None
            best_score = 0
            
            for col in columns:
                col_lower = col.lower().strip()
                score = 0
                
                for variant in variants:
                    variant_lower = variant.lower()
                    
                    # Точное совпадение
                    if col_lower == variant_lower:
                        score = 100
                        break
                    
                    # Содержит вариант
                    if variant_lower in col_lower:
                        score = max(score, 85)
                    
                    # Вариант содержит название колонки
                    if col_lower in variant_lower:
                        score = max(score, 75)
                    
                    # Частичное совпадение (по словам)
                    col_words = set(col_lower.replace('_', ' ').split())
                    variant_words = set(variant_lower.replace('_', ' ').split())
                    common_words = col_words & variant_words
                    if common_words:
                        score = max(score, 60 + len(common_words) * 10)
                
                if score > best_score and score > 50:
                    best_score = score
                    best_match = col
            
            if best_match:
                mapping[target] = best_match
        
        # Для ненайденных колонок пробуем определить по содержимому
        unmapped_columns = [col for col in columns if col not in mapping.values()]
        
        for col in unmapped_columns:
            col_data = df[col]
            numeric_ratio = get_numeric_ratio(col_data)
            uniqueness = get_uniqueness(col_data)
            avg_length = get_avg_length(col_data)
            
            # Определяем цену
            if 'price' not in mapping and numeric_ratio > 0.8:
                numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
                if len(numeric_data) > 0:
                    mean_val = numeric_data.mean()
                    # Цены обычно положительные и не слишком большие
                    if 0 < mean_val < 10000000:
                        # Проверяем, есть ли десятичные знаки
                        has_decimals = any('.' in str(x) for x in numeric_data.head(20))
                        if has_decimals or not is_integer_column(col_data):
                            mapping['price'] = col
                            continue
            
            # Определяем остаток
            if 'stock' not in mapping and numeric_ratio > 0.8:
                numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
                if len(numeric_data) > 0:
                    mean_val = numeric_data.mean()
                    # Остатки обычно целые и не слишком большие
                    if 0 <= mean_val < 100000 and is_integer_column(col_data):
                        mapping['stock'] = col
                        continue
            
            # Определяем артикул
            if 'sku' not in mapping and uniqueness > 0.5 and avg_length < 30:
                # Артикулы обычно короткие и уникальные
                if uniqueness > 0.7 or (uniqueness > 0.5 and avg_length < 15):
                    mapping['sku'] = col
                    continue
            
            # Определяем название
            if 'name' not in mapping and avg_length > 20 and uniqueness > 0.5:
                mapping['name'] = col
                continue
            
            # Определяем описание
            if 'description' not in mapping and avg_length > 50:
                mapping['description'] = col
                continue
            
            # Определяем вес
            if 'weight' not in mapping and numeric_ratio > 0.8:
                numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
                if len(numeric_data) > 0:
                    mean_val = numeric_data.mean()
                    # Вес обычно положительный и небольшой
                    if 0 < mean_val < 1000:
                        mapping['weight'] = col
                        continue
        
        return mapping
    
    def create_mapping(self, df: pd.DataFrame, selected_columns: Dict[str, str],
                      template_name: Optional[str] = None) -> Dict[str, str]:
        """
        Создание маппинга с возможностью использования шаблонов
        
        Args:
            df: DataFrame с данными
            selected_columns: Словарь выбранных пользователем колонок
            template_name: Имя шаблона (опционально)
            
        Returns:
            Словарь маппинга
        """
        # Если указан шаблон, применяем его как основу
        if template_name and template_name in self.mapping_templates:
            template = self.mapping_templates[template_name].copy()
            # Объединяем шаблон с выбранными колонками (выбранные имеют приоритет)
            mapping = template.copy()
            for target, source in selected_columns.items():
                if source and source in df.columns:
                    mapping[target] = source
                elif source == '':
                    # Явно удаляем поле из маппинга если выбрана пустая строка
                    if target in mapping:
                        del mapping[target]
        else:
            # Создаем новый маппинг из выбранных колонок
            mapping = {}
            for target, source in selected_columns.items():
                if source and source in df.columns:
                    mapping[target] = source
        
        return mapping
    
    def validate_mapping(self, df: pd.DataFrame, mapping: Dict[str, str]) -> ValidationResult:
        """
        Расширенная валидация маппинга
        
        Args:
            df: DataFrame с данными
            mapping: Словарь маппинга
            
        Returns:
            ValidationResult с результатами валидации
        """
        result = ValidationResult()
        
        # Проверяем обязательные поля
        required_fields = ['sku', 'price']
        for field in required_fields:
            if not mapping.get(field):
                result.add_error(f"Не указан маппинг для обязательного поля: {field}")
        
        # Проверяем существование колонок
        for target, source in mapping.items():
            if source and source not in df.columns:
                result.add_error(f"Колонка '{source}' (для поля '{target}') не найдена в файле")
        
        if not result.is_valid:
            return result
        
        # Применяем маппинг и проверяем результат
        try:
            mapped_df = self.parser._apply_mapping(df.copy(), mapping)
            
            if mapped_df.empty:
                result.add_error("После применения маппинга DataFrame пуст")
                return result
            
            # Проверяем качество данных
            if 'sku' in mapped_df.columns:
                total_rows = len(mapped_df)
                empty_sku = (mapped_df['sku'].isna() | (mapped_df['sku'] == '')).sum()
                duplicate_sku = mapped_df['sku'].duplicated().sum()
                
                if empty_sku > 0:
                    result.add_warning(f"Найдено {empty_sku} пустых артикулов ({empty_sku / total_rows * 100:.1f}%)")
                
                if duplicate_sku > 0:
                    result.add_warning(f"Найдено {duplicate_sku} дубликатов артикулов")
                
                result.statistics['total_rows'] = total_rows
                result.statistics['valid_skus'] = total_rows - empty_sku
                result.statistics['unique_skus'] = mapped_df['sku'].nunique()
            
            if 'price' in mapped_df.columns:
                total_rows = len(mapped_df)
                zero_prices = (mapped_df['price'] == 0).sum()
                invalid_prices = mapped_df['price'].isna().sum()
                negative_prices = (mapped_df['price'] < 0).sum()
                
                if invalid_prices > 0:
                    result.add_error(f"Найдено {invalid_prices} некорректных цен")
                
                if zero_prices > 0:
                    result.add_warning(f"Найдено {zero_prices} нулевых цен ({zero_prices / total_rows * 100:.1f}%)")
                
                if negative_prices > 0:
                    result.add_warning(f"Найдено {negative_prices} отрицательных цен")
                
                valid_prices = mapped_df[mapped_df['price'] > 0]['price']
                if len(valid_prices) > 0:
                    result.statistics['price_range'] = {
                        'min': float(valid_prices.min()),
                        'max': float(valid_prices.max()),
                        'mean': float(valid_prices.mean()),
                        'median': float(valid_prices.median())
                    }
            
            if 'stock' in mapped_df.columns:
                negative_stock = (mapped_df['stock'] < 0).sum()
                if negative_stock > 0:
                    result.add_warning(f"Найдено {negative_stock} отрицательных остатков")
                
                total_stock = int(mapped_df['stock'].sum())
                with_stock = int((mapped_df['stock'] > 0).sum())
                result.statistics['stock_stats'] = {
                    'total': total_stock,
                    'with_stock': with_stock,
                    'without_stock': len(mapped_df) - with_stock
                }
            
            result.statistics['mapped_fields'] = list(mapping.keys())
            result.statistics['valid_rows'] = len(mapped_df)
            result.statistics['total_columns_in_file'] = len(df.columns)
            result.statistics['mapped_columns_count'] = len(mapping)
            
        except Exception as e:
            result.add_error(f"Ошибка применения маппинга: {str(e)}")
        
        return result
    
    def save_mapping(self, supplier_name: str, mapping: Dict[str, str],
                    created_by: str = 'user', comment: str = '') -> bool:
        """
        Сохранение маппинга для поставщика с историей
        
        Args:
            supplier_name: Имя поставщика
            mapping: Словарь маппинга
            created_by: Кто создал (user, auto, template, import)
            comment: Комментарий к изменению
            
        Returns:
            True если сохранение успешно
        """
        return self.config.set_supplier_mapping(supplier_name, mapping, created_by, comment)
    
    def get_mapping_suggestions(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Получение предложений по маппингу с оценкой уверенности
        
        Args:
            df: DataFrame с данными
            
        Returns:
            Список предложений с оценками уверенности
        """
        suggestions = []
        
        for target, variants in self.config.column_mapping.items():
            best_match = None
            best_score = 0
            best_reasons = []
            
            for col in df.columns:
                col_lower = col.lower().strip()
                col_data = df[col].dropna()
                
                if len(col_data) == 0:
                    continue
                
                score = 0
                reasons = []
                
                # Проверка по названию колонки
                for variant in variants:
                    variant_lower = variant.lower()
                    
                    if variant_lower == col_lower:
                        score += 50
                        reasons.append(f'Точное совпадение названия: "{variant}"')
                        break
                    elif variant_lower in col_lower:
                        score += 40
                        reasons.append(f'Содержит: "{variant}"')
                    elif col_lower in variant_lower:
                        score += 35
                        reasons.append(f'Часть от: "{variant}"')
                    
                    # Проверка по словам
                    col_words = set(col_lower.replace('_', ' ').replace('-', ' ').split())
                    variant_words = set(variant_lower.replace('_', ' ').replace('-', ' ').split())
                    common = col_words & variant_words
                    if common:
                        score += len(common) * 15
                        reasons.append(f'Общие слова: {", ".join(common)}')
                
                # Проверка по содержимому колонки
                numeric_ratio = 0
                try:
                    numeric_data = pd.to_numeric(col_data, errors='coerce')
                    numeric_ratio = numeric_data.notna().sum() / len(col_data)
                except:
                    pass
                
                uniqueness = col_data.nunique() / len(col_data) if len(col_data) > 0 else 0
                avg_length = col_data.astype(str).str.len().mean() if len(col_data) > 0 else 0
                
                if target == 'price' and numeric_ratio > 0.8:
                    mean_val = pd.to_numeric(col_data, errors='coerce').mean()
                    if 0 < mean_val < 10000000:
                        score += 30
                        reasons.append('Похоже на цены (числовые значения)')
                
                elif target == 'stock' and numeric_ratio > 0.8:
                    numeric_data = pd.to_numeric(col_data, errors='coerce').dropna()
                    if len(numeric_data) > 0:
                        if (numeric_data == numeric_data.astype(int)).mean() > 0.9:
                            mean_val = numeric_data.mean()
                            if mean_val < 100000:
                                score += 25
                                reasons.append('Похоже на остатки (целые числа)')
                
                elif target == 'sku' and uniqueness > 0.5 and avg_length < 30:
                    score += 20
                    reasons.append(f'Похоже на артикулы (уникальность: {uniqueness:.1%})')
                
                elif target == 'name' and avg_length > 20 and uniqueness > 0.5:
                    score += 15
                    reasons.append(f'Похоже на названия (средняя длина: {avg_length:.0f})')
                
                elif target == 'description' and avg_length > 50:
                    score += 15
                    reasons.append(f'Похоже на описания (средняя длина: {avg_length:.0f})')
                
                elif target == 'weight' and numeric_ratio > 0.8:
                    mean_val = pd.to_numeric(col_data, errors='coerce').mean()
                    if 0 < mean_val < 1000:
                        score += 15
                        reasons.append('Похоже на вес (небольшие числа)')
                
                if score > best_score:
                    best_score = score
                    best_match = col
                    best_reasons = reasons
            
            if best_match and best_score >= 30:
                suggestions.append({
                    'target': target,
                    'column': best_match,
                    'confidence': min(best_score, 100),
                    'reasons': best_reasons,
                    'confidence_level': (
                        'high' if best_score >= 70 else
                        'medium' if best_score >= 50 else
                        'low'
                    )
                })
        
        # Сортируем по уверенности
        return sorted(suggestions, key=lambda x: x['confidence'], reverse=True)
    
    def compare_mappings(self, mapping1: Dict[str, str], mapping2: Dict[str, str]) -> Dict:
        """
        Сравнение двух маппингов
        
        Args:
            mapping1: Первый маппинг
            mapping2: Второй маппинг
            
        Returns:
            Словарь с информацией об изменениях
        """
        all_keys = set(list(mapping1.keys()) + list(mapping2.keys()))
        
        changes = []
        for key in sorted(all_keys):
            old_value = mapping1.get(key, '')
            new_value = mapping2.get(key, '')
            
            if old_value != new_value:
                change_type = 'modified'
                if old_value == '' and new_value != '':
                    change_type = 'added'
                elif old_value != '' and new_value == '':
                    change_type = 'removed'
                
                changes.append({
                    'field': key,
                    'old_value': old_value or '(не задано)',
                    'new_value': new_value or '(не задано)',
                    'change_type': change_type
                })
        
        return {
            'changes': changes,
            'total_changes': len(changes),
            'added_fields': sum(1 for c in changes if c['change_type'] == 'added'),
            'removed_fields': sum(1 for c in changes if c['change_type'] == 'removed'),
            'modified_fields': sum(1 for c in changes if c['change_type'] == 'modified'),
            'has_changes': len(changes) > 0
        }
    
    def export_mapping_config(self, supplier_name: str, format: str = 'json') -> str:
        """
        Экспорт конфигурации маппинга в файл
        
        Args:
            supplier_name: Имя поставщика
            format: Формат экспорта ('json' или 'csv')
            
        Returns:
            Путь к экспортированному файлу
        """
        mapping = self.config.get_supplier_mapping(supplier_name)
        if not mapping:
            return ''
        
        export_dir = Path(self.config.exports_dir) / 'mappings'
        export_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'json':
            filename = f"mapping_{supplier_name}_{timestamp}.json"
            filepath = export_dir / filename
            
            export_data = {
                'supplier': supplier_name,
                'mapping': mapping,
                'exported_at': datetime.now().isoformat(),
                'version': self.config.config_version,
                'template_compatible': True
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            return str(filepath)
        
        elif format == 'csv':
            filename = f"mapping_{supplier_name}_{timestamp}.csv"
            filepath = export_dir / filename
            
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Поле', 'Колонка в файле', 'Описание'])
                
                field_descriptions = {
                    'sku': 'Артикул товара',
                    'price': 'Цена',
                    'stock': 'Остаток',
                    'brand': 'Бренд',
                    'name': 'Название',
                    'category': 'Категория',
                    'description': 'Описание',
                    'weight': 'Вес',
                    'barcode': 'Штрихкод',
                    'country': 'Страна производства',
                    'warranty': 'Гарантия',
                    'dimensions': 'Размеры',
                    'min_order_qty': 'Минимальный заказ',
                    'currency': 'Валюта',
                    'vat': 'НДС'
                }
                
                for target, source in mapping.items():
                    description = field_descriptions.get(target, '')
                    writer.writerow([target, source, description])
            
            return str(filepath)
        
        return ''
    
    def import_mapping_config(self, filepath: str) -> Optional[Dict[str, str]]:
        """
        Импорт конфигурации маппинга из файла
        
        Args:
            filepath: Путь к файлу
            
        Returns:
            Словарь маппинга или None
        """
        try:
            filepath = Path(filepath)
            
            if not filepath.exists():
                self.logger.error(f"Файл не найден: {filepath}")
                return None
            
            if filepath.suffix == '.json':
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                    # Поддерживаем разные форматы JSON
                    if 'mapping' in data:
                        return data['mapping']
                    elif isinstance(data, dict):
                        # Может быть сам маппинг
                        if all(isinstance(v, str) for v in data.values()):
                            return data
            
            elif filepath.suffix == '.csv':
                mapping = {}
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        field = row.get('Поле', row.get('field', row.get('target', '')))
                        column = row.get('Колонка в файле', row.get('column', row.get('source', '')))
                        if field and column:
                            mapping[field] = column
                
                if mapping:
                    return mapping
            
            self.logger.error(f"Не удалось импортировать маппинг из файла: {filepath}")
            return None
            
        except Exception as e:
            self.logger.error(f"Ошибка импорта маппинга: {e}")
            return None
    
    def get_available_suppliers_for_email(self) -> List[str]:
        """
        Получение списка поставщиков, у которых настроена почта
        
        Returns:
            Список имен поставщиков
        """
        suppliers = []
        for supplier_data in self.config.suppliers:
            if supplier_data.get('email') and supplier_data.get('email_password'):
                suppliers.append(supplier_data.get('name', 'Unknown'))
        return suppliers
    
    def get_uploaded_files_list(self) -> List[Dict]:
        """
        Получение списка загруженных файлов
        
        Returns:
            Список словарей с информацией о файлах
        """
        files = []
        
        if self.upload_dir.exists():
            for filepath in sorted(self.upload_dir.iterdir(), key=lambda f: f.stat().st_mtime, reverse=True):
                if filepath.is_file():
                    files.append({
                        'filename': filepath.name,
                        'size': filepath.stat().st_size,
                        'size_formatted': f"{filepath.stat().st_size / 1024:.1f} KB",
                        'uploaded_at': datetime.fromtimestamp(filepath.stat().st_mtime).isoformat(),
                        'filepath': str(filepath)
                    })
        
        return files
    
    def cleanup_old_uploads(self, max_age_hours: int = 24) -> int:
        """
        Очистка старых загруженных файлов
        
        Args:
            max_age_hours: Максимальный возраст файла в часах
            
        Returns:
            Количество удаленных файлов
        """
        cleaned = 0
        cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
        
        if self.upload_dir.exists():
            for filepath in self.upload_dir.iterdir():
                if filepath.is_file():
                    mtime = datetime.fromtimestamp(filepath.stat().st_mtime)
                    if mtime < cutoff_time:
                        try:
                            filepath.unlink()
                            cleaned += 1
                        except Exception as e:
                            self.logger.warning(f"Не удалось удалить файл {filepath}: {e}")
        
        # Очищаем кэш
        self.uploaded_files_cache.clear()
        
        if cleaned > 0:
            self.logger.info(f"Очищено {cleaned} старых загруженных файлов")
        
        return cleaned
    
    def get_mapping_statistics(self, supplier_name: str) -> Dict:
        """
        Получение статистики маппинга для поставщика
        
        Args:
            supplier_name: Имя поставщика
            
        Returns:
            Словарь со статистикой
        """
        mapping = self.config.get_supplier_mapping(supplier_name)
        history = self.config.get_mapping_history(supplier_name)
        
        stats = {
            'has_mapping': mapping is not None and any(mapping.values()),
            'mapped_fields_count': sum(1 for v in mapping.values() if v) if mapping else 0,
            'total_possible_fields': len(self.config.column_mapping),
            'history_count': len(history),
            'last_updated': history[0].get('created_at', '') if history else None,
            'required_fields_mapped': False
        }
        
        if mapping:
            required = ['sku', 'price']
            stats['required_fields_mapped'] = all(mapping.get(f) for f in required)
        
        return stats


# ===================================================================
# БЛОК 11: РАСШИРЕННЫЙ МНОГОПОСТАВЩИЧНЫЙ ЗАГРУЗЧИК
# ===================================================================

class MultiSupplierDownloader:
    """Расширенный класс для загрузки прайсов от всех поставщиков"""
    
    ALLOWED_EXTENSIONS = ['.xls', '.xlsx', '.csv', '.xml', '.json', '.txt', '.xlsm', '.xlsb', '.ods']
    ARCHIVE_EXTENSIONS = ['.zip', '.tar', '.gz', '.tgz', '.bz2', '.7z', '.rar']
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.suppliers = config.get_active_suppliers()
        self.temp_dir = Path(config.temp_dir) / 'supplier_downloads'
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.download_stats: Dict[str, Dict] = {}
        self.max_concurrent_downloads = config.max_workers
        self.download_timeout = 120  # секунд на одного поставщика
    
    def download_from_supplier(self, supplier: SupplierConfig) -> Dict[str, Any]:
        """Расширенная загрузка от одного поставщика с детальной статистикой"""
        supplier_name = supplier.name
        start_time = time.time()
        
        results = {
            'supplier': supplier_name,
            'status': 'pending',
            'files_downloaded': 0,
            'files_extracted': 0,
            'files_skipped': 0,
            'total_size': 0,
            'errors': [],
            'warnings': [],
            'duration': 0,
            'downloaded_files': []
        }
        
        try:
            if not supplier.email or not supplier.email_password:
                results['status'] = 'no_credentials'
                results['warnings'].append('Не указаны email или пароль')
                return results
            
            self.logger.info(f"Загрузка прайсов от поставщика: {supplier_name} ({supplier.email})")
            
            # Подключаемся к почте поставщика
            try:
                if supplier.imap_port == 993:
                    mail = imaplib.IMAP4_SSL(supplier.imap_server, supplier.imap_port, timeout=60)
                else:
                    mail = imaplib.IMAP4(supplier.imap_server, supplier.imap_port, timeout=60)
                
                mail.login(supplier.email, supplier.email_password)
                mail.select('INBOX')
            except Exception as e:
                results['status'] = 'connection_error'
                results['errors'].append(f"Ошибка подключения: {str(e)}")
                return results
            
            try:
                # Формируем критерии поиска
                criteria = []
                
                if supplier.sender_filter:
                    criteria.append(f'FROM "{supplier.sender_filter}"')
                
                if supplier.subject_filter:
                    criteria.append(f'SUBJECT "{supplier.subject_filter}"')
                
                since_date = (datetime.now() - timedelta(days=self.config.email_search_days)).strftime("%d-%b-%Y")
                criteria.append(f'SINCE {since_date}')
                criteria.append('UNSEEN')
                
                search_criteria = ' '.join(criteria) if criteria else 'ALL'
                
                # Ищем письма
                status, messages = mail.search(None, search_criteria)
                
                if status != 'OK' or not messages[0]:
                    self.logger.info(f"Поставщик {supplier_name}: писем не найдено")
                    results['status'] = 'no_emails'
                    return results
                
                email_ids = messages[0].split()
                
                if len(email_ids) > self.config.email_max_emails:
                    email_ids = email_ids[-self.config.email_max_emails:]
                
                self.logger.info(f"Поставщик {supplier_name}: найдено {len(email_ids)} писем")
                
                # Обрабатываем каждое письмо
                for msg_id in email_ids:
                    try:
                        status, msg_data = mail.fetch(msg_id, '(RFC822)')
                        if status != 'OK':
                            continue
                        
                        msg = email.message_from_bytes(msg_data[0][1], policy=email_default_policy)
                        
                        # Получаем информацию о письме
                        subject = msg.get('Subject', '')
                        from_addr = msg.get('From', '')
                        date_str = msg.get('Date', '')
                        
                        # Обрабатываем вложения
                        for part in msg.walk():
                            content_disposition = str(part.get('Content-Disposition', ''))
                            
                            if 'attachment' in content_disposition or part.get_filename():
                                filename = part.get_filename()
                                if not filename:
                                    continue
                                
                                # Декодируем имя файла
                                try:
                                    decoded_parts = decode_header(filename)
                                    filename = ''.join(
                                        str(p[0] if isinstance(p[0], str) else p[0].decode('utf-8', errors='ignore'))
                                        for p in decoded_parts
                                    )
                                except:
                                    pass
                                
                                # Очищаем имя файла
                                filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                                
                                content = part.get_payload(decode=True)
                                if not content:
                                    continue
                                
                                ext = Path(filename).suffix.lower()
                                
                                # Проверяем, не архив ли это
                                if ext in self.ARCHIVE_EXTENSIONS:
                                    if ext == '.rar' and not RAR_SUPPORT:
                                        self.logger.warning(f"RAR не поддерживается: {filename}")
                                        results['warnings'].append(f"RAR не поддерживается: {filename}")
                                        continue
                                    
                                    self.logger.info(f"Поставщик {supplier_name}: распаковка архива {filename}")
                                    extracted = self._extract_archive(filename, content, supplier_name)
                                    
                                    for extracted_file in extracted:
                                        results['downloaded_files'].append(extracted_file)
                                        results['files_extracted'] += 1
                                        results['total_size'] += extracted_file.get('size', 0)
                                    
                                elif ext in self.ALLOWED_EXTENSIONS:
                                    file_info = {
                                        'supplier': supplier_name,
                                        'supplier_email': supplier.email,
                                        'filename': filename,
                                        'content': content,
                                        'size': len(content),
                                        'downloaded_at': datetime.now().isoformat(),
                                        'subject': subject,
                                        'from': from_addr,
                                        'date': date_str
                                    }
                                    
                                    results['downloaded_files'].append(file_info)
                                    results['files_downloaded'] += 1
                                    results['total_size'] += len(content)
                                    
                                    self.logger.info(
                                        f"Поставщик {supplier_name}: скачан файл {filename} "
                                        f"({len(content) / 1024:.1f} KB)"
                                    )
                                else:
                                    results['files_skipped'] += 1
                                    self.logger.debug(f"Пропущен файл {filename} (расширение {ext})")
                        
                    except Exception as e:
                        error_msg = f"Ошибка обработки письма: {str(e)}"
                        results['errors'].append(error_msg)
                        self.logger.error(f"Поставщик {supplier_name}: {error_msg}")
                
                # Помечаем письма как прочитанные
                if self.config.email_mark_as_read:
                    for msg_id in email_ids:
                        try:
                            mail.store(msg_id, '+FLAGS', '\\Seen')
                        except:
                            pass
                
            finally:
                try:
                    mail.close()
                    mail.logout()
                except:
                    pass
            
            # Обновляем статус
            if results['errors']:
                results['status'] = 'partial_success' if results['files_downloaded'] > 0 else 'failed'
            else:
                results['status'] = 'success'
            
            # Обновляем статистику поставщика
            supplier.update_stats(
                success=results['status'] in ['success', 'partial_success'],
                products_count=results['files_downloaded'] + results['files_extracted']
            )
            
            # Сохраняем обновленную статистику
            self.config.set_supplier_mapping(supplier_name, supplier.column_mapping)
            
        except Exception as e:
            results['status'] = 'failed'
            results['errors'].append(f"Критическая ошибка: {str(e)}")
            self.logger.error(f"Поставщик {supplier_name}: критическая ошибка: {e}")
        
        results['duration'] = time.time() - start_time
        self.download_stats[supplier_name] = results
        
        return results
    
    def _extract_archive(self, filename: str, content: bytes, 
                        supplier_name: str) -> List[Dict]:
        """Расширенная распаковка архивов с поддержкой вложенных архивов"""
        results = []
        
        try:
            # Создаем временную директорию для распаковки
            extract_id = f"{supplier_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
            extract_dir = self.temp_dir / extract_id
            extract_dir.mkdir(parents=True, exist_ok=True)
            
            temp_archive = self.temp_dir / f"temp_{extract_id}{Path(filename).suffix}"
            
            try:
                # Сохраняем архив во временный файл
                with open(temp_archive, 'wb') as f:
                    f.write(content)
                
                ext = Path(filename).suffix.lower()
                
                # Распаковываем в зависимости от типа архива
                if ext == '.zip':
                    with zipfile.ZipFile(temp_archive, 'r') as zip_ref:
                        # Проверяем на наличие вредоносных путей
                        for member in zip_ref.namelist():
                            if member.startswith('/') or '..' in member:
                                self.logger.warning(f"Подозрительный путь в архиве: {member}")
                                continue
                        zip_ref.extractall(extract_dir)
                
                elif ext in ['.tar', '.tgz', '.bz2']:
                    mode = 'r:*'
                    if ext == '.bz2':
                        mode = 'r:bz2'
                    elif ext == '.tgz':
                        mode = 'r:gz'
                    
                    with tarfile.open(temp_archive, mode) as tar_ref:
                        # Проверяем пути
                        for member in tar_ref.getmembers():
                            if member.name.startswith('/') or '..' in member.name:
                                self.logger.warning(f"Подозрительный путь в архиве: {member.name}")
                                continue
                        tar_ref.extractall(extract_dir)
                
                elif ext == '.gz' and not filename.endswith('.tgz'):
                    # Одиночный gzip файл
                    output_name = Path(filename).stem
                    output_path = extract_dir / output_name
                    
                    with gzip.open(temp_archive, 'rb') as f_in:
                        with open(output_path, 'wb') as f_out:
                            shutil.copyfileobj(f_in, f_out)
                
                elif ext == '.7z':
                    # Требуется библиотека py7zr
                    try:
                        import py7zr
                        with py7zr.SevenZipFile(temp_archive, mode='r') as z:
                            z.extractall(path=extract_dir)
                    except ImportError:
                        self.logger.warning("Библиотека py7zr не установлена, пропускаем .7z архив")
                
                elif ext == '.rar' and RAR_SUPPORT:
                    with rarfile.RarFile(temp_archive, 'r') as rf:
                        rf.extractall(extract_dir)
                
                else:
                    self.logger.warning(f"Неподдерживаемый формат архива: {ext}")
                    return results
                
                # Рекурсивно обрабатываем извлеченные файлы
                for file_path in extract_dir.rglob('*'):
                    if file_path.is_file():
                        file_ext = file_path.suffix.lower()
                        
                        # Проверяем, не архив ли внутри архива
                        if file_ext in self.ARCHIVE_EXTENSIONS:
                            with open(file_path, 'rb') as f:
                                nested_content = f.read()
                            nested_results = self._extract_archive(
                                file_path.name, nested_content, supplier_name
                            )
                            results.extend(nested_results)
                        
                        elif file_ext in self.ALLOWED_EXTENSIONS:
                            with open(file_path, 'rb') as f:
                                file_content = f.read()
                            
                            file_info = {
                                'supplier': supplier_name,
                                'supplier_email': '',
                                'filename': file_path.name,
                                'content': file_content,
                                'size': len(file_content),
                                'downloaded_at': datetime.now().isoformat(),
                                'archive_source': filename
                            }
                            
                            results.append(file_info)
                            self.logger.info(
                                f"Поставщик {supplier_name}: извлечен файл {file_path.name} "
                                f"из {filename} ({len(file_content) / 1024:.1f} KB)"
                            )
                
            finally:
                # Удаляем временный архив
                if temp_archive.exists():
                    temp_archive.unlink()
                
                # Удаляем временную директорию
                shutil.rmtree(extract_dir, ignore_errors=True)
            
        except Exception as e:
            self.logger.error(f"Ошибка распаковки архива {filename}: {e}")
        
        return results
    
    def download_all_suppliers(self) -> List[Dict]:
        """Расширенная загрузка от всех поставщиков с параллельной обработкой"""
        all_results = []
        
        if not self.suppliers:
            self.logger.warning("Список поставщиков пуст")
            return all_results
        
        self.logger.info(f"Загрузка прайсов от {len(self.suppliers)} поставщиков...")
        start_time = time.time()
        
        # Используем ThreadPoolExecutor для параллельной загрузки
        with ThreadPoolExecutor(max_workers=self.max_concurrent_downloads) as executor:
            # Создаем задачи для каждого поставщика
            future_to_supplier = {}
            
            for supplier in self.suppliers:
                future = executor.submit(self.download_from_supplier, supplier)
                future_to_supplier[future] = supplier.name
            
            # Собираем результаты по мере выполнения
            for future in as_completed(future_to_supplier):
                supplier_name = future_to_supplier[future]
                
                try:
                    result = future.result(timeout=self.download_timeout)
                    
                    # Добавляем скачанные файлы в общий список
                    if result['status'] in ['success', 'partial_success']:
                        all_results.extend(result.get('downloaded_files', []))
                    
                    self.logger.info(
                        f"Поставщик {supplier_name}: загружено {result['files_downloaded']} "
                        f"файлов, извлечено {result['files_extracted']}, "
                        f"ошибок: {len(result['errors'])}"
                    )
                    
                except TimeoutError:
                    self.logger.error(f"Поставщик {supplier_name}: превышено время ожидания")
                    self.download_stats[supplier_name] = {
                        'supplier': supplier_name,
                        'status': 'timeout',
                        'error': 'Превышено время ожидания'
                    }
                except Exception as e:
                    self.logger.error(f"Поставщик {supplier_name}: ошибка загрузки: {e}")
                    self.download_stats[supplier_name] = {
                        'supplier': supplier_name,
                        'status': 'failed',
                        'error': str(e)
                    }
        
        total_duration = time.time() - start_time
        
        # Формируем сводную статистику
        total_files = len(all_results)
        total_size = sum(f.get('size', 0) for f in all_results)
        
        self.logger.info(
            f"Загрузка завершена за {total_duration:.1f} сек. "
            f"Всего файлов: {total_files}, "
            f"общий размер: {total_size / 1024 / 1024:.1f} MB"
        )
        
        return all_results
    
    def get_download_stats(self) -> pd.DataFrame:
        """Получение статистики загрузки в виде DataFrame"""
        if not self.download_stats:
            return pd.DataFrame()
        
        stats_list = []
        for supplier_name, stats in self.download_stats.items():
            stats_list.append({
                'Поставщик': supplier_name,
                'Статус': stats.get('status', 'unknown'),
                'Файлов скачано': stats.get('files_downloaded', 0),
                'Файлов извлечено': stats.get('files_extracted', 0),
                'Пропущено': stats.get('files_skipped', 0),
                'Ошибок': len(stats.get('errors', [])),
                'Размер (MB)': round(stats.get('total_size', 0) / 1024 / 1024, 2),
                'Время (сек)': round(stats.get('duration', 0), 1)
            })
        
        return pd.DataFrame(stats_list)
    
    def retry_failed_downloads(self) -> List[Dict]:
        """Повторная загрузка для поставщиков с ошибками"""
        failed_suppliers = []
        
        for supplier_name, stats in self.download_stats.items():
            if stats.get('status') in ['failed', 'timeout', 'connection_error']:
                supplier = self.config.get_supplier_by_name(supplier_name)
                if supplier:
                    failed_suppliers.append(supplier)
        
        if not failed_suppliers:
            self.logger.info("Нет поставщиков с ошибками для повторной загрузки")
            return []
        
        self.logger.info(f"Повторная загрузка для {len(failed_suppliers)} поставщиков")
        
        all_results = []
        for supplier in failed_suppliers:
            result = self.download_from_supplier(supplier)
            if result['status'] in ['success', 'partial_success']:
                all_results.extend(result.get('downloaded_files', []))
        
        return all_results


# ===================================================================
# БЛОК 12: РАСШИРЕННЫЙ АНАЛИЗАТОР ПРАЙСОВ
# ===================================================================

class PriceAnalyzer:
    """Расширенный класс для анализа и сравнения прайсов от разных поставщиков"""
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.analysis_dir = Path(config.analysis_dir)
        self.analysis_dir.mkdir(parents=True, exist_ok=True)
        self.current_analysis_id: Optional[str] = None
    
    def analyze(self, parsed_files: Dict[str, pd.DataFrame], 
               analysis_config: Optional[Dict] = None) -> pd.DataFrame:
        """Расширенный анализ прайсов с множеством метрик"""
        if not parsed_files:
            self.logger.warning("Нет данных для анализа")
            return pd.DataFrame()
        
        self.current_analysis_id = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        self.logger.info(
            f"Анализ {len(parsed_files)} прайсов...",
            context={'analysis_id': self.current_analysis_id}
        )
        
        # Объединяем все данные
        all_data = []
        file_stats = {}
        
        for filename, df in parsed_files.items():
            if df.empty:
                continue
            
            # Получаем информацию о поставщике
            supplier = df['supplier'].iloc[0] if 'supplier' in df.columns else 'Unknown'
            supplier_email = df['supplier_email'].iloc[0] if 'supplier_email' in df.columns else ''
            
            # Сохраняем статистику по файлу
            file_stats[filename] = {
                'supplier': supplier,
                'rows': len(df),
                'columns': len(df.columns),
                'skus': df['sku'].nunique() if 'sku' in df.columns else 0
            }
            
            # Добавляем метаданные
            df_copy = df.copy()
            df_copy['supplier'] = supplier
            df_copy['supplier_email'] = supplier_email
            df_copy['source_file'] = filename
            df_copy['analysis_id'] = self.current_analysis_id
            
            all_data.append(df_copy)
        
        if not all_data:
            self.logger.warning("Нет данных для объединения")
            return pd.DataFrame()
        
        # Объединяем все DataFrame'ы
        combined_df = pd.concat(all_data, ignore_index=True)
        
        # Нормализуем SKU
        combined_df['sku'] = combined_df['sku'].astype(str).str.strip().str.upper()
        
        # Группируем по SKU и анализируем
        result = []
        
        for sku, group in combined_df.groupby('sku'):
            # Находим лучшую цену
            valid_prices = group[group['price'] > 0]
            
            if valid_prices.empty:
                continue
            
            min_price_row = valid_prices.loc[valid_prices['price'].idxmin()]
            max_price_row = valid_prices.loc[valid_prices['price'].idxmax()]
            
            # Собираем информацию о всех ценах
            all_prices = {}
            all_stocks = {}
            all_suppliers = set()
            
            for _, row in valid_prices.iterrows():
                supplier = row.get('supplier', 'Unknown')
                price = row.get('price', 0)
                stock = row.get('stock', 0)
                
                if pd.notna(price) and price > 0:
                    all_prices[supplier] = price
                    all_stocks[supplier] = stock
                    all_suppliers.add(supplier)
            
            # Рассчитываем метрики
            prices_list = list(all_prices.values())
            stocks_list = list(all_stocks.values())
            
            price_stats = {
                'min': min(prices_list) if prices_list else 0,
                'max': max(prices_list) if prices_list else 0,
                'mean': statistics.mean(prices_list) if prices_list else 0,
                'median': statistics.median(prices_list) if prices_list else 0,
                'stdev': statistics.stdev(prices_list) if len(prices_list) > 1 else 0
            }
            
            # Считаем потенциальную экономию
            potential_savings = price_stats['max'] - price_stats['min'] if len(prices_list) > 1 else 0
            savings_percent = (potential_savings / price_stats['max'] * 100) if price_stats['max'] > 0 else 0
            
            result_row = {
                'sku': sku,
                'brand': min_price_row.get('brand', ''),
                'name': min_price_row.get('name', ''),
                'category': min_price_row.get('category', ''),
                
                # Ценовая информация
                'min_price': price_stats['min'],
                'max_price': price_stats['max'],
                'avg_price': round(price_stats['mean'], 2),
                'median_price': round(price_stats['median'], 2),
                'price_std': round(price_stats['stdev'], 2),
                
                # Информация о поставщиках
                'best_supplier': min_price_row.get('supplier', 'Unknown'),
                'best_supplier_email': min_price_row.get('supplier_email', ''),
                'worst_supplier': max_price_row.get('supplier', 'Unknown'),
                'suppliers_count': len(all_suppliers),
                
                # Детальная информация
                'all_suppliers': '; '.join(sorted(all_suppliers)),
                'all_prices': '; '.join([f"{k}: {v:.2f}" for k, v in sorted(all_prices.items())]),
                'all_stocks': '; '.join([f"{k}: {v}" for k, v in sorted(all_stocks.items())]),
                
                # Метрики выгоды
                'price_diff': round(potential_savings, 2),
                'price_diff_percent': round(savings_percent, 2),
                'total_stock': sum(stocks_list),
                
                # Файлы-источники
                'best_source_file': min_price_row.get('source_file', ''),
                'analysis_id': self.current_analysis_id,
                
                # Дополнительная информация
                'description': min_price_row.get('description', ''),
                'weight': min_price_row.get('weight', 0),
                'barcode': min_price_row.get('barcode', ''),
                'country': min_price_row.get('country', ''),
                
                # Статистика
                'price_range_ratio': round(price_stats['max'] / price_stats['min'], 2) if price_stats['min'] > 0 else 0
            }
            
            result.append(result_row)
        
        # Создаем итоговый DataFrame
        result_df = pd.DataFrame(result)
        
        if not result_df.empty:
            # Сортируем по минимальной цене
            result_df = result_df.sort_values('min_price').reset_index(drop=True)
            
            # Добавляем колонку с рейтингом цены
            result_df['price_rank'] = result_df.groupby('category')['min_price'].rank(method='dense')
            
            # Выделяем лучшие предложения
            result_df['is_best_in_category'] = result_df.groupby('category')['min_price'].transform('min') == result_df['min_price']
        
        # Сохраняем результаты в БД
        db = Database(self.config.db_path)
        
        # Сохраняем сводную информацию об анализе
        db.save_analysis(
            analysis_id=self.current_analysis_id,
            supplier_count=len(file_stats),
            total_files=len(parsed_files),
            total_products=sum(s['rows'] for s in file_stats.values()),
            unique_skus=len(result_df),
            new_skus=0,
            price_changes=0,
            min_price=result_df['min_price'].min() if not result_df.empty else 0,
            max_price=result_df['max_price'].max() if not result_df.empty else 0,
            avg_price=result_df['avg_price'].mean() if not result_df.empty else 0,
            median_price=result_df['median_price'].median() if not result_df.empty else 0,
            duration=0,
            status='success',
            analysis_config=analysis_config
        )
        
        # Сохраняем сравнения цен
        for _, row in result_df.iterrows():
            suppliers = row['all_suppliers'].split('; ')
            prices = {}
            for price_info in row['all_prices'].split('; '):
                if ': ' in price_info:
                    supplier, price = price_info.split(': ')
                    prices[supplier.strip()] = float(price)
            
            for supplier in suppliers:
                supplier = supplier.strip()
                if supplier in prices:
                    db.save_price_comparison(
                        analysis_id=self.current_analysis_id,
                        sku=row['sku'],
                        supplier=supplier,
                        price=prices[supplier],
                        stock=0,
                        brand=row.get('brand', ''),
                        name=row.get('name', ''),
                        category=row.get('category', ''),
                        is_best_price=(supplier == row['best_supplier']),
                        price_rank=int(row.get('price_rank', 0))
                    )
        
        self.logger.info(
            f"Анализ завершен: {len(result_df)} уникальных SKU",
            context={'analysis_id': self.current_analysis_id}
        )
        
        return result_df
    
    def compare_with_previous_analysis(self, current_df: pd.DataFrame, 
                                      previous_analysis_id: str) -> pd.DataFrame:
        """Сравнение с предыдущим анализом"""
        db = Database(self.config.db_path)
        
        # Получаем данные предыдущего анализа
        with db.get_connection() as conn:
            previous_data = pd.read_sql_query('''
                SELECT sku, supplier, price, stock 
                FROM price_comparisons 
                WHERE analysis_id = ?
            ''', conn, params=(previous_analysis_id,))
        
        if previous_data.empty:
            return current_df
        
        # Добавляем информацию об изменениях
        changes = []
        
        for _, row in current_df.iterrows():
            sku = row['sku']
            prev_rows = previous_data[previous_data['sku'] == sku]
            
            if prev_rows.empty:
                row['change_type'] = 'new'
                row['old_min_price'] = 0
                row['price_change'] = 0
                row['price_change_percent'] = 0
            else:
                prev_best_price = prev_rows['price'].min()
                current_best_price = row['min_price']
                
                if current_best_price < prev_best_price:
                    row['change_type'] = 'decreased'
                elif current_best_price > prev_best_price:
                    row['change_type'] = 'increased'
                else:
                    row['change_type'] = 'unchanged'
                
                row['old_min_price'] = prev_best_price
                row['price_change'] = current_best_price - prev_best_price
                row['price_change_percent'] = ((current_best_price - prev_best_price) / prev_best_price * 100) if prev_best_price > 0 else 0
            
            changes.append(row)
        
        return pd.DataFrame(changes)
    
    def get_best_prices_by_category(self, df: pd.DataFrame) -> pd.DataFrame:
        """Получение лучших цен по категориям"""
        if df.empty:
            return pd.DataFrame()
        
        # Группируем по категориям и находим лучшие цены
        category_stats = df.groupby('category').agg({
            'sku': 'count',
            'min_price': ['min', 'mean', 'max'],
            'suppliers_count': 'mean',
            'price_diff': 'mean',
            'price_diff_percent': 'mean'
        }).round(2)
        
        category_stats.columns = [
            'Количество товаров', 
            'Мин. цена', 
            'Средняя цена', 
            'Макс. цена',
            'Среднее кол-во поставщиков',
            'Средняя разница цен',
            'Средняя разница (%)'
        ]
        
        return category_stats.reset_index()
    
    def get_supplier_performance(self, df: pd.DataFrame) -> pd.DataFrame:
        """Анализ эффективности поставщиков"""
        if df.empty:
            return pd.DataFrame()
        
        # Анализируем лучшие предложения по поставщикам
        supplier_stats = df['best_supplier'].value_counts().reset_index()
        supplier_stats.columns = ['Поставщик', 'Лучших предложений']
        
        # Добавляем информацию о средних ценах
        # Извлекаем цены из колонки all_prices
        supplier_prices = defaultdict(list)
        
        for _, row in df.iterrows():
            best_supplier = row['best_supplier']
            supplier_prices[best_supplier].append(row['min_price'])
        
        avg_prices = {k: statistics.mean(v) for k, v in supplier_prices.items()}
        supplier_stats['Средняя цена'] = supplier_stats['Поставщик'].map(avg_prices).round(2)
        
        # Процент от общего количества
        total_best = supplier_stats['Лучших предложений'].sum()
        supplier_stats['Доля (%)'] = (supplier_stats['Лучших предложений'] / total_best * 100).round(2)
        
        return supplier_stats.sort_values('Лучших предложений', ascending=False)
    
    def find_arbitrage_opportunities(self, df: pd.DataFrame, 
                                    min_margin_percent: float = 10.0) -> pd.DataFrame:
        """Поиск возможностей для арбитража"""
        if df.empty:
            return pd.DataFrame()
        
        opportunities = []
        
        for _, row in df.iterrows():
            # Ищем товары с большой разницей в ценах
            if row['price_diff_percent'] > min_margin_percent:
                # Парсим цены поставщиков
                prices = {}
                for price_info in row['all_prices'].split('; '):
                    if ': ' in price_info:
                        supplier, price = price_info.split(': ')
                        prices[supplier.strip()] = float(price)
                
                if len(prices) >= 2:
                    best_supplier = min(prices, key=prices.get)
                    worst_supplier = max(prices, key=prices.get)
                    
                    buy_price = prices[best_supplier]
                    sell_price = prices[worst_supplier]
                    
                    if buy_price > 0:
                        margin = sell_price - buy_price
                        margin_percent = (margin / buy_price) * 100
                        
                        if margin_percent > min_margin_percent:
                            opportunities.append({
                                'sku': row['sku'],
                                'name': row.get('name', ''),
                                'category': row.get('category', ''),
                                'buy_from': best_supplier,
                                'buy_price': buy_price,
                                'sell_price': sell_price,
                                'margin': round(margin, 2),
                                'margin_percent': round(margin_percent, 2),
                                'potential_profit': round(margin, 2)
                            })
        
        return pd.DataFrame(opportunities).sort_values('margin_percent', ascending=False)
    
    def generate_price_report(self, df: pd.DataFrame, 
                            report_type: str = 'full') -> str:
        """Генерация отчета о ценах"""
        reports_dir = Path(self.config.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"price_report_{report_type}_{timestamp}.xlsx"
        filepath = reports_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Основной лист с результатами
            df.to_excel(writer, sheet_name='Анализ цен', index=False)
            
            # Лучшие цены по категориям
            category_stats = self.get_best_prices_by_category(df)
            category_stats.to_excel(writer, sheet_name='По категориям', index=False)
            
            # Эффективность поставщиков
            supplier_stats = self.get_supplier_performance(df)
            supplier_stats.to_excel(writer, sheet_name='Поставщики', index=False)
            
            # Арбитражные возможности
            arbitrage = self.find_arbitrage_opportunities(df)
            if not arbitrage.empty:
                arbitrage.to_excel(writer, sheet_name='Арбитраж', index=False)
            
            # Форматируем листы
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Форматируем заголовки
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Автоподбор ширины
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Отчет сохранен: {filepath}")
        return str(filepath)
    
    def export_to_csv(self, df: pd.DataFrame, filename: str = None) -> str:
        """Экспорт результатов в CSV"""
        if filename is None:
            filename = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = self.analysis_dir / filename
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        self.logger.info(f"Результаты экспортированы в CSV: {filepath}")
        return str(filepath)
    
    def export_to_excel(self, df: pd.DataFrame, filename: str = None) -> str:
        """Расширенный экспорт в Excel с форматированием и графиками"""
        if filename is None:
            filename = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.analysis_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Анализ цен', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Анализ цен']
            
            # Форматирование
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Условное форматирование для цен
            price_col_letter = None
            for idx, col in enumerate(df.columns):
                if col == 'min_price':
                    price_col_letter = get_column_letter(idx + 1)
                    break
            
            if price_col_letter:
                # Цветовая шкала для цен
                worksheet.conditional_formatting.add(
                    f'{price_col_letter}2:{price_col_letter}{len(df) + 1}',
                    ColorScaleRule(
                        start_type='min', start_color='63BE7B',  # Зеленый (низкая цена)
                        mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Желтый
                        end_type='max', end_color='F8696B'  # Красный (высокая цена)
                    )
                )
            
            # Добавляем график распределения цен
            if len(df) > 1:
                chart_sheet = workbook.create_sheet('Графики')
                
                # Гистограмма цен
                chart = BarChart()
                chart.type = "col"
                chart.title = "Распределение цен"
                chart.y_axis.title = 'Количество'
                chart.x_axis.title = 'Ценовой диапазон'
                
                # Создаем данные для гистограммы
                prices = df['min_price'].dropna()
                if len(prices) > 0:
                    # Создаем вспомогательный лист для данных графика
                    data_sheet = workbook.create_sheet('_data')
                    
                    # Вычисляем бины
                    bins = np.histogram(prices, bins=10)
                    data_sheet['A1'] = 'Диапазон цен'
                    data_sheet['B1'] = 'Количество'
                    
                    for i in range(len(bins[0])):
                        data_sheet[f'A{i+2}'] = f'{bins[1][i]:.0f}-{bins[1][i+1]:.0f}'
                        data_sheet[f'B{i+2}'] = int(bins[0][i])
                    
                    data_ref = Reference(data_sheet, min_col=2, min_row=1, max_row=len(bins[0])+1)
                    cats_ref = Reference(data_sheet, min_col=1, min_row=2, max_row=len(bins[0])+1)
                    
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    
                    chart_sheet.add_chart(chart, "A1")
                    
                    # Скрываем вспомогательный лист
                    data_sheet.sheet_state = 'hidden'
            
            # Автоподбор ширины колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_length = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_length
            
            # Добавляем автофильтр
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        self.logger.info(f"Результаты экспортированы в Excel: {filepath}")
        return str(filepath)


# ===================================================================
# БЛОК 13: СИСТЕМА УВЕДОМЛЕНИЙ
# ===================================================================

class NotificationManager:
    """Класс для управления уведомлениями"""
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.notification_config = NotificationConfig.from_dict(config.notification_config)
    
    def send_email_notification(self, subject: str, body: str, 
                               recipients: Optional[List[str]] = None,
                               attachments: Optional[List[str]] = None) -> bool:
        """Отправка email-уведомления"""
        if not self.notification_config.email_notifications:
            return False
        
        try:
            recipients = recipients or self.notification_config.email_recipients
            if not recipients:
                self.logger.warning("Нет получателей для email-уведомления")
                return False
            
            # Создаем письмо
            msg = MIMEMultipart()
            msg['From'] = self.config.email_user
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = subject
            msg['Date'] = formatdate(localtime=True)
            msg['Message-ID'] = make_msgid()
            
            # Добавляем тело письма
            msg.attach(MIMEText(body, 'html', 'utf-8'))
            
            # Добавляем вложения
            if attachments:
                for filepath in attachments:
                    if os.path.exists(filepath):
                        with open(filepath, 'rb') as f:
                            attachment = MIMEApplication(f.read())
                            attachment.add_header(
                                'Content-Disposition',
                                'attachment',
                                filename=os.path.basename(filepath)
                            )
                            msg.attach(attachment)
            
            # Отправляем письмо
            with smtplib.SMTP_SSL(self.config.imap_server, 465) as server:
                server.login(self.config.email_user, self.config.email_pass)
                server.send_message(msg)
            
            self.logger.info(f"Email-уведомление отправлено: {subject}")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка отправки email: {e}")
            return False
    
    def send_telegram_notification(self, message: str, 
                                  parse_mode: str = 'HTML') -> bool:
        """Отправка уведомления в Telegram"""
        if not self.notification_config.telegram_enabled:
            return False
        
        try:
            token = self.notification_config.telegram_bot_token
            chat_id = self.notification_config.telegram_chat_id
            
            if not token or not chat_id:
                self.logger.warning("Не настроены параметры Telegram")
                return False
            
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            
            # Разбиваем длинные сообщения
            max_length = 4000
            messages = []
            
            if len(message) > max_length:
                # Разбиваем по строкам
                lines = message.split('\n')
                current_message = ''
                
                for line in lines:
                    if len(current_message) + len(line) + 1 > max_length:
                        messages.append(current_message)
                        current_message = line
                    else:
                        if current_message:
                            current_message += '\n' + line
                        else:
                            current_message = line
                
                if current_message:
                    messages.append(current_message)
            else:
                messages = [message]
            
            # Отправляем сообщения
            for i, msg_text in enumerate(messages):
                payload = {
                    'chat_id': chat_id,
                    'text': msg_text,
                    'parse_mode': parse_mode,
                    'disable_web_page_preview': True
                }
                
                response = requests.post(url, json=payload, timeout=10)
                
                if response.status_code != 200:
                    self.logger.error(f"Ошибка отправки в Telegram: {response.text}")
                    return False
                
                # Задержка между сообщениями
                if i < len(messages) - 1:
                    time.sleep(0.5)
            
            self.logger.info("Уведомление отправлено в Telegram")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка отправки в Telegram: {e}")
            return False
    
    def notify_price_changes(self, changes: List[Dict]) -> None:
        """Уведомление об изменениях цен"""
        if not self.notification_config.notify_on_price_changes:
            return
        
        # Фильтруем значительные изменения
        significant_changes = [
            c for c in changes 
            if abs(c.get('change_percent', 0)) >= self.notification_config.min_price_change_percent
        ]
        
        if not significant_changes:
            return
        
        # Формируем сообщение
        subject = f"📊 Изменение цен - {len(significant_changes)} товаров"
        
        body = f"""
        <html>
        <head><meta charset="utf-8"></head>
        <body>
        <h2>Значительные изменения цен</h2>
        <p>Найдено изменений: <b>{len(significant_changes)}</b></p>
        <table border="1" cellpadding="5" cellspacing="0">
        <tr style="background-color: #366092; color: white;">
            <th>Артикул</th>
            <th>Поставщик</th>
            <th>Старая цена</th>
            <th>Новая цена</th>
            <th>Изменение</th>
            <th>Изменение %</th>
        </tr>
        """
        
        for change in significant_changes[:50]:  # Ограничиваем 50 товарами
            color = 'red' if change.get('new_price', 0) > change.get('old_price', 0) else 'green'
            body += f"""
            <tr>
                <td>{change.get('sku', '')}</td>
                <td>{change.get('supplier', '')}</td>
                <td>{change.get('old_price', 0):.2f}</td>
                <td>{change.get('new_price', 0):.2f}</td>
                <td style="color: {color};">{change.get('change_absolute', 0):+.2f}</td>
                <td style="color: {color};">{change.get('change_percent', 0):+.1f}%</td>
            </tr>
            """
        
        body += "</table></body></html>"
        
        # Отправляем уведомления
        self.send_email_notification(subject, body)
        
        telegram_msg = f"📊 <b>Изменение цен</b>\n\n"
        telegram_msg += f"Найдено изменений: <b>{len(significant_changes)}</b>\n\n"
        
        for change in significant_changes[:10]:  # В Telegram отправляем только топ-10
            emoji = '🔺' if change.get('new_price', 0) > change.get('old_price', 0) else '🔻'
            telegram_msg += (
                f"{emoji} <code>{change.get('sku', '')}</code>: "
                f"{change.get('old_price', 0):.2f} → {change.get('new_price', 0):.2f} "
                f"({change.get('change_percent', 0):+.1f}%)\n"
            )
        
        self.send_telegram_notification(telegram_msg)
    
    def notify_run_complete(self, result: Dict) -> None:
        """Уведомление о завершении работы робота"""
        if result['status'] == 'success' and not self.notification_config.notify_on_success:
            return
        
        if result['status'] in ['failed', 'critical_failed'] and not self.notification_config.notify_on_failure:
            return
        
        # Формируем тему
        status_emoji = {
            'success': '✅',
            'no_files': 'ℹ️',
            'partial_success': '⚠️',
            'failed': '❌',
            'critical_failed': '💥'
        }
        
        emoji = status_emoji.get(result['status'], '❓')
        subject = f"{emoji} Робот завершил работу - {result['status']}"
        
        # Формируем тело письма
        body = f"""
        <html>
        <head><meta charset="utf-8"></head>
        <body>
        <h2>Результаты работы робота</h2>
        <table border="1" cellpadding="10" cellspacing="0">
        <tr><td><b>Статус</b></td><td>{result['status']}</td></tr>
        <tr><td><b>Найдено файлов</b></td><td>{result.get('files_found', 0)}</td></tr>
        <tr><td><b>Обработано файлов</b></td><td>{result.get('files_processed', 0)}</td></tr>
        <tr><td><b>Обновлено товаров</b></td><td>{result.get('products_updated', 0)}</td></tr>
        <tr><td><b>Добавлено товаров</b></td><td>{result.get('products_added', 0)}</td></tr>
        <tr><td><b>Отправлено в Яндекс</b></td><td>{result.get('offers_sent', 0)}</td></tr>
        <tr><td><b>Ошибок</b></td><td>{len(result.get('errors', []))}</td></tr>
        <tr><td><b>Время выполнения</b></td><td>{result.get('duration', 0):.1f} сек</td></tr>
        </table>
        """
        
        if result.get('errors'):
            body += "<h3>Ошибки:</h3><ul>"
            for error in result['errors'][:10]:
                body += f"<li>{error}</li>"
            body += "</ul>"
        
        body += "</body></html>"
        
        # Отправляем уведомления
        self.send_email_notification(subject, body)
        
        # Telegram
        telegram_msg = f"{emoji} <b>Робот завершил работу</b>\n\n"
        telegram_msg += f"📁 Файлов: {result.get('files_processed', 0)}\n"
        telegram_msg += f"🔄 Обновлено: {result.get('products_updated', 0)}\n"
        telegram_msg += f"📦 Отправлено: {result.get('offers_sent', 0)}\n"
        telegram_msg += f"❌ Ошибок: {len(result.get('errors', []))}\n"
        telegram_msg += f"⏱️ Время: {result.get('duration', 0):.1f} сек"
        
        self.send_telegram_notification(telegram_msg)
    
    def send_daily_summary(self, stats: Dict) -> None:
        """Отправка ежедневной сводки"""
        if not self.notification_config.daily_summary:
            return
        
        subject = f"📊 Ежедневная сводка - {datetime.now().strftime('%d.%m.%Y')}"
        
        body = f"""
        <html>
        <head><meta charset="utf-8"></head>
        <body>
        <h2>Ежедневная сводка работы робота</h2>
        <table border="1" cellpadding="10" cellspacing="0">
        <tr><td><b>Всего товаров в базе</b></td><td>{stats.get('total_products', 0)}</td></tr>
        <tr><td><b>Активных поставщиков</b></td><td>{stats.get('active_suppliers', 0)}</td></tr>
        <tr><td><b>Запусков за сегодня</b></td><td>{stats.get('runs_today', 0)}</td></tr>
        <tr><td><b>Обработано файлов</b></td><td>{stats.get('files_today', 0)}</td></tr>
        <tr><td><b>Обновлено товаров</b></td><td>{stats.get('updates_today', 0)}</td></tr>
        <tr><td><b>Ошибок</b></td><td>{stats.get('errors_today', 0)}</td></tr>
        </table>
        </body>
        </html>
        """
        
        self.send_email_notification(subject, body)

# ===================================================================
# БЛОК 14: РАСШИРЕННЫЙ ОСНОВНОЙ РОБОТ
# ===================================================================

class PriceRobot:
    """Расширенный основной класс робота с полной интеграцией всех компонентов"""
    
    def __init__(self, config: Config, logger: StreamlitLogger):
        self.config = config
        self.logger = logger
        self.db = Database(config.db_path)
        self.product_db = GoogleSheetsDatabase(config, logger)
        self.email_downloader = EmailDownloader(config, logger)
        self.price_parser = PriceParser(config, logger)
        self.yandex_client = YandexMarketClient(config, logger)
        self.notification_manager = NotificationManager(config, logger)
        self.multi_supplier_downloader = MultiSupplierDownloader(config, logger)
        self.price_analyzer = PriceAnalyzer(config, logger)
        self.mapping_constructor = MappingConstructor(config, logger)
        
        # Создаем необходимые директории
        for dir_path in [config.uploads_dir, config.archive_dir, config.temp_dir,
                        config.cache_dir, config.backup_dir, config.analysis_dir,
                        config.reports_dir, config.exports_dir]:
            Path(dir_path).mkdir(parents=True, exist_ok=True)
        
        # Статистика текущего запуска
        self.run_stats = {
            'start_time': time.time(),
            'files_found': 0,
            'files_processed': 0,
            'files_skipped': 0,
            'files_with_errors': 0,
            'products_updated': 0,
            'products_added': 0,
            'products_unchanged': 0,
            'offers_sent': 0,
            'offers_failed': 0,
            'errors': [],
            'warnings': [],
            'price_changes': [],
            'processed_suppliers': set(),
            'processed_files_info': []
        }
    
    def _save_and_archive_file(self, filename: str, content: bytes, 
                              supplier: str = '') -> str:
        """Сохранение и архивирование файла"""
        # Сохраняем в uploads
        safe_filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        upload_filename = f"{timestamp}_{supplier}_{safe_filename}" if supplier else f"{timestamp}_{safe_filename}"
        
        upload_path = Path(self.config.uploads_dir) / upload_filename
        with open(upload_path, 'wb') as f:
            f.write(content)
        
        # Архивируем
        archive_path = Path(self.config.archive_dir) / upload_filename
        shutil.copy2(upload_path, archive_path)
        
        self.logger.debug(f"Файл сохранен: {upload_path}")
        return str(upload_path)
    
    def _calculate_retail_price(self, base_price: float, supplier_markup: float = 0) -> float:
        """Расчет розничной цены с учетом наценки"""
        if supplier_markup > 0:
            markup = supplier_markup
        else:
            markup = self.config.markup_percent
        
        return round(base_price * (1 + markup / 100), 2)
    
    def _process_single_file(self, file_data: Dict) -> Dict:
        """Обработка одного файла прайса"""
        filename = file_data.get('filename', 'unknown')
        content = file_data.get('content', b'')
        supplier_name = file_data.get('supplier', 'Unknown')
        supplier_email = file_data.get('supplier_email', '')
        
        file_result = {
            'filename': filename,
            'supplier': supplier_name,
            'status': 'pending',
            'products_found': 0,
            'products_updated': 0,
            'products_added': 0,
            'products_skipped': 0,
            'errors': [],
            'warnings': [],
            'processing_time': 0
        }
        
        start_time = time.time()
        
        try:
            # Проверяем, не обработан ли уже файл
            if self.db.is_file_processed(content):
                self.logger.info(f"Файл {filename} уже был обработан, пропускаем")
                file_result['status'] = 'skipped'
                file_result['warnings'].append('Файл уже был обработан')
                self.run_stats['files_skipped'] += 1
                return file_result
            
            # Сохраняем файл
            saved_path = self._save_and_archive_file(filename, content, supplier_name)
            
            # Получаем маппинг для поставщика
            mapping = self.config.get_supplier_mapping(supplier_name)
            supplier_config = self.config.get_supplier_by_name(supplier_name)
            
            # Получаем правила обработки
            processing_rules = None
            custom_markup = 0
            if supplier_config:
                processing_rules = supplier_config.processing_rules
                custom_markup = supplier_config.get_effective_markup(self.config.markup_percent)
            
            # Парсим файл
            if mapping and any(mapping.values()):
                df = self.price_parser.parse(filename, content, mapping, processing_rules)
                self.logger.info(f"Файл {filename} распарсен с сохраненным маппингом")
            else:
                df = self.price_parser.parse(filename, content, None, processing_rules)
                if mapping:
                    self.logger.warning(f"Для поставщика {supplier_name} маппинг пуст, использовано автоопределение")
                else:
                    self.logger.info(f"Файл {filename} распарсен с автоопределением колонок")
            
            if df.empty:
                file_result['status'] = 'no_data'
                file_result['warnings'].append('Файл не содержит данных')
                return file_result
            
            file_result['products_found'] = len(df)
            
            # Валидируем данные
            validation = self.price_parser.validate_data(df)
            if not validation.is_valid:
                file_result['warnings'].extend(validation.errors)
                file_result['warnings'].extend(validation.warnings)
            
            # Обрабатываем товары
            updated, added, skipped, price_changes = self._process_products_from_df(
                df, supplier_name, custom_markup
            )
            
            file_result['products_updated'] = updated
            file_result['products_added'] = added
            file_result['products_skipped'] = skipped
            file_result['processing_time'] = time.time() - start_time
            file_result['status'] = 'processed'
            
            # Сохраняем информацию о файле в БД
            self.db.mark_file_processed(
                filename=filename,
                file_content=content,
                offers_count=len(df),
                supplier=supplier_name,
                status='processed',
                valid_offers=updated + added,
                invalid_offers=skipped,
                new_products=added,
                updated_products=updated,
                processing_time=file_result['processing_time'],
                metadata={
                    'supplier_email': supplier_email,
                    'mapping_used': mapping,
                    'validation_stats': validation.statistics
                }
            )
            
            self.run_stats['files_processed'] += 1
            self.run_stats['products_updated'] += updated
            self.run_stats['products_added'] += added
            self.run_stats['processed_suppliers'].add(supplier_name)
            self.run_stats['processed_files_info'].append(file_result)
            
            self.logger.info(
                f"Файл {filename} обработан: обновлено {updated}, "
                f"добавлено {added}, пропущено {skipped} товаров"
            )
            
        except Exception as e:
            file_result['status'] = 'error'
            file_result['errors'].append(str(e))
            file_result['processing_time'] = time.time() - start_time
            
            self.logger.error(f"Ошибка обработки файла {filename}: {e}")
            self.run_stats['files_with_errors'] += 1
            self.run_stats['errors'].append(f"Файл {filename}: {str(e)}")
            
            # Сохраняем ошибку в БД
            try:
                self.db.mark_file_processed(
                    filename=filename,
                    file_content=content,
                    offers_count=0,
                    supplier=supplier_name,
                    status='error',
                    error_message=str(e)
                )
            except:
                pass
        
        return file_result
    
    def _process_products_from_df(self, df: pd.DataFrame, supplier_name: str = '',
                                 custom_markup: float = 0) -> Tuple[int, int, int, List[Dict]]:
        """Расширенная обработка товаров из DataFrame"""
        updated = 0
        added = 0
        skipped = 0
        price_changes = []
        
        threshold = self.config.min_stock_threshold
        
        for _, row in df.iterrows():
            try:
                sku = str(row.get('sku', '')).strip().upper()
                if not sku or sku in ['', 'NAN', 'NONE', 'NULL']:
                    skipped += 1
                    continue
                
                # Получаем данные из файла
                base_price = float(row.get('price', 0))
                stock_from_file = int(row.get('stock', 0))
                brand = str(row.get('brand', ''))
                name = str(row.get('name', ''))
                category = str(row.get('category', ''))
                description = str(row.get('description', ''))
                
                # Пропускаем товары с нулевой ценой
                if base_price <= 0:
                    self.logger.debug(f"Товар {sku}: нулевая цена, пропускаем")
                    skipped += 1
                    continue
                
                # Рассчитываем розничную цену
                retail_price = self._calculate_retail_price(base_price, custom_markup)
                
                # Определяем остаток для отправки
                if stock_from_file < threshold:
                    stock_to_send = 0
                else:
                    stock_to_send = stock_from_file
                
                # Проверяем существование товара в базе
                existing = self.product_db.get_product_by_sku(sku)
                
                if existing:
                    # Товар существует - обновляем
                    old_price = float(existing.get('цена_розница', 0))
                    old_stock = int(existing.get('остаток', 0))
                    old_base_price = float(existing.get('цена_базовая', 0))
                    
                    # Определяем, что изменилось
                    update_data = {}
                    
                    if abs(base_price - old_base_price) > 0.01:
                        update_data['цена_базовая'] = base_price
                        update_data['цена_розница'] = retail_price
                        
                        # Записываем изменение цены
                        price_change = {
                            'sku': sku,
                            'supplier': supplier_name,
                            'old_price': old_price,
                            'new_price': retail_price,
                            'old_stock': old_stock,
                            'new_stock': stock_to_send,
                            'change_absolute': retail_price - old_price,
                            'change_percent': ((retail_price - old_price) / old_price * 100) if old_price > 0 else 0
                        }
                        price_changes.append(price_change)
                        
                        # Сохраняем в историю
                        self.db.add_price_history(
                            sku=sku,
                            old_price=old_price,
                            new_price=retail_price,
                            old_stock=old_stock,
                            new_stock=stock_to_send,
                            supplier=supplier_name
                        )
                    
                    if stock_from_file != old_stock:
                        update_data['остаток'] = stock_to_send
                    
                    # Обновляем дополнительные поля если они изменились
                    if brand and brand != existing.get('бренд', ''):
                        update_data['бренд'] = brand
                    
                    if name and name != existing.get('название', ''):
                        update_data['название'] = name
                    
                    if category and category != existing.get('категория', ''):
                        update_data['категория'] = category
                    
                    if description and description != existing.get('описание', ''):
                        update_data['описание'] = description
                    
                    if update_data:
                        update_data['источник'] = supplier_name
                        update_data['дата_обновления'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        self.product_db.update_product(sku, **update_data)
                        updated += 1
                    else:
                        skipped += 1
                else:
                    # Товар не существует - создаем новый
                    new_product_data = {
                        'артикул': sku,
                        'цена_базовая': base_price,
                        'цена_розница': retail_price,
                        'остаток': stock_to_send,
                        'бренд': brand,
                        'название': name,
                        'категория': category,
                        'описание': description,
                        'вес': row.get('weight', 0),
                        'штрихкод': str(row.get('barcode', '')),
                        'страна': str(row.get('country', '')),
                        'гарантия': str(row.get('warranty', '')),
                        'источник': supplier_name,
                        'дата_обновления': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    self.product_db.update_product(sku, **new_product_data)
                    added += 1
                    
                    # Записываем новый товар в историю
                    self.db.add_price_history(
                        sku=sku,
                        old_price=0,
                        new_price=retail_price,
                        old_stock=0,
                        new_stock=stock_to_send,
                        supplier=supplier_name
                    )
                
            except Exception as e:
                self.logger.error(f"Ошибка обработки товара {row.get('sku', 'unknown')}: {e}")
                skipped += 1
                continue
        
        # Сохраняем изменения цен для уведомлений
        if price_changes:
            self.run_stats['price_changes'].extend(price_changes)
        
        return updated, added, skipped, price_changes
    
    def run_full_cycle(self, progress_callback: Optional[Callable] = None) -> Dict:
        """Полный цикл работы робота: загрузка, обработка, отправка"""
        start_time = time.time()
        run_id = f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        
        result = {
            'run_id': run_id,
            'status': 'pending',
            'files_found': 0,
            'files_processed': 0,
            'files_skipped': 0,
            'files_with_errors': 0,
            'products_updated': 0,
            'products_added': 0,
            'offers_sent': 0,
            'offers_failed': 0,
            'errors': [],
            'warnings': [],
            'price_changes': [],
            'duration': 0,
            'start_time': datetime.now().isoformat(),
            'end_time': None
        }
        
        try:
            # Этап 1: Проверка подключений
            if progress_callback:
                progress_callback("Проверка подключений...")
            
            self.logger.info("=" * 60)
            self.logger.info(f"ЗАПУСК РОБОТА: {run_id}")
            self.logger.info("=" * 60)
            
            # Проверяем Google Sheets
            self.logger.info("Проверка подключения к Google Sheets...")
            if self.product_db.df is None:
                self.product_db._load_data()
            
            stats = self.product_db.get_stats()
            self.logger.info(f"Google Sheets: {stats['total_products']} товаров в базе")
            
            # Проверяем Яндекс Маркет
            if self.config.yandex_auto_update:
                self.logger.info("Проверка подключения к Яндекс Маркет...")
                yandex_ok, yandex_msg = self.yandex_client.test_connection()
                if not yandex_ok:
                    self.logger.warning(f"Яндекс Маркет недоступен: {yandex_msg}")
                    result['warnings'].append(f"Яндекс Маркет: {yandex_msg}")
            
            # Этап 2: Загрузка прайсов от поставщиков
            if progress_callback:
                progress_callback("Загрузка прайсов от поставщиков...")
            
            self.logger.info("Загрузка прайсов от всех поставщиков...")
            
            # Используем многопоставщичный загрузчик
            downloaded_files = self.multi_supplier_downloader.download_all_suppliers()
            
            result['files_found'] = len(downloaded_files)
            self.run_stats['files_found'] = len(downloaded_files)
            
            if not downloaded_files:
                self.logger.info("Новых прайсов не найдено")
                result['status'] = 'no_files'
                result['warnings'].append('Новых файлов не найдено')
            else:
                self.logger.info(f"Загружено {len(downloaded_files)} файлов от {len(self.run_stats['processed_suppliers'])} поставщиков")
                
                # Этап 3: Обработка файлов
                if progress_callback:
                    progress_callback(f"Обработка {len(downloaded_files)} файлов...")
                
                # Обрабатываем файлы (можно параллельно)
                if self.config.parallel_parsing and len(downloaded_files) > 1:
                    with ThreadPoolExecutor(max_workers=self.config.max_workers) as executor:
                        futures = []
                        for file_data in downloaded_files:
                            future = executor.submit(self._process_single_file, file_data)
                            futures.append(future)
                        
                        for i, future in enumerate(as_completed(futures)):
                            file_result = future.result()
                            if progress_callback:
                                progress_callback(f"Обработано {i+1}/{len(downloaded_files)} файлов")
                else:
                    for i, file_data in enumerate(downloaded_files):
                        if progress_callback:
                            progress_callback(f"Обработка файла {i+1}/{len(downloaded_files)}: {file_data.get('filename', 'unknown')[:50]}")
                        self._process_single_file(file_data)
                
                # Обновляем результат
                result['files_processed'] = self.run_stats['files_processed']
                result['files_skipped'] = self.run_stats['files_skipped']
                result['files_with_errors'] = self.run_stats['files_with_errors']
                result['products_updated'] = self.run_stats['products_updated']
                result['products_added'] = self.run_stats['products_added']
                result['price_changes'] = self.run_stats['price_changes']
            
            # Этап 4: Отправка в Яндекс Маркет
            if self.config.yandex_auto_update:
                if progress_callback:
                    progress_callback("Подготовка и отправка данных в Яндекс Маркет...")
                
                self.logger.info("Подготовка товаров для Яндекс Маркета...")
                
                all_offers = self.product_db.get_products_for_yandex(
                    markup_percent=self.config.markup_percent,
                    min_stock_threshold=self.config.min_stock_threshold
                )
                
                if all_offers:
                    self.logger.info(f"Отправка {len(all_offers)} товаров в Яндекс Маркет...")
                    
                    # Валидируем офферы перед отправкой
                    valid_offers = self.yandex_client.validate_offers(all_offers)
                    
                    if len(valid_offers) < len(all_offers):
                        self.logger.warning(
                            f"Отфильтровано {len(all_offers) - len(valid_offers)} некорректных офферов"
                        )
                    
                    # Отправляем цены и остатки
                    update_result = self.yandex_client.bulk_update_prices_and_stocks(valid_offers)
                    
                    result['offers_sent'] = update_result['prices']['updated']
                    result['offers_failed'] = len(update_result['prices']['errors']) + len(update_result['stocks']['errors'])
                    
                    # Сохраняем в кэш
                    for offer in valid_offers:
                        self.db.update_sent_cache(
                            sku=offer['sku'],
                            price=offer['price'],
                            stock=offer['stock']
                        )
                    
                    if update_result['prices']['errors']:
                        for err in update_result['prices']['errors']:
                            result['errors'].append(f"Цены: {err.get('error', 'Unknown error')}")
                    
                    if update_result['stocks']['errors']:
                        for err in update_result['stocks']['errors']:
                            result['errors'].append(f"Остатки: {err.get('error', 'Unknown error')}")
                    
                    self.logger.info(
                        f"Отправлено в Яндекс: {result['offers_sent']} товаров, "
                        f"ошибок: {result['offers_failed']}"
                    )
                else:
                    self.logger.warning("Нет товаров для отправки в Яндекс Маркет")
                    result['warnings'].append("База товаров пуста - нечего отправлять")
            
            # Определяем итоговый статус
            if result['errors']:
                if result['files_processed'] > 0:
                    result['status'] = 'partial_success'
                else:
                    result['status'] = 'failed'
            else:
                if result['files_processed'] > 0 or result['offers_sent'] > 0:
                    result['status'] = 'success'
                else:
                    result['status'] = 'no_files'
            
            # Сохраняем статистику запуска
            duration = time.time() - start_time
            result['duration'] = duration
            result['end_time'] = datetime.now().isoformat()
            
            self.db.add_run_stats(
                run_id=run_id,
                files_found=result['files_found'],
                files_processed=result['files_processed'],
                files_skipped=result['files_skipped'],
                files_with_errors=result['files_with_errors'],
                offers_sent=result['offers_sent'],
                offers_failed=result['offers_failed'],
                products_updated=result['products_updated'],
                products_added=result['products_added'],
                errors=len(result['errors']),
                warnings=len(result['warnings']),
                duration=duration,
                status=result['status'],
                run_type='manual',
                triggered_by='user'
            )
            
            # Отправляем уведомления
            if self.notification_manager.notification_config.enabled:
                try:
                    self.notification_manager.notify_run_complete(result)
                    
                    if result['price_changes']:
                        self.notification_manager.notify_price_changes(result['price_changes'])
                except Exception as e:
                    self.logger.error(f"Ошибка отправки уведомлений: {e}")
            
            self.logger.info("=" * 60)
            self.logger.info(f"РОБОТ ЗАВЕРШИЛ РАБОТУ: {duration:.1f} сек")
            self.logger.info(f"Статус: {result['status']}")
            self.logger.info(f"Файлов: {result['files_processed']} обработано, {result['files_skipped']} пропущено")
            self.logger.info(f"Товаров: {result['products_updated']} обновлено, {result['products_added']} добавлено")
            self.logger.info(f"Отправлено в Яндекс: {result['offers_sent']}")
            self.logger.info("=" * 60)
            
        except Exception as e:
            self.logger.critical(f"Критическая ошибка в работе робота: {e}", exc_info=True)
            result['status'] = 'critical_failed'
            result['errors'].append(f"Критическая ошибка: {str(e)}")
            result['duration'] = time.time() - start_time
            result['end_time'] = datetime.now().isoformat()
        
        return result
    
    def run_single_supplier(self, supplier_name: str) -> Dict:
        """Запуск обработки только для одного поставщика"""
        supplier = self.config.get_supplier_by_name(supplier_name)
        if not supplier:
            return {'status': 'error', 'error': f'Поставщик {supplier_name} не найден'}
        
        self.logger.info(f"Запуск обработки для поставщика: {supplier_name}")
        
        # Загружаем файлы поставщика
        downloaded_files = self.multi_supplier_downloader.download_from_supplier(supplier)
        
        if not downloaded_files.get('downloaded_files'):
            return {
                'status': 'no_files',
                'message': f'Нет новых файлов от поставщика {supplier_name}'
            }
        
        # Обрабатываем каждый файл
        results = []
        for file_data in downloaded_files['downloaded_files']:
            file_result = self._process_single_file(file_data)
            results.append(file_result)
        
        return {
            'status': 'success',
            'supplier': supplier_name,
            'files_processed': len(results),
            'results': results
        }
    
    def quick_analysis(self) -> pd.DataFrame:
        """Быстрый анализ текущих прайсов"""
        self.logger.info("Запуск быстрого анализа...")
        
        # Загружаем файлы
        downloaded_files = self.multi_supplier_downloader.download_all_suppliers()
        
        if not downloaded_files:
            return pd.DataFrame()
        
        # Парсим файлы
        parsed_files = {}
        for file_data in downloaded_files:
            try:
                filename = file_data.get('filename', 'unknown')
                content = file_data.get('content', b'')
                supplier_name = file_data.get('supplier', 'Unknown')
                
                mapping = self.config.get_supplier_mapping(supplier_name)
                
                if mapping and any(mapping.values()):
                    df = self.price_parser.parse(filename, content, mapping)
                else:
                    df = self.price_parser.parse(filename, content)
                
                if not df.empty:
                    df['supplier'] = supplier_name
                    df['supplier_email'] = file_data.get('supplier_email', '')
                    df['source_file'] = filename
                    parsed_files[filename] = df
                    
            except Exception as e:
                self.logger.error(f"Ошибка парсинга {filename}: {e}")
        
        if not parsed_files:
            return pd.DataFrame()
        
        # Анализируем
        result_df = self.price_analyzer.analyze(parsed_files)
        
        return result_df
    
    def get_run_summary(self) -> Dict:
        """Получение сводки о текущем запуске"""
        return {
            'runtime': time.time() - self.run_stats['start_time'],
            'files_found': self.run_stats['files_found'],
            'files_processed': self.run_stats['files_processed'],
            'files_skipped': self.run_stats['files_skipped'],
            'files_with_errors': self.run_stats['files_with_errors'],
            'products_updated': self.run_stats['products_updated'],
            'products_added': self.run_stats['products_added'],
            'offers_sent': self.run_stats['offers_sent'],
            'offers_failed': self.run_stats['offers_failed'],
            'errors_count': len(self.run_stats['errors']),
            'warnings_count': len(self.run_stats['warnings']),
            'suppliers_processed': len(self.run_stats['processed_suppliers']),
            'price_changes_count': len(self.run_stats['price_changes'])
        }
    
    def get_daily_stats(self) -> Dict:
        """Получение статистики за сегодня"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        stats = {
            'date': today,
            'total_products': 0,
            'active_suppliers': len(self.config.get_active_suppliers()),
            'runs_today': 0,
            'files_today': 0,
            'updates_today': 0,
            'errors_today': 0
        }
        
        # Получаем статистику из БД
        db_stats = self.db.get_stats(1)
        if not db_stats.empty:
            stats['runs_today'] = len(db_stats)
            stats['files_today'] = db_stats['files_processed'].sum()
            stats['updates_today'] = db_stats['products_updated'].sum()
            stats['errors_today'] = db_stats['errors'].sum()
        
        # Получаем статистику товаров
        product_stats = self.product_db.get_stats()
        stats['total_products'] = product_stats['total_products']
        
        return stats
    
    def cleanup_temp_files(self, max_age_hours: int = 24) -> int:
        """Очистка временных файлов"""
        cleaned = 0
        temp_dir = Path(self.config.temp_dir)
        
        if temp_dir.exists():
            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
            
            for item in temp_dir.rglob('*'):
                if item.is_file():
                    mtime = datetime.fromtimestamp(item.stat().st_mtime)
                    if mtime < cutoff_time:
                        try:
                            item.unlink()
                            cleaned += 1
                        except:
                            pass
        
        self.logger.info(f"Очищено {cleaned} временных файлов")
        return cleaned
    
    def backup_database(self) -> str:
        """Создание резервной копии базы данных"""
        backup_dir = Path(self.config.backup_dir)
        backup_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = backup_dir / f"database_backup_{timestamp}.db"
        
        shutil.copy2(self.config.db_path, backup_file)
        
        # Сжимаем бэкап
        compressed_file = backup_dir / f"database_backup_{timestamp}.db.gz"
        with open(backup_file, 'rb') as f_in:
            with gzip.open(compressed_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        
        # Удаляем несжатый файл
        backup_file.unlink()
        
        # Удаляем старые бэкапы (оставляем последние 10)
        all_backups = sorted(backup_dir.glob("database_backup_*.db.gz"))
        if len(all_backups) > 10:
            for old_backup in all_backups[:-10]:
                old_backup.unlink()
        
        self.logger.info(f"Создан бэкап базы данных: {compressed_file}")
        return str(compressed_file)
    
    def restore_database(self, backup_file: str) -> bool:
        """Восстановление базы данных из бэкапа"""
        try:
            # Создаем бэкап текущей базы перед восстановлением
            current_backup = self.backup_database()
            
            # Восстанавливаем из бэкапа
            if backup_file.endswith('.gz'):
                with gzip.open(backup_file, 'rb') as f_in:
                    with open(self.config.db_path, 'wb') as f_out:
                        shutil.copyfileobj(f_in, f_out)
            else:
                shutil.copy2(backup_file, self.config.db_path)
            
            # Пересоздаем подключение к БД
            self.db = Database(self.config.db_path)
            
            self.logger.success(f"База данных восстановлена из {backup_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка восстановления базы данных: {e}")
            return False
    
    def get_system_health(self) -> Dict:
        """Проверка здоровья системы"""
        health = {
            'timestamp': datetime.now().isoformat(),
            'status': 'healthy',
            'components': {}
        }
        
        # Проверка БД
        try:
            db_stats = self.db.get_database_stats()
            health['components']['database'] = {
                'status': 'healthy',
                'size_mb': round(db_stats.get('db_size_mb', 0), 2),
                'tables': {k: v for k, v in db_stats.items() if k.endswith('_count')}
            }
        except Exception as e:
            health['components']['database'] = {
                'status': 'unhealthy',
                'error': str(e)
            }
            health['status'] = 'degraded'
        
        # Проверка Google Sheets
        try:
            product_stats = self.product_db.get_stats()
            health['components']['google_sheets'] = {
                'status': 'healthy',
                'products_count': product_stats.get('total_products', 0)
            }
        except Exception as e:
            health['components']['google_sheets'] = {
                'status': 'unhealthy',
                'error': str(e)
            }
            health['status'] = 'degraded'
        
        # Проверка Яндекс Маркета
        if self.config.yandex_token and self.config.campaign_id:
            try:
                yandex_ok, yandex_msg = self.yandex_client.test_connection()
                health['components']['yandex_market'] = {
                    'status': 'healthy' if yandex_ok else 'unhealthy',
                    'message': yandex_msg
                }
            except Exception as e:
                health['components']['yandex_market'] = {
                    'status': 'unhealthy',
                    'error': str(e)
                }
                health['status'] = 'degraded'
        
        # Проверка дискового пространства
        try:
            import shutil
            disk_usage = shutil.disk_usage(self.config.uploads_dir)
            free_gb = disk_usage.free / (1024 ** 3)
            total_gb = disk_usage.total / (1024 ** 3)
            
            health['components']['disk'] = {
                'status': 'healthy' if free_gb > 1 else 'warning',
                'free_gb': round(free_gb, 2),
                'total_gb': round(total_gb, 2),
                'usage_percent': round((1 - free_gb / total_gb) * 100, 2)
            }
            
            if free_gb < 1:
                health['status'] = 'degraded'
                health['warnings'] = health.get('warnings', []) + ['Мало свободного места на диске']
        except Exception as e:
            health['components']['disk'] = {
                'status': 'unknown',
                'error': str(e)
            }
        
        return health


# ===================================================================
# БЛОК 15: ИНТЕРФЕЙС STREAMLIT - РАСШИРЕННЫЙ КОНСТРУКТОР МАППИНГА
# С ПОДДЕРЖКОЙ ЛОКАЛЬНОЙ ЗАГРУЗКИ ФАЙЛОВ
# ===================================================================

def render_mapping_constructor():
    """Расширенный интерфейс конструктора маппинга колонок с поддержкой локальной загрузки"""
    st.subheader("🛠️ Конструктор маппинга колонок")
    
    st.markdown("""
    **Расширенные возможности:**
    1. 📤 **Локальная загрузка** — загрузите файл прайса с компьютера
    2. 📧 **Загрузка из почты** — загрузите файл из почты поставщика
    3. 🔗 **Загрузка по URL** — загрузите файл по ссылке
    4. 📂 **Образец из архива** — используйте ранее загруженный файл
    5. 🔍 Автоматическое определение колонок
    6. 🎯 Ручная настройка соответствия полей
    7. 📋 Использование готовых шаблонов
    8. ✅ Валидация маппинга перед сохранением
    9. 📜 Просмотр истории изменений
    """)
    
    # Инициализация конструктора маппинга
    if 'mapping_constructor' not in st.session_state:
        st.session_state.mapping_constructor = MappingConstructor(
            st.session_state.config,
            st.session_state.logger
        )
    
    mapping_constructor = st.session_state.mapping_constructor
    
    # Вкладки конструктора
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📤 Загрузка и настройка",
        "📧 Загрузка из почты",
        "🔍 Автоопределение",
        "📋 Шаблоны",
        "📜 История"
    ])
    
    # ===================================================================
    # ВКЛАДКА 1: ЗАГРУЗКА И НАСТРОЙКА (ЛОКАЛЬНАЯ ЗАГРУЗКА)
    # ===================================================================
    with tab1:
        st.markdown("## 📤 Загрузка прайса для настройки маппинга")
        st.markdown("Загрузите файл прайса с компьютера для настройки соответствия колонок.")
        
        # Выбор поставщика
        suppliers = st.session_state.config.suppliers
        supplier_names = [s.get('name', 'Unknown') for s in suppliers]
        
        if not supplier_names:
            st.warning("⚠️ Сначала добавьте поставщика в разделе 'Поставщики'")
            return
        
        selected_supplier = st.selectbox(
            "Выберите поставщика для настройки",
            supplier_names,
            key="mapping_supplier_select_local"
        )
        
        # Информация о текущем маппинге
        current_mapping = st.session_state.config.get_supplier_mapping(selected_supplier)
        mapping_stats = mapping_constructor.get_mapping_statistics(selected_supplier)
        
        # Отображаем статус маппинга
        if current_mapping and any(current_mapping.values()):
            with st.expander("✅ Текущий маппинг поставщика", expanded=True):
                status_color = "green" if mapping_stats['required_fields_mapped'] else "orange"
                st.markdown(f"**Статус:** 🔴 Обязательные поля не настроены" if not mapping_stats['required_fields_mapped'] else "**Статус:** 🟢 Все обязательные поля настроены")
                st.markdown(f"**Настроено полей:** {mapping_stats['mapped_fields_count']} из {mapping_stats['total_possible_fields']}")
                st.markdown(f"**Последнее обновление:** {mapping_stats.get('last_updated', 'Никогда')[:19] if mapping_stats.get('last_updated') else 'Никогда'}")
                
                st.markdown("---")
                st.markdown("**Соответствие колонок:**")
                
                col1, col2 = st.columns(2)
                mapped_items = [(target, source) for target, source in current_mapping.items() if source]
                
                for i, (target, source) in enumerate(mapped_items):
                    if i % 2 == 0:
                        with col1:
                            st.markdown(f"• **{target}** → `{source}`")
                    else:
                        with col2:
                            st.markdown(f"• **{target}** → `{source}`")
        else:
            st.info("ℹ️ Маппинг для этого поставщика еще не настроен")
        
        st.divider()
        
        # ===================================================================
        # СПОСОБЫ ЗАГРУЗКИ ФАЙЛА
        # ===================================================================
        st.subheader("📂 Выберите способ загрузки файла")
        
        load_method = st.radio(
            "Способ загрузки прайса:",
            [
                "📤 Загрузить с компьютера (локально)",
                "📂 Использовать образец из архива",
                "🔗 Загрузить по URL"
            ],
            key="load_method_local",
            horizontal=True
        )
        
        file_content = None
        file_filename = None
        file_loaded = False
        
        # ===================================================================
        # СПОСОБ 1: ЗАГРУЗКА С КОМПЬЮТЕРА
        # ===================================================================
        if load_method == "📤 Загрузить с компьютера (локально)":
            st.markdown("### 📤 Загрузка файла с компьютера")
            st.markdown("Поддерживаемые форматы: **XLSX, XLS, CSV, XML, JSON, TXT, ODS**")
            
            uploaded_file = st.file_uploader(
                "Выберите файл прайса",
                type=['xlsx', 'xls', 'csv', 'xml', 'json', 'txt', 'xlsm', 'ods', 'xlsb'],
                help="Загрузите файл прайса с вашего компьютера",
                key="local_file_uploader"
            )
            
            if uploaded_file is not None:
                file_content = uploaded_file.read()
                file_filename = uploaded_file.name
                file_loaded = True
                
                # Сохраняем файл локально для истории
                try:
                    saved_path = mapping_constructor.save_uploaded_file(file_content, file_filename)
                    st.success(f"✅ Файл сохранен: `{saved_path}`")
                except Exception as e:
                    st.warning(f"⚠️ Не удалось сохранить файл локально: {e}")
        
        # ===================================================================
        # СПОСОБ 2: ИСПОЛЬЗОВАТЬ ОБРАЗЕЦ ИЗ АРХИВА
        # ===================================================================
        elif load_method == "📂 Использовать образец из архива":
            st.markdown("### 📂 Образец из архива поставщика")
            st.markdown("Будет загружен последний сохраненный файл от этого поставщика.")
            
            if st.button("🔍 Найти образец в архиве", key="find_sample_btn", use_container_width=True):
                with st.spinner("Поиск файлов в архиве..."):
                    sample = mapping_constructor.load_sample_from_supplier(selected_supplier)
                    
                    if sample:
                        file_content, file_filename = sample
                        file_loaded = True
                        st.success(f"✅ Найден файл: `{file_filename}` ({len(file_content) / 1024:.1f} KB)")
                    else:
                        st.warning(f"⚠️ Не найдено сохраненных файлов для поставщика '{selected_supplier}'")
                        st.info("💡 Сначала загрузите файл через почту или локально, чтобы он появился в архиве.")
            
            # Показываем список доступных файлов
            with st.expander("📋 Доступные файлы в архиве"):
                archive_files = mapping_constructor.get_uploaded_files_list()
                
                if archive_files:
                    st.markdown(f"Найдено файлов: {len(archive_files)}")
                    
                    for file_info in archive_files[:20]:  # Показываем последние 20
                        col1, col2, col3 = st.columns([3, 2, 1])
                        
                        with col1:
                            st.markdown(f"📄 `{file_info['filename']}`")
                        
                        with col2:
                            st.markdown(f"📏 {file_info['size_formatted']}")
                        
                        with col3:
                            if st.button("📥 Загрузить", key=f"load_{file_info['filename'][:20]}", use_container_width=True):
                                try:
                                    loaded_content, loaded_filename = mapping_constructor.load_file_from_local(
                                        file_info['filepath']
                                    )
                                    file_content = loaded_content
                                    file_filename = loaded_filename
                                    file_loaded = True
                                    st.success(f"✅ Файл загружен!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Ошибка: {e}")
                else:
                    st.info("Нет сохраненных файлов")
        
        # ===================================================================
        # СПОСОБ 3: ЗАГРУЗКА ПО URL
        # ===================================================================
        elif load_method == "🔗 Загрузить по URL":
            st.markdown("### 🔗 Загрузка файла по ссылке")
            st.markdown("Введите прямую ссылку на файл прайса.")
            
            url_input = st.text_input(
                "URL файла",
                placeholder="https://example.com/pricelist.xlsx",
                key="url_input_mapping"
            )
            
            custom_filename = st.text_input(
                "Имя файла (опционально)",
                placeholder="Если не указано, будет извлечено из URL",
                key="custom_filename_mapping"
            )
            
            if st.button("🔗 Загрузить по URL", key="load_url_btn", use_container_width=True):
                if url_input:
                    with st.spinner("Загрузка файла по URL..."):
                        try:
                            filename = custom_filename if custom_filename else None
                            file_content, file_filename = mapping_constructor.load_file_from_url(
                                url_input, 
                                filename
                            )
                            file_loaded = True
                            st.success(f"✅ Файл загружен: `{file_filename}` ({len(file_content) / 1024:.1f} KB)")
                        except Exception as e:
                            st.error(f"❌ Ошибка загрузки: {str(e)}")
                else:
                    st.error("❌ Введите URL файла")
        
        st.divider()
        
        # ===================================================================
        # ОБРАБОТКА ЗАГРУЖЕННОГО ФАЙЛА
        # ===================================================================
        if file_loaded and file_content is not None and file_filename is not None:
            try:
                # Получаем предпросмотр
                with st.spinner("Анализ файла..."):
                    df, columns, metadata = mapping_constructor.preview_file(file_content, file_filename)
                
                if df.empty:
                    st.warning("⚠️ Не удалось распарсить файл или файл пуст")
                    st.stop()
                
                # ===================================================================
                # ИНФОРМАЦИЯ О ФАЙЛЕ
                # ===================================================================
                with st.expander("📊 Информация о файле", expanded=True):
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("📄 Строк", metadata.get('row_count', 0))
                    
                    with col2:
                        st.metric("📊 Колонок", metadata.get('column_count', 0))
                    
                    with col3:
                        st.metric("💾 Размер", metadata.get('file_size_formatted', '0 KB'))
                    
                    with col4:
                        st.metric("📝 Кодировка", metadata.get('encoding', 'unknown'))
                    
                    # Типы колонок
                    if metadata.get('column_types'):
                        st.markdown("**Типы колонок (автоопределение):**")
                        type_cols = st.columns(3)
                        for i, (col, col_type) in enumerate(metadata['column_types'].items()):
                            type_emoji = {
                                'integer': '🔢',
                                'decimal': '💲',
                                'short_text': '📝',
                                'text': '📄',
                                'long_text': '📑',
                                'empty': '🫗'
                            }.get(col_type, '❓')
                            
                            with type_cols[i % 3]:
                                st.markdown(f"{type_emoji} **{col}**: *{col_type}*")
                
                # ===================================================================
                # ПРЕДПРОСМОТР ДАННЫХ
                # ===================================================================
                st.subheader("📋 Предпросмотр данных")
                
                preview_rows = st.slider(
                    "Количество строк для предпросмотра",
                    min_value=5,
                    max_value=100,
                    value=20,
                    step=5,
                    key="preview_rows_slider"
                )
                
                st.dataframe(
                    df.head(preview_rows),
                    use_container_width=True,
                    height=400
                )
                
                # Информация о колонках
                with st.expander("📊 Статистика по колонкам"):
                    col_stats = []
                    for col in df.columns:
                        col_stats.append({
                            'Колонка': col,
                            'Тип': metadata['column_types'].get(col, 'unknown'),
                            'Непустых': len(df) - metadata['null_counts'].get(col, 0),
                            'Пустых': metadata['null_counts'].get(col, 0),
                            'Уникальных': metadata['unique_counts'].get(col, 0),
                            'Пример': ', '.join([str(x) for x in metadata['sample_values'].get(col, [])[:3]])
                        })
                    
                    st.dataframe(
                        pd.DataFrame(col_stats),
                        use_container_width=True,
                        hide_index=True
                    )
                
                # ===================================================================
                # НАСТРОЙКА МАППИНГА
                # ===================================================================
                st.subheader("🎯 Настройка маппинга колонок")
                st.markdown("Выберите, какая колонка в файле соответствует каждому полю:")
                
                # Список доступных колонок для выбора
                columns_options = [''] + df.columns.tolist()
                
                # Функция для получения индекса колонки в списке
                def get_column_index(column_name: str) -> int:
                    if column_name and column_name in columns_options:
                        return columns_options.index(column_name)
                    return 0
                
                # ===================================================================
                # ОБЯЗАТЕЛЬНЫЕ ПОЛЯ
                # ===================================================================
                with st.expander("📌 Обязательные поля", expanded=True):
                    st.markdown("*Поля, отмеченные звёздочкой (*), обязательны для заполнения*")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        sku_col = st.selectbox(
                            "📌 Артикул (SKU) *",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('sku', '')),
                            key="mapping_sku_local",
                            help="Уникальный идентификатор товара. Обязательное поле."
                        )
                    
                    with col2:
                        price_col = st.selectbox(
                            "💰 Цена *",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('price', '')),
                            key="mapping_price_local",
                            help="Цена товара. Обязательное поле."
                        )
                
                # ===================================================================
                # ОСНОВНЫЕ ПОЛЯ
                # ===================================================================
                with st.expander("📦 Основные поля", expanded=True):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        stock_col = st.selectbox(
                            "📦 Остаток",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('stock', '')),
                            key="mapping_stock_local",
                            help="Количество товара на складе"
                        )
                        
                        brand_col = st.selectbox(
                            "🏷️ Бренд",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('brand', '')),
                            key="mapping_brand_local",
                            help="Производитель или бренд товара"
                        )
                        
                        category_col = st.selectbox(
                            "📂 Категория",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('category', '')),
                            key="mapping_category_local",
                            help="Категория или раздел товара"
                        )
                    
                    with col2:
                        name_col = st.selectbox(
                            "📝 Название",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('name', '')),
                            key="mapping_name_local",
                            help="Наименование товара"
                        )
                        
                        description_col = st.selectbox(
                            "📄 Описание",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('description', '')),
                            key="mapping_description_local",
                            help="Описание товара"
                        )
                        
                        barcode_col = st.selectbox(
                            "🔢 Штрихкод",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('barcode', '')),
                            key="mapping_barcode_local",
                            help="Штрихкод или EAN товара"
                        )
                
                # ===================================================================
                # ДОПОЛНИТЕЛЬНЫЕ ПОЛЯ
                # ===================================================================
                with st.expander("📋 Дополнительные поля", expanded=False):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        weight_col = st.selectbox(
                            "⚖️ Вес",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('weight', '')),
                            key="mapping_weight_local",
                            help="Вес товара"
                        )
                        
                        country_col = st.selectbox(
                            "🌍 Страна",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('country', '')),
                            key="mapping_country_local",
                            help="Страна производства"
                        )
                        
                        warranty_col = st.selectbox(
                            "🛡️ Гарантия",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('warranty', '')),
                            key="mapping_warranty_local",
                            help="Срок гарантии"
                        )
                        
                        currency_col = st.selectbox(
                            "💱 Валюта",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('currency', '')),
                            key="mapping_currency_local",
                            help="Валюта цены"
                        )
                    
                    with col2:
                        dimensions_col = st.selectbox(
                            "📐 Размеры",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('dimensions', '')),
                            key="mapping_dimensions_local",
                            help="Габариты товара"
                        )
                        
                        min_order_col = st.selectbox(
                            "📦 Мин. заказ",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('min_order_qty', '')),
                            key="mapping_min_order_local",
                            help="Минимальное количество для заказа"
                        )
                        
                        vat_col = st.selectbox(
                            "🧾 НДС",
                            options=columns_options,
                            index=get_column_index(current_mapping.get('vat', '')),
                            key="mapping_vat_local",
                            help="Ставка НДС"
                        )
                
                # ===================================================================
                # ПРОВЕРКА ОБЯЗАТЕЛЬНЫХ ПОЛЕЙ
                # ===================================================================
                if not sku_col or not price_col:
                    st.error("❌ **Артикул (SKU)** и **Цена** — обязательные поля. Выберите соответствующие колонки.")
                else:
                    # ===================================================================
                    # ФОРМИРОВАНИЕ МАППИНГА
                    # ===================================================================
                    new_mapping = {
                        'sku': sku_col,
                        'price': price_col,
                        'stock': stock_col if stock_col else '',
                        'brand': brand_col if brand_col else '',
                        'name': name_col if name_col else '',
                        'category': category_col if category_col else '',
                        'description': description_col if description_col else '',
                        'weight': weight_col if weight_col else '',
                        'barcode': barcode_col if barcode_col else '',
                        'country': country_col if country_col else '',
                        'warranty': warranty_col if warranty_col else '',
                        'dimensions': dimensions_col if dimensions_col else '',
                        'min_order_qty': min_order_col if min_order_col else '',
                        'currency': currency_col if currency_col else '',
                        'vat': vat_col if vat_col else ''
                    }
                    
                    # Удаляем пустые поля
                    new_mapping = {k: v for k, v in new_mapping.items() if v}
                    
                    # ===================================================================
                    # ВАЛИДАЦИЯ МАППИНГА
                    # ===================================================================
                    validation = mapping_constructor.validate_mapping(df, new_mapping)
                    
                    if not validation.is_valid:
                        st.error("❌ **Ошибки валидации маппинга:**")
                        for error in validation.errors:
                            st.error(f"• {error}")
                    
                    if validation.warnings:
                        st.warning("⚠️ **Предупреждения:**")
                        for warning in validation.warnings:
                            st.warning(f"• {warning}")
                    
                    # ===================================================================
                    # ПРЕДПРОСМОТР РЕЗУЛЬТАТА МАППИНГА
                    # ===================================================================
                    if validation.is_valid:
                        st.subheader("🔍 Результат маппинга (предпросмотр)")
                        
                        # Применяем маппинг
                        result_df = mapping_constructor.parser._apply_mapping(df.copy(), new_mapping)
                        
                        if not result_df.empty:
                            # Статистика результата
                            col1, col2, col3, col4 = st.columns(4)
                            
                            with col1:
                                st.metric("📦 Товаров", len(result_df))
                            
                            with col2:
                                valid_prices = (result_df['price'] > 0).sum()
                                st.metric("💰 С ценой", valid_prices)
                            
                            with col3:
                                with_stock = (result_df['stock'] > 0).sum()
                                st.metric("📦 В наличии", with_stock)
                            
                            with col4:
                                unique_skus = result_df['sku'].nunique()
                                st.metric("🔑 Уникальных SKU", unique_skus)
                            
                            # Таблица с результатом
                            st.dataframe(
                                result_df.head(20),
                                use_container_width=True,
                                height=400,
                                column_config={
                                    "sku": "Артикул",
                                    "price": st.column_config.NumberColumn("Цена", format="%.2f"),
                                    "stock": "Остаток",
                                    "brand": "Бренд",
                                    "name": "Название",
                                    "category": "Категория",
                                    "description": "Описание"
                                }
                            )
                            
                            # ===================================================================
                            # СРАВНЕНИЕ С ТЕКУЩИМ МАППИНГОМ
                            # ===================================================================
                            if current_mapping and any(current_mapping.values()):
                                comparison = mapping_constructor.compare_mappings(current_mapping, new_mapping)
                                
                                if comparison['has_changes']:
                                    with st.expander("🔄 Изменения относительно текущего маппинга", expanded=False):
                                        st.markdown(f"**Всего изменений:** {comparison['total_changes']}")
                                        st.markdown(f"• Добавлено полей: {comparison['added_fields']}")
                                        st.markdown(f"• Удалено полей: {comparison['removed_fields']}")
                                        st.markdown(f"• Изменено полей: {comparison['modified_fields']}")
                                        
                                        for change in comparison['changes']:
                                            emoji = {
                                                'added': '➕',
                                                'removed': '➖',
                                                'modified': '✏️'
                                            }.get(change['change_type'], '•')
                                            
                                            st.markdown(
                                                f"{emoji} **{change['field']}**: "
                                                f"`{change['old_value']}` → `{change['new_value']}`"
                                            )
                        
                        # ===================================================================
                        # СОХРАНЕНИЕ МАППИНГА
                        # ===================================================================
                        st.divider()
                        st.subheader("💾 Сохранение маппинга")
                        
                        col1, col2, col3 = st.columns([2, 1, 1])
                        
                        with col1:
                            save_comment = st.text_input(
                                "Комментарий к сохранению",
                                placeholder="Например: Обновлены колонки после нового прайса",
                                key="save_comment_local"
                            )
                        
                        with col2:
                            if st.button("💾 Сохранить маппинг", type="primary", use_container_width=True, key="save_mapping_local_btn"):
                                if mapping_constructor.save_mapping(
                                    selected_supplier,
                                    new_mapping,
                                    created_by='user',
                                    comment=save_comment
                                ):
                                    st.success(f"✅ Маппинг сохранен для поставщика **{selected_supplier}**!")
                                    st.balloons()
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("❌ Ошибка сохранения маппинга")
                        
                        with col3:
                            if current_mapping and any(current_mapping.values()):
                                if st.button("🔄 Сбросить", type="secondary", use_container_width=True, key="reset_mapping_local_btn"):
                                    if st.session_state.get(f"confirm_reset_local", False):
                                        if mapping_constructor.save_mapping(
                                            selected_supplier,
                                            {},
                                            created_by='user',
                                            comment='Маппинг сброшен'
                                        ):
                                            st.success("✅ Маппинг сброшен")
                                            st.session_state[f"confirm_reset_local"] = False
                                            time.sleep(1)
                                            st.rerun()
                                    else:
                                        st.session_state[f"confirm_reset_local"] = True
                                        st.warning("⚠️ Нажмите еще раз для подтверждения сброса")
            
            except Exception as e:
                st.error(f"❌ Ошибка обработки файла: {str(e)}")
                st.exception(e)
    
    # ===================================================================
    # ВКЛАДКА 2: ЗАГРУЗКА ИЗ ПОЧТЫ ПОСТАВЩИКА
    # ===================================================================
    with tab2:
        st.markdown("## 📧 Загрузка прайса из почты поставщика")
        st.markdown("Загрузите файл прайса напрямую из почтового ящика поставщика.")
        
        # Выбор поставщика с настроенной почтой
        email_suppliers = mapping_constructor.get_available_suppliers_for_email()
        
        if not email_suppliers:
            st.warning("⚠️ Нет поставщиков с настроенной почтой. Добавьте email и пароль в разделе 'Поставщики'.")
            st.info("💡 Вы можете использовать локальную загрузку файла на вкладке '📤 Загрузка и настройка'.")
        else:
            email_supplier = st.selectbox(
                "Выберите поставщика",
                email_suppliers,
                key="email_supplier_select"
            )
            
            max_emails = st.slider(
                "Максимальное количество писем для проверки",
                min_value=1,
                max_value=50,
                value=10,
                key="max_emails_slider"
            )
            
            if st.button("📧 Загрузить файлы из почты", type="primary", use_container_width=True, key="load_from_email_btn"):
                with st.spinner(f"Подключение к почте поставщика {email_supplier}..."):
                    try:
                        files = mapping_constructor.load_file_from_email(email_supplier, max_emails)
                        
                        if not files:
                            st.warning(f"⚠️ Не найдено файлов в почте поставщика {email_supplier}")
                        else:
                            st.success(f"✅ Загружено файлов: {len(files)}")
                            
                            # Показываем список загруженных файлов
                            for i, (content, filename, subject) in enumerate(files):
                                with st.expander(f"📄 {filename} ({len(content) / 1024:.1f} KB)", expanded=(i == 0)):
                                    st.markdown(f"**Из письма:** {subject[:100]}")
                                    st.markdown(f"**Размер:** {len(content) / 1024:.1f} KB")
                                    
                                    if st.button("📥 Использовать для настройки", key=f"use_email_file_{i}", use_container_width=True):
                                        # Переключаемся на вкладку локальной загрузки
                                        st.session_state['email_file_content'] = content
                                        st.session_state['email_file_filename'] = filename
                                        st.success(f"✅ Файл `{filename}` выбран для настройки. Перейдите на вкладку '📤 Загрузка и настройка'.")
                                        st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Ошибка загрузки из почты: {str(e)}")
    
    # ===================================================================
    # ВКЛАДКА 3: АВТООПРЕДЕЛЕНИЕ
    # ===================================================================
    with tab3:
        st.subheader("🔍 Автоматическое определение маппинга")
        st.markdown("""
        Загрузите файл, и система автоматически определит соответствие колонок
        на основе анализа содержимого и названий колонок.
        """)
        
        auto_file = st.file_uploader(
            "📤 Загрузите файл для автоопределения",
            type=['xlsx', 'xls', 'csv', 'xml', 'json', 'txt', 'ods'],
            key="auto_mapping_file_uploader"
        )
        
        if auto_file is not None:
            try:
                content = auto_file.read()
                filename = auto_file.name
                
                # Сохраняем файл
                mapping_constructor.save_uploaded_file(content, filename)
                
                # Парсим файл
                with st.spinner("Анализ файла..."):
                    df = mapping_constructor.parser.parse(filename, content)
                
                if not df.empty:
                    st.subheader("📊 Предпросмотр файла")
                    st.dataframe(df.head(10), use_container_width=True)
                    
                    # Получаем предложения
                    with st.spinner("Автоопределение колонок..."):
                        suggestions = mapping_constructor.get_mapping_suggestions(df)
                    
                    if suggestions:
                        st.subheader("💡 Предложенный маппинг")
                        
                        # Отображаем предложения
                        suggestion_data = []
                        for s in suggestions:
                            confidence_emoji = (
                                '🟢' if s['confidence'] >= 80 else
                                '🟡' if s['confidence'] >= 50 else
                                '🟠' if s['confidence'] >= 30 else '🔴'
                            )
                            suggestion_data.append({
                                'Поле': s['target'],
                                'Колонка в файле': s['column'],
                                'Уверенность': f"{confidence_emoji} {s['confidence']}%",
                                'Уровень': s['confidence_level'],
                                'Причины': '; '.join(s['reasons'])
                            })
                        
                        suggestion_df = pd.DataFrame(suggestion_data)
                        st.dataframe(suggestion_df, use_container_width=True, hide_index=True)
                        
                        # Визуализация уверенности
                        st.subheader("📊 Уверенность определения")
                        
                        fig = go.Figure()
                        fig.add_trace(go.Bar(
                            x=[s['target'] for s in suggestions],
                            y=[s['confidence'] for s in suggestions],
                            text=[f"{s['confidence']}%" for s in suggestions],
                            textposition='outside',
                            marker_color=[
                                'green' if s['confidence'] >= 80 else
                                'yellow' if s['confidence'] >= 50 else
                                'orange' if s['confidence'] >= 30 else 'red'
                                for s in suggestions
                            ]
                        ))
                        fig.update_layout(
                            title='Уверенность автоопределения по полям',
                            xaxis_title='Поле',
                            yaxis_title='Уверенность (%)',
                            yaxis_range=[0, 105],
                            height=400
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Кнопка применения
                        if st.button("✅ Применить предложенный маппинг", type="primary", use_container_width=True):
                            auto_mapping = {s['target']: s['column'] for s in suggestions}
                            
                            if mapping_constructor.save_mapping(
                                selected_supplier,
                                auto_mapping,
                                created_by='auto',
                                comment='Автоматически определенный маппинг'
                            ):
                                st.success("✅ Автоматический маппинг применен!")
                                st.balloons()
                                time.sleep(1)
                                st.rerun()
                    else:
                        st.warning("⚠️ Не удалось автоматически определить маппинг")
                        st.info("💡 Попробуйте настроить маппинг вручную на вкладке '📤 Загрузка и настройка'")
            
            except Exception as e:
                st.error(f"❌ Ошибка: {str(e)}")
    
    # ===================================================================
    # ВКЛАДКА 4: ШАБЛОНЫ
    # ===================================================================
    with tab4:
        st.subheader("📋 Шаблоны маппинга")
        st.markdown("Используйте готовые шаблоны для быстрой настройки маппинга.")
        
        # Отображаем доступные шаблоны
        templates = mapping_constructor.mapping_templates
        template_names = list(templates.keys())
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            selected_template = st.selectbox(
                "Выберите шаблон",
                template_names,
                key="template_select_mapping"
            )
        
        if selected_template:
            template = templates[selected_template]
            
            # Описание шаблона
            template_descriptions = {
                'standard': 'Стандартный набор полей для большинства поставщиков',
                'minimal': 'Минимальный набор (только обязательные поля)',
                'extended': 'Расширенный набор со всеми возможными полями',
                'yandex_market': 'Формат Яндекс Маркета',
                'ozon': 'Формат OZON',
                'wildberries': 'Формат Wildberries',
                '1c_export': 'Формат выгрузки из 1С'
            }
            
            description = template_descriptions.get(selected_template, 'Пользовательский шаблон')
            st.markdown(f"**{description}**")
            
            st.write("**Поля шаблона:**")
            template_data = []
            for field, column in template.items():
                template_data.append({
                    'Поле': field,
                    'Типовая колонка': column,
                    'Обязательное': '✅' if field in ['sku', 'price'] else '❌'
                })
            
            st.dataframe(
                pd.DataFrame(template_data),
                use_container_width=True,
                hide_index=True
            )
            
            with col2:
                if st.button("📥 Применить шаблон", use_container_width=True, key="apply_template_btn"):
                    if mapping_constructor.save_mapping(
                        selected_supplier,
                        template,
                        created_by='template',
                        comment=f'Применен шаблон: {selected_template}'
                    ):
                        st.success(f"✅ Шаблон '{selected_template}' применен!")
                        st.balloons()
                        time.sleep(1)
                        st.rerun()
        
        # Создание нового шаблона
        with st.expander("➕ Создать новый шаблон", expanded=False):
            new_template_name = st.text_input("Название нового шаблона", key="new_template_name")
            
            st.write("Заполните поля шаблона (типовые названия колонок):")
            
            col1, col2 = st.columns(2)
            
            with col1:
                t_sku = st.text_input("SKU", value="Артикул", key="t_sku")
                t_price = st.text_input("Цена", value="Цена", key="t_price")
                t_stock = st.text_input("Остаток", value="Остаток", key="t_stock")
                t_brand = st.text_input("Бренд", value="Бренд", key="t_brand")
                t_name = st.text_input("Название", value="Название", key="t_name")
                t_category = st.text_input("Категория", value="Категория", key="t_category")
            
            with col2:
                t_description = st.text_input("Описание", value="Описание", key="t_description")
                t_barcode = st.text_input("Штрихкод", value="Штрихкод", key="t_barcode")
                t_weight = st.text_input("Вес", value="Вес", key="t_weight")
                t_country = st.text_input("Страна", value="Страна", key="t_country")
                t_warranty = st.text_input("Гарантия", value="Гарантия", key="t_warranty")
                t_dimensions = st.text_input("Размеры", value="Размеры", key="t_dimensions")
            
            if st.button("💾 Сохранить шаблон", use_container_width=True, key="save_template_btn"):
                if new_template_name:
                    new_template = {
                        'sku': t_sku,
                        'price': t_price,
                        'stock': t_stock,
                        'brand': t_brand,
                        'name': t_name,
                        'category': t_category,
                        'description': t_description,
                        'barcode': t_barcode,
                        'weight': t_weight,
                        'country': t_country,
                        'warranty': t_warranty,
                        'dimensions': t_dimensions
                    }
                    
                    mapping_constructor.mapping_templates[new_template_name] = new_template
                    mapping_constructor.save_templates()
                    st.success(f"✅ Шаблон '{new_template_name}' сохранен!")
                    st.rerun()
                else:
                    st.error("❌ Введите название шаблона")
        
        # Управление шаблонами
        with st.expander("🗑️ Управление пользовательскими шаблонами", expanded=False):
            system_templates = ['standard', 'minimal', 'extended', 'yandex_market', 'ozon', 'wildberries', '1c_export']
            custom_templates = {k: v for k, v in templates.items() if k not in system_templates}
            
            if custom_templates:
                template_to_delete = st.selectbox(
                    "Выберите шаблон для удаления",
                    list(custom_templates.keys()),
                    key="delete_template_select"
                )
                
                if st.button("🗑️ Удалить шаблон", type="secondary", use_container_width=True, key="delete_template_btn"):
                    if template_to_delete in mapping_constructor.mapping_templates:
                        del mapping_constructor.mapping_templates[template_to_delete]
                        mapping_constructor.save_templates()
                        st.success(f"✅ Шаблон '{template_to_delete}' удален")
                        st.rerun()
            else:
                st.info("Нет пользовательских шаблонов")
    
    # ===================================================================
    # ВКЛАДКА 5: ИСТОРИЯ
    # ===================================================================
    with tab5:
        st.subheader("📜 История изменений маппинга")
        
        if selected_supplier:
            history = st.session_state.config.get_mapping_history(selected_supplier)
            
            if history:
                st.markdown(f"**История изменений для поставщика: {selected_supplier}**")
                st.markdown(f"**Всего версий:** {len(history)}")
                
                for record in reversed(history):
                    with st.expander(
                        f"📝 Версия {record.get('version', '?')} — "
                        f"{record.get('created_at', '')[:19]} "
                        f"({record.get('created_by', 'system')})",
                        expanded=(record == history[-1])  # Последняя версия раскрыта
                    ):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown(f"**Автор:** {record.get('created_by', 'system')}")
                        
                        with col2:
                            st.markdown(f"**Дата:** {record.get('created_at', '')[:19]}")
                        
                        if record.get('comment'):
                            st.markdown(f"**Комментарий:** {record['comment']}")
                        
                        if record.get('changes'):
                            st.markdown("**Изменения в этой версии:**")
                            for change in record['changes']:
                                emoji = {
                                    'added': '➕',
                                    'removed': '➖',
                                    'modified': '✏️'
                                }.get(change.get('change_type', 'modified'), '•')
                                
                                st.markdown(
                                    f"{emoji} **{change['field']}**: "
                                    f"`{change['old_value']}` → `{change['new_value']}`"
                                )
                        
                        if record.get('mapping'):
                            st.markdown("**Итоговый маппинг:**")
                            mapped_fields = {k: v for k, v in record['mapping'].items() if v}
                            
                            if mapped_fields:
                                mapping_df = pd.DataFrame(
                                    [{'Поле': k, 'Колонка': v} for k, v in mapped_fields.items()]
                                )
                                st.dataframe(mapping_df, use_container_width=True, hide_index=True)
                            else:
                                st.info("Маппинг пуст")
            else:
                st.info(f"История изменений для поставщика '{selected_supplier}' пуста")
                st.markdown("""
                💡 **История появляется при сохранении маппинга.**
                
                Попробуйте:
                1. Загрузить файл прайса
                2. Настроить маппинг
                3. Сохранить изменения
                """)
        
        # Экспорт/импорт маппинга
        st.divider()
        st.subheader("📥 Экспорт / Импорт маппинга")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Экспорт маппинга:**")
            
            export_format = st.selectbox(
                "Формат экспорта",
                ['json', 'csv'],
                key="export_format_select"
            )
            
            if st.button("📤 Экспортировать", use_container_width=True, key="export_mapping_btn"):
                if current_mapping and any(current_mapping.values()):
                    filepath = mapping_constructor.export_mapping_config(selected_supplier, export_format)
                    if filepath:
                        with open(filepath, 'rb') as f:
                            file_data = f.read()
                        
                        mime_type = "application/json" if export_format == 'json' else "text/csv"
                        
                        st.download_button(
                            f"📥 Скачать {export_format.upper()}",
                            file_data,
                            f"mapping_{selected_supplier}.{export_format}",
                            mime_type,
                            key="download_exported_mapping"
                        )
                else:
                    st.warning("⚠️ Нет маппинга для экспорта")
        
        with col2:
            st.markdown("**Импорт маппинга:**")
            
            imported_file = st.file_uploader(
                "Загрузите файл маппинга",
                type=['json', 'csv'],
                key="import_mapping_file_uploader"
            )
            
            if imported_file is not None:
                # Сохраняем во временный файл
                temp_dir = Path(st.session_state.config.temp_dir)
                temp_dir.mkdir(parents=True, exist_ok=True)
                temp_path = temp_dir / f"import_mapping_{uuid.uuid4().hex}.{imported_file.name.split('.')[-1]}"
                
                with open(temp_path, 'wb') as f:
                    f.write(imported_file.read())
                
                imported_mapping = mapping_constructor.import_mapping_config(str(temp_path))
                
                if imported_mapping:
                    st.success("✅ Маппинг успешно импортирован!")
                    
                    st.markdown("**Импортированные поля:**")
                    import_df = pd.DataFrame(
                        [{'Поле': k, 'Колонка': v} for k, v in imported_mapping.items()]
                    )
                    st.dataframe(import_df, use_container_width=True, hide_index=True)
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        if st.button("✅ Применить импортированный маппинг", type="primary", use_container_width=True, key="apply_imported_mapping_btn"):
                            if mapping_constructor.save_mapping(
                                selected_supplier,
                                imported_mapping,
                                created_by='import',
                                comment=f'Импортирован из файла {imported_file.name}'
                            ):
                                st.success("✅ Маппинг применен!")
                                st.balloons()
                                time.sleep(1)
                                st.rerun()
                    
                    with col2:
                        if st.button("❌ Отменить", use_container_width=True, key="cancel_import_mapping_btn"):
                            st.rerun()
                else:
                    st.error("❌ Не удалось импортировать маппинг из файла")
                
                # Удаляем временный файл
                try:
                    temp_path.unlink(missing_ok=True)
                except:
                    pass
        
        # Очистка старых загрузок
        st.divider()
        with st.expander("🧹 Очистка временных файлов", expanded=False):
            uploaded_files = mapping_constructor.get_uploaded_files_list()
            
            if uploaded_files:
                st.markdown(f"**Временных файлов:** {len(uploaded_files)}")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    max_age = st.number_input(
                        "Максимальный возраст (часов)",
                        value=24,
                        min_value=1,
                        max_value=168,
                        key="max_age_hours_input"
                    )
                
                with col2:
                    if st.button("🧹 Очистить старые файлы", use_container_width=True, key="cleanup_uploads_btn"):
                        cleaned = mapping_constructor.cleanup_old_uploads(max_age)
                        st.success(f"✅ Очищено {cleaned} файлов")
                        st.rerun()
            else:
                st.info("Нет временных файлов")


# ===================================================================
# БЛОК 16: ИНТЕРФЕЙС STREAMLIT - ОСНОВНОЙ
# ===================================================================

def init_session_state() -> None:
    """Расширенная инициализация состояния сессии"""
    if 'logger' not in st.session_state:
        st.session_state.logger = StreamlitLogger()
    
    if 'config' not in st.session_state:
        st.session_state.config = Config.load()
    
    if 'product_db' not in st.session_state:
        try:
            st.session_state.product_db = GoogleSheetsDatabase(
                st.session_state.config,
                st.session_state.logger
            )
        except Exception as e:
            st.session_state.product_db = None
            st.error(f"Ошибка подключения к Google Sheets: {e}")
    
    if 'analysis_result' not in st.session_state:
        st.session_state.analysis_result = None
    
    if 'analyzer' not in st.session_state:
        st.session_state.analyzer = None
    
    if 'is_running' not in st.session_state:
        st.session_state.is_running = False
    
    if 'result' not in st.session_state:
        st.session_state.result = None
    
    if 'progress_message' not in st.session_state:
        st.session_state.progress_message = ""
    
    if 'progress_value' not in st.session_state:
        st.session_state.progress_value = 0
    
    if 'robot' not in st.session_state:
        st.session_state.robot = None
    
    if 'db' not in st.session_state:
        st.session_state.db = Database(st.session_state.config.db_path)
    
    if 'notification_manager' not in st.session_state:
        st.session_state.notification_manager = NotificationManager(
            st.session_state.config,
            st.session_state.logger
        )
    
    if 'selected_supplier' not in st.session_state:
        st.session_state.selected_supplier = None
    
    if 'auto_refresh' not in st.session_state:
        st.session_state.auto_refresh = False


def save_config() -> None:
    """Сохранение конфигурации"""
    try:
        st.session_state.config.save()
        st.success("✅ Конфигурация сохранена")
    except Exception as e:
        st.error(f"❌ Ошибка сохранения: {e}")


def run_robot() -> None:
    """Запуск робота"""
    if st.session_state.is_running:
        st.warning("⚠️ Робот уже запущен")
        return
    
    config = st.session_state.config
    
    if not config.email_user or not config.email_pass:
        st.error("❌ Заполните настройки почты")
        return
    
    if not config.yandex_token or config.campaign_id == 0:
        st.warning("⚠️ Не заполнены настройки Яндекс Маркета (будет пропущена отправка)")
    
    # Создаем робота
    st.session_state.robot = PriceRobot(config, st.session_state.logger)
    st.session_state.is_running = True
    st.session_state.result = None
    st.session_state.progress_value = 0
    st.session_state.logger.clear()
    
    def run_thread() -> None:
        """Фоновый поток выполнения"""
        try:
            result = st.session_state.robot.run_full_cycle(
                progress_callback=lambda msg: update_progress(msg)
            )
            st.session_state.result = result
        except Exception as e:
            st.session_state.logger.error(f"Ошибка в потоке: {str(e)}")
            st.session_state.result = {
                'status': 'failed',
                'errors': [str(e)],
                'files_found': 0,
                'files_processed': 0,
                'products_updated': 0,
                'products_added': 0,
                'offers_sent': 0,
                'duration': 0
            }
        finally:
            st.session_state.is_running = False
            st.session_state.progress_value = 100
    
    def update_progress(msg: str) -> None:
        """Обновление прогресса"""
        st.session_state.progress_message = msg
        # Увеличиваем прогресс
        if st.session_state.progress_value < 90:
            st.session_state.progress_value += 5
    
    thread = threading.Thread(target=run_thread, daemon=True)
    thread.start()


def run_analysis() -> None:
    """Запуск анализа прайсов"""
    try:
        config = st.session_state.config
        logger = st.session_state.logger
        
        with st.spinner("Загрузка прайсов от поставщиков..."):
            downloader = MultiSupplierDownloader(config, logger)
            downloaded_files = downloader.download_all_suppliers()
        
        if not downloaded_files:
            st.warning("⚠️ Не удалось загрузить ни одного прайса")
            return
        
        with st.spinner("Парсинг файлов..."):
            parser = PriceParser(config, logger)
            parsed_files = {}
            
            for file_data in downloaded_files:
                try:
                    filename = file_data.get('filename', 'unknown')
                    content = file_data.get('content', b'')
                    supplier_name = file_data.get('supplier', 'Unknown')
                    
                    if not content:
                        continue
                    
                    mapping = config.get_supplier_mapping(supplier_name)
                    
                    if mapping and any(mapping.values()):
                        df = parser.parse(filename, content, mapping)
                    else:
                        df = parser.parse(filename, content)
                    
                    if not df.empty:
                        df['supplier'] = supplier_name
                        df['supplier_email'] = file_data.get('supplier_email', '')
                        df['source_file'] = filename
                        parsed_files[filename] = df
                        
                except Exception as e:
                    logger.error(f"Ошибка парсинга {file_data.get('filename', 'unknown')}: {e}")
        
        if not parsed_files:
            st.warning("⚠️ Не удалось распарсить ни одного файла")
            return
        
        with st.spinner("Анализ цен..."):
            analyzer = PriceAnalyzer(config, logger)
            st.session_state.analyzer = analyzer
            result_df = analyzer.analyze(parsed_files)
        
        if result_df.empty:
            st.warning("⚠️ Анализ не дал результатов")
            return
        
        st.session_state.analysis_result = result_df
        
        supplier_count = len(set(f.get('supplier', 'Unknown') for f in downloaded_files))
        total_products = sum(len(df) for df in parsed_files.values())
        unique_skus = len(result_df)
        
        st.success(
            f"✅ Анализ завершен! Найдено {unique_skus} уникальных товаров "
            f"от {supplier_count} поставщиков"
        )
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Ошибка анализа: {str(e)}")
        st.session_state.logger.error(f"Ошибка анализа: {e}")


def render_supplier_settings() -> None:
    """Расширенный интерфейс управления поставщиками"""
    st.subheader("📋 Управление поставщиками")
    
    suppliers = st.session_state.config.suppliers
    
    # Статистика
    if suppliers:
        active_count = sum(1 for s in suppliers if s.get('enabled', True))
        with_mapping = sum(1 for s in suppliers if any(s.get('column_mapping', {}).values()))
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Всего поставщиков", len(suppliers))
        with col2:
            st.metric("Активных", active_count)
        with col3:
            st.metric("С настроенным маппингом", with_mapping)
    
    # Список поставщиков
    if suppliers:
        st.subheader("📋 Список поставщиков")
        
        for i, supplier in enumerate(suppliers):
            with st.expander(
                f"{'✅' if supplier.get('enabled', True) else '❌'} {supplier.get('name', 'Unknown')}",
                expanded=False
            ):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write(f"**Email:** {supplier.get('email', '')}")
                    st.write(f"**IMAP:** {supplier.get('imap_server', '')}:{supplier.get('imap_port', 993)}")
                    st.write(f"**Тема:** {supplier.get('subject_filter', '(все)')}")
                    st.write(f"**Отправитель:** {supplier.get('sender_filter', '(все)')}")
                
                with col2:
                    mapping = supplier.get('column_mapping', {})
                    if mapping and any(mapping.values()):
                        st.write("**Маппинг:**")
                        for field, col in mapping.items():
                            if col:
                                st.write(f"  • {field} → `{col}`")
                    else:
                        st.warning("⚠️ Маппинг не настроен")
                    
                    # Статистика
                    stats = supplier.get('stats', {})
                    if stats:
                        st.write(f"**Успешных синхронизаций:** {stats.get('total_files_processed', 0)}")
                        st.write(f"**Загружено товаров:** {stats.get('total_products_loaded', 0)}")
                        success_rate = stats.get('success_rate', 0)
                        st.progress(success_rate / 100, text=f"Успешность: {success_rate:.1f}%")
                
                # Действия
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("🔄 Синхронизировать", key=f"sync_{i}", use_container_width=True):
                        robot = PriceRobot(st.session_state.config, st.session_state.logger)
                        result = robot.run_single_supplier(supplier.get('name', ''))
                        if result['status'] == 'success':
                            st.success(f"✅ Синхронизировано файлов: {result.get('files_processed', 0)}")
                        else:
                            st.error(f"❌ Ошибка: {result.get('error', 'Неизвестная ошибка')}")
                
                with col2:
                    new_enabled = not supplier.get('enabled', True)
                    if st.button(
                        "🔴 Деактивировать" if supplier.get('enabled', True) else "🟢 Активировать",
                        key=f"toggle_{i}",
                        use_container_width=True
                    ):
                        supplier['enabled'] = new_enabled
                        st.session_state.config._save_suppliers()
                        st.rerun()
                
                with col3:
                    if st.button("🗑️ Удалить", key=f"delete_{i}", use_container_width=True):
                        if st.session_state.get(f"confirm_delete_{i}", False):
                            st.session_state.config.suppliers.pop(i)
                            st.session_state.config._save_suppliers()
                            st.success("✅ Поставщик удален")
                            st.rerun()
                        else:
                            st.session_state[f"confirm_delete_{i}"] = True
                            st.warning("⚠️ Нажмите еще раз для подтверждения удаления")
    
    # Добавление нового поставщика
    with st.expander("➕ Добавить нового поставщика", expanded=not suppliers):
        col1, col2 = st.columns(2)
        
        with col1:
            new_name = st.text_input("Название поставщика*", key="new_supplier_name")
            new_email = st.text_input("Email*", key="new_supplier_email")
            new_password = st.text_input("Пароль*", type="password", key="new_supplier_password")
            new_imap = st.text_input("IMAP сервер", value="imap.mail.ru", key="new_supplier_imap")
            new_port = st.number_input("IMAP порт", value=993, step=1, key="new_supplier_port")
        
        with col2:
            new_subject = st.text_input("Фильтр по теме", placeholder="прайс", key="new_supplier_subject")
            new_sender = st.text_input("Фильтр по отправителю", key="new_supplier_sender")
            new_priority = st.number_input("Приоритет", value=0, step=1, key="new_supplier_priority")
            new_markup = st.number_input("Индивидуальная наценка (%)", value=0.0, step=0.5, key="new_supplier_markup")
            new_enabled = st.checkbox("Активен", value=True, key="new_supplier_enabled")
        
        if st.button("✅ Добавить поставщика", type="primary", use_container_width=True):
            if not new_name or not new_email or not new_password:
                st.error("❌ Заполните обязательные поля (Название, Email, Пароль)")
            else:
                new_supplier = SupplierConfig(
                    name=new_name,
                    email=new_email,
                    email_password=new_password,
                    imap_server=new_imap,
                    imap_port=new_port,
                    subject_filter=new_subject,
                    sender_filter=new_sender,
                    enabled=new_enabled,
                    priority=new_priority,
                    custom_markup=new_markup
                )
                
                st.session_state.config.suppliers.append(new_supplier.to_dict())
                st.session_state.config._save_suppliers()
                st.success(f"✅ Поставщик {new_name} добавлен!")
                st.rerun()


def render_analysis_tab() -> None:
    """Расширенный интерфейс аналитики"""
    st.subheader("📊 Аналитика прайсов поставщиков")
    
    if st.button("🚀 Запустить анализ", type="primary", use_container_width=True):
        if not st.session_state.config.suppliers:
            st.error("❌ Сначала добавьте поставщиков в настройках")
        else:
            run_analysis()
    
    st.divider()
    
    if 'analysis_result' in st.session_state and st.session_state.analysis_result is not None:
        df = st.session_state.analysis_result
        
        # Ключевые метрики
        col1, col2, col3, col4, col5, col6 = st.columns(6)
        
        with col1:
            st.metric("📦 Уникальных SKU", len(df))
        with col2:
            st.metric("🏷️ Поставщиков", df['best_supplier'].nunique())
        with col3:
            st.metric("💰 Средняя цена", f"{df['min_price'].mean():.2f} ₽")
        with col4:
            st.metric("⬇️ Мин. цена", f"{df['min_price'].min():.2f} ₽")
        with col5:
            st.metric("📊 Медиана", f"{df['median_price'].median():.2f} ₽")
        with col6:
            st.metric("💵 Экономия", f"{df['price_diff'].sum():.2f} ₽")
        
        # Фильтры
        st.subheader("🔍 Фильтры")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            search_query = st.text_input("Поиск по SKU", placeholder="Введите артикул...")
        
        with col2:
            suppliers_list = ['Все'] + sorted(df['best_supplier'].unique().tolist())
            selected_supplier = st.selectbox("Поставщик", suppliers_list)
        
        with col3:
            categories = ['Все'] + sorted(df['category'].dropna().unique().tolist())
            selected_category = st.selectbox("Категория", categories)
        
        with col4:
            min_diff = st.number_input("Мин. разница цен (%)", value=0.0, step=1.0)
        
        # Применяем фильтры
        filtered_df = df.copy()
        
        if search_query:
            filtered_df = filtered_df[filtered_df['sku'].str.contains(search_query, case=False, na=False)]
        
        if selected_supplier != 'Все':
            filtered_df = filtered_df[filtered_df['best_supplier'] == selected_supplier]
        
        if selected_category != 'Все':
            filtered_df = filtered_df[filtered_df['category'] == selected_category]
        
        if min_diff > 0:
            filtered_df = filtered_df[filtered_df['price_diff_percent'] >= min_diff]
        
        st.subheader(f"📋 Результаты ({len(filtered_df)} товаров)")
        
        # Вкладки с разными представлениями
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Таблица",
            "📈 Графики",
            "🏆 Лучшие предложения",
            "💡 Арбитраж"
        ])
        
        with tab1:
            st.dataframe(
                filtered_df,
                use_container_width=True,
                height=600,
                column_config={
                    "sku": "Артикул",
                    "brand": "Бренд",
                    "name": "Название",
                    "min_price": st.column_config.NumberColumn("Мин. цена", format="%.2f ₽"),
                    "best_supplier": "Лучший поставщик",
                    "price_diff": st.column_config.NumberColumn("Разница", format="%.2f ₽"),
                    "price_diff_percent": st.column_config.NumberColumn("Разница %", format="%.1f%%"),
                    "suppliers_count": "Поставщиков"
                },
                hide_index=True
            )
        
        with tab2:
            # Графики
            col1, col2 = st.columns(2)
            
            with col1:
                # Топ-10 самых дешевых товаров
                top_cheap = filtered_df.nsmallest(10, 'min_price')
                fig = px.bar(
                    top_cheap,
                    x='sku',
                    y='min_price',
                    title='Топ-10 самых дешевых товаров',
                    labels={'min_price': 'Цена (₽)', 'sku': 'Артикул'},
                    color='min_price',
                    color_continuous_scale='greens'
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # Распределение цен
                fig = px.histogram(
                    filtered_df,
                    x='min_price',
                    title='Распределение цен',
                    labels={'min_price': 'Цена (₽)'},
                    nbins=30,
                    color_discrete_sequence=['blue']
                )
                st.plotly_chart(fig, use_container_width=True)
            
            # Сравнение поставщиков
            supplier_performance = filtered_df['best_supplier'].value_counts().head(10)
            fig = px.pie(
                values=supplier_performance.values,
                names=supplier_performance.index,
                title='Доля лучших предложений по поставщикам'
            )
            st.plotly_chart(fig, use_container_width=True)
        
        with tab3:
            if st.session_state.analyzer:
                best_prices = st.session_state.analyzer.get_best_prices_by_category(filtered_df)
                if not best_prices.empty:
                    st.dataframe(best_prices, use_container_width=True, hide_index=True)
                
                supplier_perf = st.session_state.analyzer.get_supplier_performance(filtered_df)
                if not supplier_perf.empty:
                    st.subheader("🏆 Эффективность поставщиков")
                    st.dataframe(supplier_perf, use_container_width=True, hide_index=True)
        
        with tab4:
            if st.session_state.analyzer:
                opportunities = st.session_state.analyzer.find_arbitrage_opportunities(
                    filtered_df, 
                    min_margin_percent=10.0
                )
                
                if not opportunities.empty:
                    st.success(f"Найдено {len(opportunities)} возможностей для арбитража")
                    st.dataframe(
                        opportunities,
                        use_container_width=True,
                        column_config={
                            "sku": "Артикул",
                            "buy_from": "Купить у",
                            "buy_price": st.column_config.NumberColumn("Цена покупки", format="%.2f ₽"),
                            "sell_price": st.column_config.NumberColumn("Цена продажи", format="%.2f ₽"),
                            "margin": st.column_config.NumberColumn("Маржа", format="%.2f ₽"),
                            "margin_percent": st.column_config.NumberColumn("Маржа %", format="%.1f%%")
                        },
                        hide_index=True
                    )
                else:
                    st.info("Возможностей для арбитража не найдено (минимальная маржа: 10%)")
        
        # Экспорт
        st.subheader("📥 Экспорт результатов")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📥 CSV", use_container_width=True):
                if st.session_state.analyzer:
                    filepath = st.session_state.analyzer.export_to_csv(filtered_df)
                    with open(filepath, 'rb') as f:
                        st.download_button(
                            "📥 Скачать CSV",
                            f.read(),
                            os.path.basename(filepath),
                            "text/csv"
                        )
        
        with col2:
            if st.button("📥 Excel", use_container_width=True):
                if st.session_state.analyzer:
                    filepath = st.session_state.analyzer.export_to_excel(filtered_df)
                    with open(filepath, 'rb') as f:
                        st.download_button(
                            "📥 Скачать Excel",
                            f.read(),
                            os.path.basename(filepath),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        
        with col3:
            if st.button("📊 Полный отчет", use_container_width=True):
                if st.session_state.analyzer:
                    filepath = st.session_state.analyzer.generate_price_report(filtered_df)
                    with open(filepath, 'rb') as f:
                        st.download_button(
                            "📥 Скачать отчет",
                            f.read(),
                            os.path.basename(filepath),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


# ===================================================================
# БЛОК 17: ГЛАВНАЯ ФУНКЦИЯ STREAMLIT
# ===================================================================

def render_dashboard() -> None:
    """Расширенная панель управления"""
    st.subheader("📊 Панель управления")
    
    # Системная информация
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        db_size = os.path.getsize(st.session_state.config.db_path) / (1024 * 1024)
        st.metric("💾 Размер БД", f"{db_size:.1f} MB")
    
    with col2:
        temp_size = sum(f.stat().st_size for f in Path(st.session_state.config.temp_dir).rglob('*') if f.is_file())
        st.metric("📁 Временные файлы", f"{temp_size / 1024 / 1024:.1f} MB")
    
    with col3:
        archive_count = len(list(Path(st.session_state.config.archive_dir).glob('*')))
        st.metric("📦 Файлов в архиве", archive_count)
    
    with col4:
        backup_count = len(list(Path(st.session_state.config.backup_dir).glob('*')))
        st.metric("💿 Резервных копий", backup_count)
    
    # Статистика запусков
    st.subheader("📈 Статистика запусков")
    
    db = st.session_state.db
    stats = db.get_stats(30)
    
    if not stats.empty:
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("🚀 Всего запусков", len(stats))
        with col2:
            st.metric("📁 Обработано файлов", int(stats['files_processed'].sum()))
        with col3:
            st.metric("📦 Отправлено товаров", int(stats['offers_sent'].sum()))
        with col4:
            st.metric("❌ Всего ошибок", int(stats['errors'].sum()))
        with col5:
            avg_time = stats['duration_seconds'].mean()
            st.metric("⏱️ Среднее время", f"{avg_time:.1f}с")
        
        # График запусков
        if len(stats) > 1:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            
            fig.add_trace(
                go.Bar(x=stats['run_date'], y=stats['files_processed'], name="Файлы"),
                secondary_y=False
            )
            
            fig.add_trace(
                go.Scatter(x=stats['run_date'], y=stats['offers_sent'], name="Товары", mode='lines+markers'),
                secondary_y=True
            )
            
            fig.update_layout(
                title='Активность робота',
                xaxis_title='Дата',
                height=400
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    # Системные предупреждения
    alerts = db.get_unresolved_alerts()
    if not alerts.empty:
        st.subheader("⚠️ Системные предупреждения")
        for _, alert in alerts.head(5).iterrows():
            severity_color = {
                'critical': '🔴',
                'error': '🟠',
                'warning': '🟡',
                'info': '🔵'
            }.get(alert['severity'], '⚪')
            
            st.warning(f"{severity_color} **{alert['title']}** - {alert['created_at'][:19]}")


def render_settings_sidebar() -> None:
    """Расширенная боковая панель настроек"""
    with st.sidebar:
        st.header("⚙️ Настройки")
        
        # Вкладки настроек
        settings_tab1, settings_tab2, settings_tab3, settings_tab4, settings_tab5 = st.tabs([
            "📧 Почта", "🏪 Маркет", "📊 Sheets", "💰 Цены", "🔔 Уведомления"
        ])
        
        with settings_tab1:
            st.text_input("IMAP сервер", key="imap_server", 
                         value=st.session_state.config.imap_server)
            st.number_input("IMAP порт", key="imap_port", 
                          value=st.session_state.config.imap_port, step=1)
            st.text_input("Email", key="email_user", 
                         value=st.session_state.config.email_user)
            st.text_input("Пароль", key="email_pass", 
                         value=st.session_state.config.email_pass, type="password")
            st.number_input("Искать за дней", key="email_search_days", 
                          value=st.session_state.config.email_search_days, step=1)
            st.number_input("Макс. писем", key="email_max_emails", 
                          value=st.session_state.config.email_max_emails, step=5)
            st.checkbox("Помечать прочитанными", key="email_mark_as_read",
                       value=st.session_state.config.email_mark_as_read)
        
        with settings_tab2:
            st.text_input("OAuth токен", key="yandex_token", 
                         value=st.session_state.config.yandex_token, type="password")
            st.number_input("ID кампании", key="campaign_id", 
                          value=st.session_state.config.campaign_id, step=1)
            st.number_input("ID склада", key="warehouse_id", 
                          value=st.session_state.config.warehouse_id, step=1)
            st.checkbox("Автообновление", key="yandex_auto_update",
                       value=st.session_state.config.yandex_auto_update)
        
        with settings_tab3:
            st.text_input("ID таблицы", key="google_sheet_id", 
                         value=st.session_state.config.google_sheet_id)
            st.text_input("Название листа", key="google_sheet_name", 
                         value=st.session_state.config.google_sheet_name)
            
            if os.path.exists(st.session_state.config.google_credentials_json):
                st.success(f"✅ Файл ключей найден")
            else:
                st.warning(f"⚠️ Файл ключей не найден")
        
        with settings_tab4:
            st.number_input("Наценка (%)", key="markup_percent", 
                          value=st.session_state.config.markup_percent, step=0.5)
            st.number_input("Мин. остаток", key="min_stock_threshold", 
                          value=st.session_state.config.min_stock_threshold, step=1)
            st.number_input("Товаров в запросе", key="max_offers_per_request", 
                          value=st.session_state.config.max_offers_per_request, step=50)
            st.number_input("Потоков", key="max_workers", 
                          value=st.session_state.config.max_workers, step=1)
        
        with settings_tab5:
            notification_config = st.session_state.config.notification_config
            st.checkbox("Включить уведомления", key="notifications_enabled",
                       value=notification_config.get('enabled', False))
            st.checkbox("Email уведомления", key="email_notifications",
                       value=notification_config.get('email_notifications', False))
            st.checkbox("Telegram уведомления", key="telegram_enabled",
                       value=notification_config.get('telegram_enabled', False))
        
        if st.button("💾 Сохранить настройки", use_container_width=True, type="primary"):
            config = st.session_state.config
            
            # Почта
            config.imap_server = st.session_state.imap_server
            config.imap_port = int(st.session_state.imap_port)
            config.email_user = st.session_state.email_user
            config.email_pass = st.session_state.email_pass
            config.email_search_days = int(st.session_state.email_search_days)
            config.email_max_emails = int(st.session_state.email_max_emails)
            config.email_mark_as_read = st.session_state.email_mark_as_read
            
            # Яндекс
            config.yandex_token = st.session_state.yandex_token
            config.campaign_id = int(st.session_state.campaign_id)
            config.warehouse_id = int(st.session_state.warehouse_id)
            config.yandex_auto_update = st.session_state.yandex_auto_update
            
            # Google Sheets
            config.google_sheet_id = st.session_state.google_sheet_id
            config.google_sheet_name = st.session_state.google_sheet_name
            
            # Обработка
            config.markup_percent = float(st.session_state.markup_percent)
            config.min_stock_threshold = int(st.session_state.min_stock_threshold)
            config.max_offers_per_request = int(st.session_state.max_offers_per_request)
            config.max_workers = int(st.session_state.max_workers)
            
            # Уведомления
            config.notification_config = {
                'enabled': st.session_state.notifications_enabled,
                'email_notifications': st.session_state.email_notifications,
                'telegram_enabled': st.session_state.telegram_enabled
            }
            
            save_config()
            st.rerun()
        
        st.divider()
        
        # Управление роботом
        if st.session_state.is_running:
            st.warning("🔄 Робот выполняется...")
            st.progress(st.session_state.progress_value / 100)
            st.info(f"📌 {st.session_state.progress_message}")
            
            if st.button("⏹️ Остановить", use_container_width=True):
                st.session_state.is_running = False
                st.rerun()
        else:
            if st.button("🚀 Запустить робота", use_container_width=True, type="primary"):
                run_robot()
                st.rerun()
        
        # Системные действия
        with st.expander("🔧 Системные действия"):
            if st.button("🧹 Очистить временные файлы", use_container_width=True):
                if st.session_state.robot:
                    cleaned = st.session_state.robot.cleanup_temp_files()
                    st.success(f"Очищено {cleaned} файлов")
                else:
                    robot = PriceRobot(st.session_state.config, st.session_state.logger)
                    cleaned = robot.cleanup_temp_files()
                    st.success(f"Очищено {cleaned} файлов")
            
            if st.button("💿 Создать бэкап БД", use_container_width=True):
                if st.session_state.robot:
                    backup_path = st.session_state.robot.backup_database()
                else:
                    robot = PriceRobot(st.session_state.config, st.session_state.logger)
                    backup_path = robot.backup_database()
                st.success(f"Бэкап создан: {backup_path}")
            
            if st.button("🔄 Обновить данные из Sheets", use_container_width=True):
                if st.session_state.product_db:
                    st.session_state.product_db.refresh_data()
                    st.success("Данные обновлены")


def render_main_content() -> None:
    """Основное содержимое страницы"""
    st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            font-weight: bold;
            background: linear-gradient(90deg, #1f77b4, #2ca02c);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 1rem;
        }
        .stButton > button {
            border-radius: 8px;
            font-weight: bold;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="main-header">🤖 Робот для загрузки и анализа прайсов v11.0</div>', 
               unsafe_allow_html=True)
    
    # Результат последнего запуска
    if st.session_state.result:
        result = st.session_state.result
        
        status_emoji = {
            'success': '✅',
            'no_files': 'ℹ️',
            'partial_success': '⚠️',
            'failed': '❌',
            'critical_failed': '💥'
        }
        
        emoji = status_emoji.get(result['status'], '❓')
        
        if result['status'] == 'success':
            st.success(f"{emoji} Робот успешно выполнил задачу!")
        elif result['status'] == 'no_files':
            st.info(f"{emoji} Новых файлов не найдено")
        elif result['status'] == 'partial_success':
            st.warning(f"{emoji} Робот выполнен с предупреждениями")
        else:
            st.error(f"{emoji} Робот выполнен с ошибками")
        
        # Детальные метрики
        cols = st.columns(8)
        metrics = [
            ("📁 Файлов", result.get('files_processed', 0)),
            ("⏭️ Пропущено", result.get('files_skipped', 0)),
            ("❌ С ошибками", result.get('files_with_errors', 0)),
            ("🔄 Обновлено", result.get('products_updated', 0)),
            ("➕ Добавлено", result.get('products_added', 0)),
            ("📦 Отправлено", result.get('offers_sent', 0)),
            ("⏱️ Время", f"{result.get('duration', 0):.1f}с"),
            ("❌ Ошибок", len(result.get('errors', [])))
        ]
        
        for col, (label, value) in zip(cols, metrics):
            with col:
                st.metric(label, value)
        
        # Ошибки
        if result.get('errors'):
            with st.expander(f"❌ Ошибки ({len(result['errors'])})", expanded=len(result['errors']) > 0):
                for error in result['errors']:
                    st.error(f"• {error}")
        
        # Предупреждения
        if result.get('warnings'):
            with st.expander(f"⚠️ Предупреждения ({len(result['warnings'])})"):
                for warning in result['warnings']:
                    st.warning(f"• {warning}")
    
    # Основные вкладки
    main_tab1, main_tab2, main_tab3, main_tab4, main_tab5 = st.tabs([
        "📊 Панель управления",
        "📊 Аналитика прайсов",
        "🛠️ Маппинг",
        "📦 База товаров",
        "📝 Логи"
    ])
    
    with main_tab1:
        render_dashboard()
    
    with main_tab2:
        analysis_subtab1, analysis_subtab2 = st.tabs(["📊 Анализ цен", "⚙️ Поставщики"])
        
        with analysis_subtab1:
            render_analysis_tab()
        
        with analysis_subtab2:
            render_supplier_settings()
    
    with main_tab3:
        render_mapping_constructor()
    
    with main_tab4:
        st.subheader("📦 База товаров (Google Sheets)")
        
        if st.session_state.product_db is None:
            st.error("❌ Не удалось подключиться к Google Sheets")
        else:
            df = st.session_state.product_db.get_all_products()
            
            if df.empty:
                st.warning("База товаров пуста")
            else:
                # Статистика
                stats = st.session_state.product_db.get_stats()
                
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("📦 Всего", stats['total_products'])
                with col2:
                    st.metric("🏷️ Брендов", stats['total_brands'])
                with col3:
                    st.metric("💰 Средняя цена", f"{stats['avg_price']:.0f} ₽")
                with col4:
                    st.metric("📦 В наличии", stats['products_with_stock'])
                with col5:
                    st.metric("💵 Общая стоимость", f"{stats['total_value']:,.0f} ₽")
                
                # Ссылка на Google Sheets
                if st.session_state.config.google_sheet_id:
                    st.markdown(
                        f"🔗 [Открыть Google Sheets](https://docs.google.com/spreadsheets/d/"
                        f"{st.session_state.config.google_sheet_id})"
                    )
                
                # Поиск
                search_query = st.text_input("🔍 Поиск по базе", 
                                           placeholder="Введите артикул, бренд или название...")
                
                if search_query:
                    df = st.session_state.product_db.search_products(search_query)
                    st.write(f"Найдено: {len(df)} товаров")
                
                # Таблица товаров
                st.dataframe(
                    df,
                    use_container_width=True,
                    column_config={
                        "артикул": "Артикул",
                        "бренд": "Бренд",
                        "название": "Название",
                        "цена_розница": st.column_config.NumberColumn("Цена", format="%.2f ₽"),
                        "остаток": "Остаток",
                        "категория": "Категория"
                    },
                    height=500,
                    hide_index=True
                )
                
                # Экспорт
                if st.button("📥 Экспортировать базу в Excel"):
                    filepath = st.session_state.product_db.export_to_excel()
                    with open(filepath, 'rb') as f:
                        st.download_button(
                            "📥 Скачать Excel",
                            f.read(),
                            os.path.basename(filepath),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
    
    with main_tab5:
        st.subheader("📝 Логи выполнения")
        
        # Фильтры логов
        col1, col2, col3 = st.columns(3)
        
        with col1:
            log_level_filter = st.selectbox(
                "Уровень",
                ['ALL', 'INFO', 'WARNING', 'ERROR', 'CRITICAL', 'SUCCESS', 'DEBUG'],
                key="log_level_filter"
            )
        
        with col2:
            log_count = st.number_input("Количество", value=100, step=50, key="log_count")
        
        with col3:
            log_search = st.text_input("Поиск в логах", key="log_search")
        
        # Отображение логов
        log_container = st.container(height=500)
        
        with log_container:
            logs = st.session_state.logger.get_logs(log_count)
            
            if log_level_filter != 'ALL':
                logs = [log for log in logs if log['level'] == log_level_filter]
            
            if log_search:
                logs = [log for log in logs if log_search.lower() in log['message'].lower()]
            
            for log in logs:
                level = log['level']
                message = log['message']
                timestamp = log['timestamp'][:19]
                
                if level in ['ERROR', 'CRITICAL']:
                    st.error(f"[{timestamp}] {message}")
                elif level == 'WARNING':
                    st.warning(f"[{timestamp}] {message}")
                elif level == 'SUCCESS':
                    st.success(f"[{timestamp}] {message}")
                elif level == 'DEBUG':
                    st.caption(f"[{timestamp}] {message}")
                else:
                    st.info(f"[{timestamp}] {message}")
        
        # Экспорт логов
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📥 Экспорт логов (JSON)", use_container_width=True):
                filepath = st.session_state.logger.export_logs('json')
                with open(filepath, 'rb') as f:
                    st.download_button(
                        "📥 Скачать JSON",
                        f.read(),
                        os.path.basename(filepath),
                        "application/json"
                    )
        
        with col2:
            if st.button("🔄 Обновить логи", use_container_width=True):
                st.rerun()


# ===================================================================
# БЛОК 18: ГЛАВНАЯ ФУНКЦИЯ ПРИЛОЖЕНИЯ
# ===================================================================

def main() -> None:
    """Главная функция приложения"""
    try:
        # Инициализация
        init_session_state()
        
        # Настройка страницы
        st.set_page_config(
            page_title="🤖 Робот для прайсов v11.0",
            page_icon="🤖",
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
        # Боковая панель
        render_settings_sidebar()
        
        # Основное содержимое
        render_main_content()
        
        # Автообновление если включено
        if st.session_state.get('auto_refresh', False):
            time.sleep(5)
            st.rerun()
        
    except Exception as e:
        st.error(f"❌ Критическая ошибка в приложении: {e}")
        st.exception(e)
        
        # Попытка восстановления
        if st.button("🔄 Перезапустить приложение"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()


# ===================================================================
# БЛОК 19: ТОЧКА ВХОДА И ЗАПУСК ИЗ КОМАНДНОЙ СТРОКИ
# ===================================================================

def run_cli():
    """Запуск из командной строки"""
    parser = argparse.ArgumentParser(description='Робот для загрузки и анализа прайсов')
    
    parser.add_argument('--mode', choices=['run', 'analyze', 'cleanup', 'backup', 'health'],
                       default='run', help='Режим работы')
    
    parser.add_argument('--supplier', type=str, help='Имя поставщика для обработки')
    parser.add_argument('--config', type=str, default='config.json', help='Путь к файлу конфигурации')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                       default='INFO', help='Уровень логирования')
    
    args = parser.parse_args()
    
    # Загружаем конфигурацию
    config = Config.load()
    logger = StreamlitLogger()
    
    # Создаем робота
    robot = PriceRobot(config, logger)
    
    if args.mode == 'run':
        if args.supplier:
            result = robot.run_single_supplier(args.supplier)
        else:
            result = robot.run_full_cycle()
        
        print(f"Статус: {result['status']}")
        print(f"Файлов обработано: {result.get('files_processed', 0)}")
        print(f"Товаров обновлено: {result.get('products_updated', 0)}")
        print(f"Ошибок: {len(result.get('errors', []))}")
        
    elif args.mode == 'analyze':
        df = robot.quick_analysis()
        if not df.empty:
            filepath = robot.price_analyzer.export_to_excel(df)
            print(f"Анализ сохранен в: {filepath}")
        else:
            print("Нет данных для анализа")
    
    elif args.mode == 'cleanup':
        cleaned = robot.cleanup_temp_files()
        print(f"Очищено файлов: {cleaned}")
    
    elif args.mode == 'backup':
        backup_path = robot.backup_database()
        print(f"Бэкап создан: {backup_path}")
    
    elif args.mode == 'health':
        health = robot.get_system_health()
        print(json.dumps(health, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    # Определяем режим запуска
    if len(sys.argv) > 1:
        # Запуск из командной строки
        run_cli()
    else:
        # Запуск Streamlit приложения
        main()
